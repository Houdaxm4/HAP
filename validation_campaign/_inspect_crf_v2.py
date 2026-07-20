"""Refined Custom_Run_Filter structure inventory for CustomRunData design."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"C:\Users\G\HAP\validation_campaign\universe")
TICKERS = ["AAPL", "MSFT", "AMZN", "TJX"]
OUT = Path(r"C:\Users\G\HAP\validation_campaign\_crf_structure.json")
MD = Path(r"C:\Users\G\HAP\validation_campaign\_crf_structure.md")


def cell_str(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v)


def is_numeric_label(s: str | None) -> bool:
    if s is None:
        return False
    return bool(re.fullmatch(r"-?\d+(\.\d+)?", s.strip()))


def find_crf(ticker: str) -> Path:
    for p in (ROOT / ticker).glob("Custom_Run_Filter*.xlsx"):
        if not p.name.startswith("~$"):
            return p
    raise FileNotFoundError(ticker)


def load_sheet_rows(path: Path, sheet: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [cell_str(v) for v in row]
        # trim trailing empties
        last = 0
        for j, v in enumerate(cells, start=1):
            if v is not None:
                last = j
        rows.append({"row": i, "cells": cells[:last], "width": last})
    wb.close()
    return rows


def analyze_ticker_sheet(ticker: str, rows: list[dict]) -> dict:
    # Section detection
    # 1) Title row 1 often ['', 'TICKER US Equity'] or similar
    # 2) Scalar KV until blank(s) then date header
    # 3) Quarterly time series block starting at 'date' or 'Fiscal Quarter'
    # 4) Eventually long stretch of numeric col-A (daily matrix / index dump)

    meta_kv = []
    date_row = None
    fq_row = None
    fy_row = None
    first_numeric_run_start = None
    ts_metrics = []  # metric name rows in quarterly block
    post_ts_scalars = []  # after fy? or after series with mostly A+B

    # Find date / fiscal quarter / fiscal year rows
    for r in rows:
        a = r["cells"][0] if r["cells"] else None
        if a == "date" and date_row is None:
            date_row = r
        elif a == "Fiscal Quarter" and fq_row is None:
            fq_row = r
        elif a == "Fiscal Year" and fy_row is None:
            fy_row = r

    # Find first long numeric label run (index column dump)
    numeric_run = 0
    for r in rows:
        a = r["cells"][0] if r["cells"] else None
        if is_numeric_label(a) and r["width"] >= 2:
            numeric_run += 1
            if numeric_run >= 20 and first_numeric_run_start is None:
                # backtrack to start
                start = r["row"] - numeric_run + 1
                first_numeric_run_start = start
                break
        else:
            numeric_run = 0

    # If not found with width>=2, try any numeric A
    if first_numeric_run_start is None:
        numeric_run = 0
        for r in rows:
            a = r["cells"][0] if r["cells"] else None
            if is_numeric_label(a):
                numeric_run += 1
                if numeric_run >= 20 and first_numeric_run_start is None:
                    first_numeric_run_start = r["row"] - numeric_run + 1
                    break
            else:
                numeric_run = 0

    # Partition
    series_start = date_row["row"] if date_row else None
    trailer_start = first_numeric_run_start

    # Meta: from row 2 until series_start-1 (skip empties)
    for r in rows:
        if series_start and r["row"] >= series_start:
            break
        if r["row"] == 1:
            continue
        if not r["cells"]:
            continue
        a = r["cells"][0]
        b = r["cells"][1] if len(r["cells"]) > 1 else None
        if a and not is_numeric_label(a) and r["width"] <= 3:
            meta_kv.append({"row": r["row"], "label": a, "value": b})

    # Collect period headers from date / fiscal quarter / fiscal year
    def periods(row_obj):
        if not row_obj:
            return []
        return [c for c in row_obj["cells"][1:] if c is not None]

    date_periods = periods(date_row)
    fq_periods = periods(fq_row)
    fy_periods = periods(fy_row)

    # Time series metrics: after date/fq headers until trailer or until we enter a sparse scalar region
    # Heuristic: in series block, width is large (>=10) OR known metric names with many values
    in_series = False
    for r in rows:
        a = r["cells"][0] if r["cells"] else None
        if not a:
            continue
        if series_start and r["row"] < series_start:
            continue
        if trailer_start and r["row"] >= trailer_start:
            break
        if a in ("date", "Fiscal Quarter", "Fiscal Year"):
            in_series = True
            continue
        if is_numeric_label(a):
            continue
        # metric row
        values = [c for c in r["cells"][1:] if c is not None]
        n_vals = len(values)
        if n_vals >= 5:
            ts_metrics.append(
                {
                    "row": r["row"],
                    "label": a,
                    "n_values": n_vals,
                    "first3": values[:3],
                    "last3": values[-3:],
                }
            )
            in_series = True
        elif n_vals <= 2 and r["row"] > (fy_row["row"] if fy_row else (series_start or 0)):
            # scalar-like after series
            post_ts_scalars.append({"row": r["row"], "label": a, "value": r["cells"][1] if len(r["cells"]) > 1 else None, "n_values": n_vals})
        elif n_vals < 5:
            # short series or scalar with a few extras
            if n_vals == 0:
                post_ts_scalars.append({"row": r["row"], "label": a, "value": None, "n_values": 0})
            elif n_vals <= 2:
                post_ts_scalars.append({"row": r["row"], "label": a, "value": r["cells"][1] if len(r["cells"]) > 1 else None, "n_values": n_vals})
            else:
                ts_metrics.append(
                    {
                        "row": r["row"],
                        "label": a,
                        "n_values": n_vals,
                        "first3": values[:3],
                        "last3": values[-3:],
                    }
                )

    # Trailer summary
    trailer_rows = 0
    trailer_max_width = 0
    trailer_sample = []
    if trailer_start:
        for r in rows:
            if r["row"] < trailer_start:
                continue
            trailer_rows += 1
            trailer_max_width = max(trailer_max_width, r["width"])
            if len(trailer_sample) < 5 or r["row"] >= rows[-1]["row"] - 2:
                trailer_sample.append(
                    {
                        "row": r["row"],
                        "a": r["cells"][0] if r["cells"] else None,
                        "width": r["width"],
                        "first5": r["cells"][:5],
                    }
                )

    # Unique real labels (non-numeric)
    real_labels = []
    seen = set()
    for r in rows:
        a = r["cells"][0] if r["cells"] else None
        if a and not is_numeric_label(a) and a not in seen:
            seen.add(a)
            real_labels.append(a)

    # Preview first 15 rows (up to 30 cols)
    preview = []
    for r in rows[:15]:
        preview.append({"row": r["row"], "cells": r["cells"][:30]})

    # Peek around fy and trailer
    def window(center, before=3, after=8):
        if not center:
            return []
        out = []
        for r in rows:
            if center - before <= r["row"] <= center + after:
                out.append({"row": r["row"], "width": r["width"], "a": r["cells"][0] if r["cells"] else None, "b": r["cells"][1] if len(r["cells"]) > 1 else None, "sample": r["cells"][:8]})
        return out

    return {
        "ticker": ticker,
        "total_rows": len(rows),
        "max_width": max((r["width"] for r in rows), default=0),
        "preview_first_15": preview,
        "meta_kv": meta_kv,
        "date_row": date_row["row"] if date_row else None,
        "fiscal_quarter_row": fq_row["row"] if fq_row else None,
        "fiscal_year_row": fy_row["row"] if fy_row else None,
        "n_date_periods": len(date_periods),
        "date_periods_first10": date_periods[:10],
        "date_periods_last5": date_periods[-5:],
        "n_fq_periods": len(fq_periods),
        "fq_periods_first10": fq_periods[:10],
        "fq_periods_last5": fq_periods[-5:],
        "n_fy_periods": len(fy_periods),
        "fy_periods_first10": fy_periods[:10],
        "fy_periods_last5": fy_periods[-5:],
        "trailer_numeric_start_row": trailer_start,
        "trailer_rows": trailer_rows,
        "trailer_max_width": trailer_max_width,
        "trailer_sample": trailer_sample,
        "time_series_metric_count": len(ts_metrics),
        "time_series_metric_labels": [m["label"] for m in ts_metrics],
        "time_series_metrics": ts_metrics,
        "post_series_scalar_count": len(post_ts_scalars),
        "post_series_scalar_labels": [m["label"] for m in post_ts_scalars],
        "post_series_scalars": post_ts_scalars,
        "all_real_col_a_labels": real_labels,
        "real_label_count": len(real_labels),
        "window_around_fy": window(fy_row["row"] if fy_row else None),
        "window_around_trailer": window(trailer_start, before=5, after=3) if trailer_start else [],
        "title_row": rows[0] if rows else None,
    }


def analyze_summary(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [cell_str(v) for v in rows[0]] if rows else []
    # drop trailing Nones
    while headers and headers[-1] is None:
        headers.pop()
    values = [cell_str(v) for v in rows[1][: len(headers)]] if len(rows) > 1 else []
    field_map = []
    for h, v in zip(headers, values):
        field_map.append({"label": h, "value": v, "is_year_header": bool(re.fullmatch(r"(FY\s*)?(19|20)\d{2}", h or "", re.I))})
    return {
        "n_rows": len(rows),
        "n_headers": len(headers),
        "headers": headers,
        "row2_values": values,
        "fields": field_map,
    }


def analyze_template_aapl() -> dict:
    path = ROOT / "AAPL" / "AAPL 2026 Q2 - Industrial Template v27.6.xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return {
        "filename": path.name,
        "sheet_count": len(names),
        "first_20": names[:20],
        "all": names,
        "has_income": any("income" in n.lower() for n in names),
        "has_balance": any("balance" in n.lower() for n in names),
        "has_cashflow": any("cash flow" in n.lower() or n.lower().startswith("cf") for n in names),
    }


def main():
    inventories = {}
    summaries = {}
    for t in TICKERS:
        path = find_crf(t)
        print(f"Loading {t} ticker sheet...", flush=True)
        rows = load_sheet_rows(path, t)
        inventories[t] = {
            "path": str(path),
            "filename": path.name,
            "sheet_names": [t, "Summary"],
            "ticker_sheet": analyze_ticker_sheet(t, rows),
        }
        print(f"Loading {t} Summary...", flush=True)
        summaries[t] = analyze_summary(path)
        inventories[t]["summary"] = summaries[t]

    # Compare real labels (non-numeric) across tickers
    aapl_real = set(inventories["AAPL"]["ticker_sheet"]["all_real_col_a_labels"])
    aapl_ts = set(inventories["AAPL"]["ticker_sheet"]["time_series_metric_labels"])
    aapl_sum = summaries["AAPL"]["headers"]

    comparison = {
        "summary_headers_identical": all(summaries[t]["headers"] == aapl_sum for t in TICKERS),
        "summary_header_count": {t: summaries[t]["n_headers"] for t in TICKERS},
        "ticker_sheet_dims": {
            t: {
                "rows": inventories[t]["ticker_sheet"]["total_rows"],
                "max_width": inventories[t]["ticker_sheet"]["max_width"],
                "date_row": inventories[t]["ticker_sheet"]["date_row"],
                "fq_row": inventories[t]["ticker_sheet"]["fiscal_quarter_row"],
                "fy_row": inventories[t]["ticker_sheet"]["fiscal_year_row"],
                "n_date_periods": inventories[t]["ticker_sheet"]["n_date_periods"],
                "n_fq_periods": inventories[t]["ticker_sheet"]["n_fq_periods"],
                "n_fy_periods": inventories[t]["ticker_sheet"]["n_fy_periods"],
                "ts_metric_count": inventories[t]["ticker_sheet"]["time_series_metric_count"],
                "post_scalar_count": inventories[t]["ticker_sheet"]["post_series_scalar_count"],
                "real_label_count": inventories[t]["ticker_sheet"]["real_label_count"],
                "trailer_start": inventories[t]["ticker_sheet"]["trailer_numeric_start_row"],
                "trailer_rows": inventories[t]["ticker_sheet"]["trailer_rows"],
                "meta_kv_count": len(inventories[t]["ticker_sheet"]["meta_kv"]),
            }
            for t in TICKERS
        },
        "meta_kv_labels_vs_aapl": {
            t: {
                "missing": sorted(
                    set(m["label"] for m in inventories["AAPL"]["ticker_sheet"]["meta_kv"])
                    - set(m["label"] for m in inventories[t]["ticker_sheet"]["meta_kv"])
                ),
                "extra": sorted(
                    set(m["label"] for m in inventories[t]["ticker_sheet"]["meta_kv"])
                    - set(m["label"] for m in inventories["AAPL"]["ticker_sheet"]["meta_kv"])
                ),
            }
            for t in TICKERS
        },
        "ts_labels_vs_aapl": {
            t: {
                "missing": sorted(aapl_ts - set(inventories[t]["ticker_sheet"]["time_series_metric_labels"])),
                "extra": sorted(set(inventories[t]["ticker_sheet"]["time_series_metric_labels"]) - aapl_ts),
            }
            for t in TICKERS
        },
        "real_labels_vs_aapl": {
            t: {
                "missing": sorted(aapl_real - set(inventories[t]["ticker_sheet"]["all_real_col_a_labels"])),
                "extra": sorted(set(inventories[t]["ticker_sheet"]["all_real_col_a_labels"]) - aapl_real),
            }
            for t in TICKERS
        },
    }

    template = analyze_template_aapl()
    folder_listings = {}
    for t in TICKERS:
        folder_listings[t] = [
            {"name": p.name, "bytes": p.stat().st_size}
            for p in sorted((ROOT / t).iterdir())
            if not p.name.startswith("~$")
        ]

    payload = {
        "folder_listings": folder_listings,
        "inventories": inventories,
        "comparison": comparison,
        "aapl_industrial_template": template,
        "parser_notes": {
            "layout": [
                "Two sheets only: <TICKER> (proprietary long form) + Summary (1 header + 1 data row).",
                "Ticker sheet layout:",
                "  row1: title (e.g. 'AAPL US Equity' in col B)",
                "  rows ~2-12: scalar key-value in cols A/B (company meta, live price, sector, EV, next earnings)",
                "  blank rows",
                "  'date' row: quarterly period end dates across columns B..",
                "  'Fiscal Quarter' row: e.g. '2001 Q1'",
                "  many metric rows: col A = metric label, B.. = quarterly values (~100 cols / ~25 years)",
                "  'Fiscal Year' row appears mid/late aligning annual bins",
                "  additional current/derived scalar metrics (narrow A/B) interspersed or after",
                "  large trailing block (~3900 rows) with numeric indices in col A — appears to be a transposed daily/auxiliary matrix; IGNORE for metric catalog",
                "Summary sheet: identical 111-column schema across companies; one row of current scalars (overlaps ticker-sheet meta + derived scores).",
            ],
            "scalar_vs_timeseries": {
                "scalars": "Summary all columns; ticker-sheet top meta KV; many 'Current *' / ranks / PE10 entry rules with single value",
                "timeseries_quarterly": "Ticker sheet rows under date/Fiscal Quarter headers with n_values >= ~5",
                "timeseries_annual": "Values aligned to Fiscal Year header row periods",
            },
            "recommended_CustomRunData_shape": {
                "summary": "dict[str, Any] keyed by exact Summary header",
                "meta": "dict from ticker top KV",
                "periods_quarterly": "list[date|str] from date row",
                "periods_fiscal_quarter": "list[str] from Fiscal Quarter row",
                "periods_fiscal_year": "list[str|int] from Fiscal Year row",
                "series": "dict[str, list[Optional[float]]] keyed by metric label",
                "scalars_ticker": "dict for narrow post/mid series A/B fields",
                "ignore": "numeric-index trailer rows",
            },
        },
    }

    OUT.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {OUT}", flush=True)

    # Markdown digest focused on parser design
    lines = []
    lines.append("# Custom_Run_Filter Structure Inventory (parser design)")
    lines.append("")
    lines.append("## Folder listings")
    for t in TICKERS:
        lines.append(f"### {t}")
        for f in folder_listings[t]:
            lines.append(f"- `{f['name']}` ({f['bytes']:,} bytes)")
        lines.append("")

    lines.append("## Sheet roles")
    lines.append("| Company | Proprietary (~4k rows) | Summary | Summary cols | Ticker rows | Ticker cols | Trailer start |")
    lines.append("|---|---|---|---|---|---|---|")
    for t in TICKERS:
        ts = inventories[t]["ticker_sheet"]
        lines.append(
            f"| {t} | `{t}` | `Summary` | {summaries[t]['n_headers']} | {ts['total_rows']} | {ts['max_width']} | {ts['trailer_numeric_start_row']} |"
        )
    lines.append("")

    lines.append("## Cross-company consistency")
    lines.append(f"- Summary headers identical across all 4: **{comparison['summary_headers_identical']}**")
    lines.append(f"- Sheet pattern: `[TICKER, Summary]` for all (TICKER name differs).")
    lines.append("")
    lines.append("### Dimension comparison")
    lines.append("```json")
    lines.append(json.dumps(comparison["ticker_sheet_dims"], indent=2))
    lines.append("```")
    lines.append("")
    lines.append("### Meta KV diffs vs AAPL")
    lines.append("```json")
    lines.append(json.dumps(comparison["meta_kv_labels_vs_aapl"], indent=2))
    lines.append("```")
    lines.append("")
    lines.append("### Time-series label diffs vs AAPL")
    lines.append("```json")
    lines.append(json.dumps(comparison["ts_labels_vs_aapl"], indent=2))
    lines.append("```")
    lines.append("")
    lines.append("### Real (non-numeric) label diffs vs AAPL")
    lines.append("```json")
    lines.append(json.dumps(comparison["real_labels_vs_aapl"], indent=2))
    lines.append("```")
    lines.append("")

    # AAPL deep dive
    a = inventories["AAPL"]["ticker_sheet"]
    lines.append("## AAPL deep dive (ticker sheet)")
    lines.append(f"- Title: {a['title_row']}")
    lines.append(f"- date row={a['date_row']} ({a['n_date_periods']} periods)")
    lines.append(f"  - first10: {a['date_periods_first10']}")
    lines.append(f"  - last5: {a['date_periods_last5']}")
    lines.append(f"- Fiscal Quarter row={a['fiscal_quarter_row']} ({a['n_fq_periods']} periods)")
    lines.append(f"  - first10: {a['fq_periods_first10']}")
    lines.append(f"  - last5: {a['fq_periods_last5']}")
    lines.append(f"- Fiscal Year row={a['fiscal_year_row']} ({a['n_fy_periods']} periods)")
    lines.append(f"  - first10: {a['fy_periods_first10']}")
    lines.append(f"  - last5: {a['fy_periods_last5']}")
    lines.append(f"- Trailer numeric block starts row {a['trailer_numeric_start_row']} ({a['trailer_rows']} rows, max_width={a['trailer_max_width']})")
    lines.append("")
    lines.append("### Meta scalars (A/B)")
    for m in a["meta_kv"]:
        lines.append(f"- r{m['row']}: `{m['label']}` = {m['value']}")
    lines.append("")
    lines.append(f"### Time-series metrics ({a['time_series_metric_count']})")
    for m in a["time_series_metrics"]:
        lines.append(f"- r{m['row']}: `{m['label']}` (n={m['n_values']}) first={m['first3']} last={m['last3']}")
    lines.append("")
    lines.append(f"### Post/narrow scalars ({a['post_series_scalar_count']})")
    for m in a["post_series_scalars"]:
        lines.append(f"- r{m['row']}: `{m['label']}` = {m['value']}")
    lines.append("")
    lines.append("### All real col-A labels (excluding numeric trailer indices)")
    for lab in a["all_real_col_a_labels"]:
        lines.append(f"- {lab}")
    lines.append("")
    lines.append("### First 15 rows preview (≤30 cols)")
    for p in a["preview_first_15"]:
        lines.append(f"- r{p['row']}: {p['cells']}")
    lines.append("")

    lines.append("## Summary sheet — full field catalog (identical across AAPL/MSFT/AMZN/TJX)")
    lines.append(f"- Rows: 2 (header + 1 company). Columns: {summaries['AAPL']['n_headers']}")
    lines.append("")
    lines.append("| # | Field | AAPL value | Kind |")
    lines.append("|---|---|---|---|")
    for i, f in enumerate(summaries["AAPL"]["fields"], start=1):
        kind = "scalar"
        lines.append(f"| {i} | {f['label']} | {f['value']} | {kind} |")
    lines.append("")

    lines.append("## AAPL Industrial Template (brief)")
    lines.append(f"- File: `{template['filename']}`")
    lines.append(f"- Sheets ({template['sheet_count']}): first 20 = {template['first_20']}")
    lines.append(f"- All sheets: {template['all']}")
    lines.append(f"- Income-style: {template['has_income']}; Balance-style: {template['has_balance']}; Cash-flow-style: {template['has_cashflow']}")
    lines.append("")

    lines.append("## Parser recommendations")
    for note in payload["parser_notes"]["layout"]:
        lines.append(f"- {note}")
    lines.append("")
    lines.append("### Suggested `CustomRunData`")
    lines.append("```json")
    lines.append(json.dumps(payload["parser_notes"]["recommended_CustomRunData_shape"], indent=2))
    lines.append("```")

    MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {MD}", flush=True)


if __name__ == "__main__":
    main()
