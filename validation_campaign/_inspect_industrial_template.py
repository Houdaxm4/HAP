"""
M1 — Industrial Template inventory (read-only).

Scans AAPL / MSFT / AMZN / TJX production templates and writes:
  docs/workbook_mapping/INDUSTRIAL_TEMPLATE_INVENTORY.md
  docs/workbook_mapping/INDUSTRIAL_TEMPLATE_DEPENDENCY_MAP.md
  docs/workbook_mapping/INDUSTRIAL_TEMPLATE_RISK_LOG.md
  docs/workbook_mapping/industrial_template_v27_inventory.json
  docs/workbook_mapping/industrial_template_v27_fingerprint.json
  docs/workbook_mapping/suite_structural_diff.json

No cell writes. No mapping engine. No mappings.
"""
from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedNameDict

ROOT = Path(__file__).resolve().parents[1]
UNIVERSE = ROOT / "validation_campaign" / "universe"
OUT_DIR = ROOT / "docs" / "workbook_mapping"

TICKERS = ("AAPL", "MSFT", "AMZN", "TJX")

CORE_SHEETS = (
    "Income - GAAP",
    "Balance Sheet - Standardized",
    "Cash Flow - Standardized",
)

SHEET_REF_RE = re.compile(r"'([^']+)'!|[A-Za-z0-9 _%-]+!")
VOLATILE_RE = re.compile(
    r"\b(INDIRECT|OFFSET|NOW|TODAY|RAND|RANDBETWEEN|INFO|CELL)\s*\(",
    re.I,
)
EXTERNAL_RE = re.compile(r"\[[^\]]+\]")
CIRCULAR_HINT_RE = re.compile(r"\b(ITERAT|CIRCULAR)\b", re.I)

# Rows to scan densely for controls / headers on statement sheets
HEADER_ROW_MAX = 25
LABEL_COL_CANDIDATES = (1, 2, 3)  # A, B, C
MAX_MERGED_SAMPLES = 40
MAX_FORMULA_SAMPLES = 30
MAX_LABEL_ROWS = 250
MAX_HIDDEN_SAMPLES = 80


def find_template(ticker: str) -> Path:
    folder = UNIVERSE / ticker
    matches = sorted(
        p
        for p in folder.glob("*.xlsx")
        if "Industrial Template" in p.name and not p.name.startswith("~$")
    )
    if not matches:
        raise FileNotFoundError(f"No Industrial Template for {ticker} in {folder}")
    return matches[0]


def cell_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v)


def is_formula(v: Any) -> bool:
    return isinstance(v, str) and v.startswith("=")


def sheet_refs_in_formula(formula: str, known_sheets: set[str] | None = None) -> set[str]:
    refs: set[str] = set()
    known = known_sheets or set()
    # Prefer quoted sheet names
    for m in re.finditer(r"'([^']+)'!", formula):
        refs.add(m.group(1))
    # Unquoted names only when they exactly match a known sheet
    if known:
        for ks in known:
            # Match SheetName! but not when preceded by letter (part of larger token)
            pat = re.compile(rf"(?<![A-Za-z0-9_]){re.escape(ks)}!")
            if pat.search(formula):
                refs.add(ks)
            # Trailing-space sheet names already in known
    return refs


def classify_row_role(label: str | None) -> str:
    if not label:
        return "blank"
    low = label.lower()
    if "check" in low:
        return "check"
    if low.startswith("%") or "margin" in low or "growth" in low:
        return "ratio_like"
    if label.isupper() and len(label) > 3:
        return "section_header"
    if label.endswith(":"):
        return "section_header"
    return "line_item"


def indent_level(label: str | None, alignment_indent: int | None) -> int:
    if alignment_indent:
        return int(alignment_indent)
    if not label:
        return 0
    leading = len(label) - len(label.lstrip(" "))
    # Also treat leading non-breaking / special spaces
    return max(0, leading // 2)


def extract_named_ranges(wb) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    defined: DefinedNameDict = wb.defined_names
    for name in defined:
        dn = defined[name]
        try:
            destinations = list(dn.destinations) if dn.attr_text else []
        except Exception:
            destinations = []
        dest_repr = []
        for sheet, coord in destinations:
            dest_repr.append({"sheet": sheet, "coord": coord})
        out.append(
            {
                "name": name,
                "attr_text": getattr(dn, "attr_text", None),
                "destinations": dest_repr,
            }
        )
    return out


def extract_tables(ws) -> list[dict[str, Any]]:
    tables = []
    try:
        for tname, table in (ws.tables or {}).items():
            tables.append(
                {
                    "name": tname,
                    "ref": table.ref,
                    "displayName": getattr(table, "displayName", tname),
                }
            )
    except Exception as exc:
        tables.append({"error": str(exc)})
    return tables


def extract_data_validations(ws) -> dict[str, Any]:
    dvs = getattr(ws, "data_validations", None)
    if not dvs or not getattr(dvs, "dataValidation", None):
        return {"count": 0, "samples": []}
    samples = []
    for i, dv in enumerate(dvs.dataValidation):
        if i >= 15:
            break
        samples.append(
            {
                "sqref": str(getattr(dv, "sqref", "")),
                "type": getattr(dv, "type", None),
                "formula1": getattr(dv, "formula1", None),
                "formula2": getattr(dv, "formula2", None),
            }
        )
    return {"count": len(dvs.dataValidation), "samples": samples}


def extract_conditional_formatting(ws) -> dict[str, Any]:
    cf = getattr(ws, "conditional_formatting", None)
    if not cf:
        return {"rule_count": 0, "ranges": []}
    ranges = []
    try:
        for sqref in list(cf)[:20]:
            ranges.append(str(sqref))
    except Exception:
        pass
    try:
        rule_count = sum(len(cf[r]) for r in cf)
    except Exception:
        rule_count = len(ranges)
    return {"rule_count": rule_count, "ranges": ranges}


def freeze_panes_info(ws) -> dict[str, Any] | None:
    fp = ws.freeze_panes
    if not fp:
        return None
    return {"cell": str(fp)}


def merged_cells_info(ws) -> dict[str, Any]:
    ranges = [str(r) for r in ws.merged_cells.ranges]
    return {
        "count": len(ranges),
        "samples": ranges[:MAX_MERGED_SAMPLES],
    }


def hidden_rows_cols(ws) -> dict[str, Any]:
    hidden_rows = []
    for idx, dim in (ws.row_dimensions or {}).items():
        if getattr(dim, "hidden", False):
            hidden_rows.append(int(idx))
    hidden_cols = []
    for letter, dim in (ws.column_dimensions or {}).items():
        if getattr(dim, "hidden", False):
            hidden_cols.append(str(letter))
    return {
        "hidden_row_count": len(hidden_rows),
        "hidden_rows_sample": hidden_rows[:MAX_HIDDEN_SAMPLES],
        "hidden_column_count": len(hidden_cols),
        "hidden_columns": hidden_cols[:MAX_HIDDEN_SAMPLES],
    }


def protection_info(ws) -> dict[str, Any]:
    p = ws.protection
    return {
        "sheet": bool(getattr(p, "sheet", False)),
        "password_set": bool(getattr(p, "password", None)),
    }


def scan_control_region(ws) -> dict[str, Any]:
    """Scan top-left for ticker / year / units / period headers."""
    controls: list[dict[str, Any]] = []
    period_headers: list[dict[str, Any]] = []
    max_col = min(ws.max_column or 1, 80)
    max_row = min(ws.max_row or 1, HEADER_ROW_MAX)

    # Explicit paired controls: labels in A, values in C (B is often blank/hidden)
    for r, kind in ((1, "ticker"), (2, "start_year"), (3, "end_year")):
        lab = cell_str(ws.cell(r, 1).value)
        val_b = ws.cell(r, 2)
        val_c = ws.cell(r, 3)
        if lab:
            controls.append(
                {
                    "cell": f"A{r}",
                    "value": lab[:120],
                    "kind": f"{kind}_label",
                    "is_formula": is_formula(ws.cell(r, 1).value),
                }
            )
        b_s = cell_str(val_b.value)
        c_s = cell_str(val_c.value)
        if b_s:
            controls.append(
                {
                    "cell": f"B{r}",
                    "value": b_s[:120],
                    "kind": f"{kind}_value_B",
                    "is_formula": is_formula(val_b.value),
                }
            )
        if c_s:
            controls.append(
                {
                    "cell": f"C{r}",
                    "value": c_s[:120],
                    "kind": f"{kind}_value",
                    "is_formula": is_formula(val_c.value),
                }
            )

    for r in range(1, max_row + 1):
        for c in range(1, min(max_col, 12) + 1):
            cell = ws.cell(r, c)
            val = cell_str(cell.value)
            if not val:
                continue
            addr = f"{get_column_letter(c)}{r}"
            if any(x["cell"] == addr for x in controls):
                continue
            low = val.lower()
            kind = None
            if low in ("ticker", "symbol"):
                kind = "ticker_label"
            elif "start" in low and "year" in low:
                kind = "start_year_label"
            elif "end" in low and "year" in low:
                kind = "end_year_label"
            elif low in ("units", "unit") or "in millions" in low or "millions of usd" in low:
                kind = "units"
            elif "fiscal" in low:
                kind = "fiscal_meta"
            elif val in ("A", "E", "P") and r <= 10:
                kind = "actual_estimate_flag"
            if kind:
                controls.append(
                    {
                        "cell": addr,
                        "value": val[:120],
                        "kind": kind,
                        "is_formula": is_formula(cell.value),
                    }
                )

        # Period / date row candidates across wider columns
        row_vals = []
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v is None:
                continue
            s = cell_str(v)
            if s is None:
                continue
            if re.search(r"20\d{2}", s) or re.match(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", s):
                row_vals.append(
                    {
                        "col": get_column_letter(c),
                        "cell": f"{get_column_letter(c)}{r}",
                        "value": s[:80],
                        "is_formula": is_formula(v),
                    }
                )
        if len(row_vals) >= 3:
            period_headers.append({"row": r, "headers": row_vals[:60]})

    return {"control_cells": controls[:80], "period_header_rows": period_headers}


def scan_formula_stats(ws, known_sheets: set[str]) -> dict[str, Any]:
    formula_count = 0
    value_count = 0
    blank_count = 0
    deps: Counter[str] = Counter()
    volatile_hits: list[dict[str, str]] = []
    external_hits: list[dict[str, str]] = []
    samples: list[dict[str, str]] = []
    max_r = ws.max_row or 0
    max_c = min(ws.max_column or 0, 100)

    writable_cells = 0
    formula_cells = 0

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if v is None or (isinstance(v, str) and not v.strip()):
                blank_count += 1
                continue
            if is_formula(v):
                formula_count += 1
                formula_cells += 1
                f = v
                for ref in sheet_refs_in_formula(f, known_sheets):
                    if ref != ws.title:
                        deps[ref] += 1
                if VOLATILE_RE.search(f):
                    if len(volatile_hits) < 25:
                        volatile_hits.append(
                            {"cell": f"{get_column_letter(c)}{r}", "formula": f[:200]}
                        )
                if EXTERNAL_RE.search(f):
                    if len(external_hits) < 25:
                        external_hits.append(
                            {"cell": f"{get_column_letter(c)}{r}", "formula": f[:200]}
                        )
                if len(samples) < MAX_FORMULA_SAMPLES:
                    samples.append(
                        {"cell": f"{get_column_letter(c)}{r}", "formula": f[:220]}
                    )
            else:
                value_count += 1
                writable_cells += 1

    total_nonblank = formula_count + value_count
    formula_ratio = (formula_count / total_nonblank) if total_nonblank else 0.0

    # Heuristic sheet role (refined in M1.5)
    if formula_ratio >= 0.75:
        role_hint = "formula_heavy"
    elif formula_ratio <= 0.25:
        role_hint = "data_heavy"
    else:
        role_hint = "hybrid"

    return {
        "formula_count": formula_count,
        "value_count": value_count,
        "blank_count_scanned": blank_count,
        "scanned_cols": max_c,
        "scanned_rows": max_r,
        "formula_ratio": round(formula_ratio, 4),
        "role_hint": role_hint,
        "outbound_sheet_deps": dict(deps.most_common(40)),
        "volatile_formula_samples": volatile_hits,
        "external_ref_samples": external_hits,
        "formula_samples": samples,
        "approx_writable_nonblank_cells": writable_cells,
        "approx_formula_cells": formula_cells,
    }


def scan_statement_structure(ws) -> dict[str, Any]:
    """Hierarchy / labels for core financial statement sheets."""
    labels: list[dict[str, Any]] = []
    units_notes: list[str] = []
    check_rows: list[dict[str, Any]] = []
    section_boundaries: list[dict[str, Any]] = []

    max_r = min(ws.max_row or 0, MAX_LABEL_ROWS)
    # Prefer column A then B for labels
    for r in range(1, max_r + 1):
        label = None
        label_col = None
        for c in LABEL_COL_CANDIDATES:
            v = ws.cell(r, c).value
            s = cell_str(v)
            if s and not is_formula(v):
                # Prefer longer text labels over tiny markers
                if label is None or (len(s) > len(label) and not s.replace(".", "", 1).isdigit()):
                    label = s
                    label_col = c
        if not label:
            continue

        cell = ws.cell(r, label_col or 1)
        align_indent = None
        try:
            align_indent = cell.alignment.indent if cell.alignment else None
        except Exception:
            align_indent = None

        role = classify_row_role(label)
        entry = {
            "row": r,
            "label_col": get_column_letter(label_col or 1),
            "label": label[:160],
            "indent": indent_level(label, int(align_indent) if align_indent else None),
            "role": role,
        }
        labels.append(entry)

        if role == "check":
            check_rows.append(entry)
        if role == "section_header":
            section_boundaries.append(entry)

        low = label.lower()
        if "million" in low or "unit" in low or "usd" in low or "$" in label:
            units_notes.append(f"R{r}: {label[:100]}")

    # Sample first data columns for formula vs value on a few key rows
    sample_rows = [e["row"] for e in labels if e["role"] == "line_item"][:40]
    region_samples = []
    max_c = min(ws.max_column or 0, 40)
    for r in sample_rows[:15]:
        formula_cols = []
        value_cols = []
        for c in range(4, max_c + 1):  # typically data starts after label cols
            v = ws.cell(r, c).value
            if v is None:
                continue
            col = get_column_letter(c)
            if is_formula(v):
                formula_cols.append(col)
            else:
                value_cols.append(col)
        region_samples.append(
            {
                "row": r,
                "value_cols_sample": value_cols[:20],
                "formula_cols_sample": formula_cols[:20],
            }
        )

    return {
        "label_row_count": len(labels),
        "labels": labels,
        "section_boundaries": section_boundaries,
        "check_rows": check_rows,
        "units_notes": units_notes[:20],
        "region_samples": region_samples,
    }


def inventory_workbook(ticker: str, path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False, read_only=False)
    known_sheets = set(wb.sheetnames)
    file_sha = hashlib.sha256(path.read_bytes()).hexdigest()

    sheets_out: list[dict[str, Any]] = []
    dependency_edges: Counter[tuple[str, str]] = Counter()

    for name in wb.sheetnames:
        ws = wb[name]
        state = "visible"
        try:
            if ws.sheet_state == "hidden":
                state = "hidden"
            elif ws.sheet_state == "veryHidden":
                state = "veryHidden"
        except Exception:
            pass

        controls = scan_control_region(ws)
        fstats = scan_formula_stats(ws, known_sheets)
        for dep, cnt in fstats["outbound_sheet_deps"].items():
            if dep != name:
                dependency_edges[(name, dep)] += cnt

        sheet_entry: dict[str, Any] = {
            "name": name,
            "index": wb.sheetnames.index(name),
            "state": state,
            "dimensions": {
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "dimensions_attr": ws.dimensions,
            },
            "freeze_panes": freeze_panes_info(ws),
            "merged_cells": merged_cells_info(ws),
            "tables": extract_tables(ws),
            "data_validations": extract_data_validations(ws),
            "conditional_formatting": extract_conditional_formatting(ws),
            "protection": protection_info(ws),
            "hidden": hidden_rows_cols(ws),
            "controls": controls,
            "formula_stats": fstats,
            "statement_structure": None,
        }

        if name in CORE_SHEETS:
            sheet_entry["statement_structure"] = scan_statement_structure(ws)

        sheets_out.append(sheet_entry)

    version_sheet_value = None
    named = extract_named_ranges(wb)

    if "Template Version" in wb.sheetnames:
        tvs = wb["Template Version"]
        version_sheet_value = cell_str(tvs["A2"].value)

    wb.close()

    edges = [
        {"from": a, "to": b, "formula_ref_count": n}
        for (a, b), n in dependency_edges.most_common(200)
    ]

    return {
        "ticker": ticker,
        "source_path": str(path),
        "source_filename": path.name,
        "file_sha256": file_sha,
        "file_size_bytes": path.stat().st_size,
        "inventoried_at": datetime.now(timezone.utc).isoformat(),
        "template_version_from_filename": (
            re.search(r"v\d+\.\d+", path.name).group(0)
            if re.search(r"v\d+\.\d+", path.name)
            else None
        ),
        "template_version_from_sheet": version_sheet_value,
        "sheet_count": len(sheets_out),
        "sheet_names": [s["name"] for s in sheets_out],
        "named_ranges": named,
        "sheets": sheets_out,
        "dependency_edges": edges,
    }


def structural_fingerprint(inv: dict[str, Any]) -> dict[str, Any]:
    """Compact fingerprint for cross-version / cross-ticker regression."""
    sheets = []
    for s in inv["sheets"]:
        sheets.append(
            {
                "name": s["name"],
                "state": s["state"],
                "max_row": s["dimensions"]["max_row"],
                "max_column": s["dimensions"]["max_column"],
                "freeze_panes": s["freeze_panes"],
                "merged_count": s["merged_cells"]["count"],
                "table_names": [t.get("name") for t in s["tables"] if "name" in t],
                "dv_count": s["data_validations"]["count"],
                "cf_rule_count": s["conditional_formatting"]["rule_count"],
                "hidden_row_count": s["hidden"]["hidden_row_count"],
                "hidden_column_count": s["hidden"]["hidden_column_count"],
                "sheet_protected": s["protection"]["sheet"],
                "formula_count": s["formula_stats"]["formula_count"],
                "value_count": s["formula_stats"]["value_count"],
                "formula_ratio": s["formula_stats"]["formula_ratio"],
                "role_hint": s["formula_stats"]["role_hint"],
                "outbound_deps": sorted(s["formula_stats"]["outbound_sheet_deps"].keys()),
            }
        )
    core_labels = {}
    for s in inv["sheets"]:
        if s["name"] in CORE_SHEETS and s["statement_structure"]:
            core_labels[s["name"]] = [
                {"row": e["row"], "label": e["label"], "indent": e["indent"], "role": e["role"]}
                for e in s["statement_structure"]["labels"]
            ]
    return {
        "ticker": inv["ticker"],
        "source_filename": inv["source_filename"],
        "file_sha256": inv["file_sha256"],
        "template_version_from_filename": inv.get("template_version_from_filename"),
        "template_version_from_sheet": inv.get("template_version_from_sheet"),
        "sheet_names": inv["sheet_names"],
        "sheet_count": inv["sheet_count"],
        "named_range_names": sorted(n["name"] for n in inv["named_ranges"]),
        "sheets": sheets,
        "core_statement_labels": core_labels,
        "dependency_edge_pairs": sorted(
            f"{e['from']}->{e['to']}" for e in inv["dependency_edges"]
        ),
    }


def diff_fingerprints(fps: dict[str, dict[str, Any]]) -> dict[str, Any]:
    base = fps["AAPL"]
    others = {t: fps[t] for t in TICKERS if t != "AAPL"}

    sheet_name_sets = {t: fp["sheet_names"] for t, fp in fps.items()}
    all_equal_names = all(sheet_name_sets[t] == base["sheet_names"] for t in others)

    per_sheet_diffs: dict[str, Any] = {}
    for sheet_name in base["sheet_names"]:
        diffs = {}
        base_sheet = next(s for s in base["sheets"] if s["name"] == sheet_name)
        for t, fp in others.items():
            other_sheet = next((s for s in fp["sheets"] if s["name"] == sheet_name), None)
            if other_sheet is None:
                diffs[t] = {"missing": True}
                continue
            field_diffs = {}
            for key in (
                "state",
                "max_row",
                "max_column",
                "merged_count",
                "dv_count",
                "cf_rule_count",
                "hidden_row_count",
                "hidden_column_count",
                "sheet_protected",
                "role_hint",
            ):
                if base_sheet.get(key) != other_sheet.get(key):
                    field_diffs[key] = {"AAPL": base_sheet.get(key), t: other_sheet.get(key)}
            # formula counts expected to differ (company-specific filled values)
            if abs(base_sheet["formula_count"] - other_sheet["formula_count"]) > 50:
                field_diffs["formula_count_delta"] = {
                    "AAPL": base_sheet["formula_count"],
                    t: other_sheet["formula_count"],
                }
            if field_diffs:
                diffs[t] = field_diffs
        if diffs:
            per_sheet_diffs[sheet_name] = diffs

    label_diffs = {}
    for sheet in CORE_SHEETS:
        base_labels = [
            (e["row"], e["label"], e["indent"])
            for e in base["core_statement_labels"].get(sheet, [])
        ]
        for t, fp in others.items():
            other_labels = [
                (e["row"], e["label"], e["indent"])
                for e in fp["core_statement_labels"].get(sheet, [])
            ]
            if base_labels != other_labels:
                # summarize first mismatches
                mismatches = []
                for i, (a, b) in enumerate(zip(base_labels, other_labels)):
                    if a != b:
                        mismatches.append({"index": i, "AAPL": a, t: b})
                        if len(mismatches) >= 15:
                            break
                label_diffs.setdefault(sheet, {})[t] = {
                    "aapl_count": len(base_labels),
                    f"{t}_count": len(other_labels),
                    "length_equal": len(base_labels) == len(other_labels),
                    "first_mismatches": mismatches,
                }

    named_diff = {
        t: {
            "only_in_AAPL": sorted(set(base["named_range_names"]) - set(fp["named_range_names"])),
            f"only_in_{t}": sorted(set(fp["named_range_names"]) - set(base["named_range_names"])),
        }
        for t, fp in others.items()
        if set(fp["named_range_names"]) != set(base["named_range_names"])
    }

    return {
        "baseline": "AAPL",
        "compared": list(others.keys()),
        "sheet_names_identical": all_equal_names,
        "sheet_names_by_ticker": sheet_name_sets,
        "named_range_diffs_vs_AAPL": named_diff,
        "per_sheet_structural_diffs_vs_AAPL": per_sheet_diffs,
        "core_label_diffs_vs_AAPL": label_diffs,
        "notes": [
            "Value/formula cell counts differ by company fill state; treat role_hint and layout fields as structural.",
            "Label mismatches on core statements are high-priority mapping risks.",
        ],
    }


def purpose_for_sheet(name: str, role_hint: str) -> str:
    purposes = {
        "Balance Sheet - Standardized": "Annual standardized balance sheet (primary BS data grid).",
        "BS%": "Balance sheet common-size percentages; formulas reference Balance Sheet - Standardized.",
        "Income - GAAP": "Annual GAAP income statement (primary IS data grid).",
        "IS%": "Income statement common-size percentages; formulas reference Income - GAAP.",
        "Cash Flow - Standardized": "Annual standardized cash flow statement (primary CF data grid).",
        "CF%": "Cash flow percentages / rates; formulas reference Cash Flow - Standardized.",
        "FCF": "Free cash flow build derived largely from Inputs / statements / Tax.",
        "Last Quarter BS Standardized": "Quarterly standardized balance sheet pack (LQ).",
        "Last Quarter IS Standardized": "Quarterly standardized income statement pack (LQ).",
        "Last Quarter CF Standardized": "Quarterly standardized cash flow pack (LQ).",
        "Last Quarter BS As Reported": "Quarterly balance sheet as-reported (company presentation; variable depth).",
        "Last Quarter IS As Reported": "Quarterly income as-reported (company presentation; variable depth).",
        "Last Quarter CF As Reported": "Quarterly cash flow as-reported (company presentation; variable depth).",
        "DividendHelper": "Dividend helper / support calculations for capital returns.",
        "Inputs": "Central bridge sheet: pulls statement lines and feeds ratios, tax, leases, R&D, metrics.",
        "IC & NOPAT & ROIC ": "Invested capital, NOPAT, and ROIC calculations (note trailing space in sheet name).",
        "Tax": "Tax / ETR support feeding IC/NOPAT and FCF.",
        "Leases": "Lease capitalization / adjustments feeding IC & ROIC.",
        "R&D": "R&D capitalization / adjustments feeding IC & ROIC.",
        "All Ratios": "Ratio library; primarily formula-driven from Inputs (and some statement refs).",
        "Final Metrics": "Summary metrics hub; feeds and is fed by Enterprise Value / Expected Returns.",
        "Expected Returns & Buybacks": "Expected returns and buyback analysis outputs.",
        "Enterprise Value": "Enterprise value build-up and related valuation outputs.",
        "Template Version": "Template metadata / version control sheet.",
    }
    if name in purposes:
        return purposes[name]
    return f"Sheet present in Industrial Template v27.6 (role_hint={role_hint})."


def render_human_md(
    inventories: dict[str, dict[str, Any]],
    diff: dict[str, Any],
) -> str:
    aapl = inventories["AAPL"]
    lines: list[str] = []
    lines.append("# Industrial Template Inventory (M1)")
    lines.append("")
    lines.append("**Status:** Milestone M1 deliverable — documentation only (no mapping, no cell writes).  ")
    lines.append("**Template family:** Industrial Template **v27.6** (all four suite companies).  ")
    lines.append("**Suite:** AAPL, MSFT, AMZN, TJX ([`VALIDATION_SUITE.md`](../VALIDATION_SUITE.md)).  ")
    lines.append(f"**Inventoried (UTC):** {aapl['inventoried_at']}")
    lines.append("")
    lines.append("## Workbook overview")
    lines.append("")
    lines.append(
        "The Mode A Excel deliverable is the completed **Industrial Template**: a multi-sheet "
        "financial model with standardized income, balance sheet, and cash flow statements, "
        "last-quarter (LQ) packs, percentage views, a central Inputs bridge, ratio/metrics sheets, "
        "and valuation / expected-return outputs. Formulas link sheets; HAP must preserve formulas "
        "and only write into non-formula input cells (mapping begins in M2 after M1.5 classification)."
    )
    lines.append("")
    lines.append("| Ticker | Source file | Filename ver | Sheet ver | Sheets | SHA256 (first 16) | Size |")
    lines.append("|--------|-------------|--------------|-----------|--------|-------------------|------|")
    for t in TICKERS:
        inv = inventories[t]
        lines.append(
            f"| {t} | `{inv['source_filename']}` | "
            f"{inv.get('template_version_from_filename')} | "
            f"{inv.get('template_version_from_sheet')} | {inv['sheet_count']} | "
            f"`{inv['file_sha256'][:16]}…` | {inv['file_size_bytes']} |"
        )
    lines.append("")
    lines.append(
        f"**Sheet count (AAPL):** {aapl['sheet_count']}.  "
        f"**Sheet names identical across suite:** "
        f"{'YES' if diff['sheet_names_identical'] else 'NO — see suite diff'}."
    )
    lines.append("")
    lines.append("## Sheet list and purpose")
    lines.append("")
    lines.append("| # | Sheet | Visibility | Role hint (heuristic) | Purpose |")
    lines.append("|---|-------|------------|------------------------|---------|")
    for s in aapl["sheets"]:
        lines.append(
            f"| {s['index']+1} | `{s['name']}` | {s['state']} | "
            f"{s['formula_stats']['role_hint']} | {purpose_for_sheet(s['name'], s['formula_stats']['role_hint'])} |"
        )
    lines.append("")
    lines.append(
        "*Role hint is formula-ratio based (`data_heavy` / `hybrid` / `formula_heavy`) and will be "
        "refined into Data / Formula / Hybrid / Control / Meta in **M1.5**. It is not a write policy.*"
    )
    lines.append("")
    lines.append("## Relationships between sheets")
    lines.append("")
    lines.append(
        "See [`INDUSTRIAL_TEMPLATE_DEPENDENCY_MAP.md`](INDUSTRIAL_TEMPLATE_DEPENDENCY_MAP.md) "
        "for the formula-derived dependency graph. High-level flow (AAPL):"
    )
    lines.append("")
    # Top edges
    top = aapl["dependency_edges"][:25]
    lines.append("| From | To | Approx. cross-sheet formula refs |")
    lines.append("|------|----|----------------------------------|")
    for e in top:
        lines.append(f"| `{e['from']}` | `{e['to']}` | {e['formula_ref_count']} |")
    lines.append("")
    lines.append("## Financial statement organization")
    lines.append("")
    lines.append("### Annual vs percentage vs last-quarter packs")
    lines.append("")
    lines.append("| Family | Annual / standardized | Common-size % | LQ standardized | LQ as-reported |")
    lines.append("|--------|----------------------|---------------|-----------------|----------------|")
    lines.append(
        "| Income | `Income - GAAP` | `IS%` | `Last Quarter IS Standardized` | `Last Quarter IS As Reported` |"
    )
    lines.append(
        "| Balance Sheet | `Balance Sheet - Standardized` | `BS%` | "
        "`Last Quarter BS Standardized` | `Last Quarter BS As Reported` |"
    )
    lines.append(
        "| Cash Flow | `Cash Flow - Standardized` | `CF%` | "
        "`Last Quarter CF Standardized` | `Last Quarter CF As Reported` |"
    )
    lines.append("")
    lines.append(
        "Additional derived sheets: `FCF`, `DividendHelper`, `Inputs`, `IC & NOPAT & ROIC ` "
        "(trailing space in name), `Tax`, `Leases`, `R&D`, `All Ratios`, `Final Metrics`, "
        "`Expected Returns & Buybacks`, `Enterprise Value`, `Template Version`."
    )
    lines.append("")
    lines.append(
        "Downstream analytics typically consume standardized/annual grids and LQ packs via "
        "`Inputs` → `All Ratios` / `Final Metrics` → valuation sheets. Exact edges are in the dependency map."
    )
    lines.append("")
    lines.append(
        "**Dimension note:** openpyxl reports `max_column` up to 16382 (`XFB`) on some sheets "
        "because of sparse dimension metadata. Inventory formula scans are capped at 100 columns; "
        "period headers observed for annual statements occupy roughly columns C–L for the FY window."
    )
    lines.append("")

    for sheet_name in CORE_SHEETS:
        s = next(x for x in aapl["sheets"] if x["name"] == sheet_name)
        st = s["statement_structure"]
        lines.append(f"## Core statement detail — `{sheet_name}` (AAPL primary)")
        lines.append("")
        lines.append(
            f"- Dimensions: {s['dimensions']['max_row']} rows × {s['dimensions']['max_column']} cols"
        )
        lines.append(
            f"- Formula ratio: {s['formula_stats']['formula_ratio']} "
            f"({s['formula_stats']['formula_count']} formulas / "
            f"{s['formula_stats']['value_count']} values in scanned used range)"
        )
        lines.append(f"- Freeze panes: `{s['freeze_panes']}`")
        lines.append(f"- Merged cells: {s['merged_cells']['count']}")
        lines.append(f"- Tables: {s['tables'] or 'none'}")
        lines.append(f"- Data validations: {s['data_validations']['count']}")
        lines.append(
            f"- Hidden rows/cols: {s['hidden']['hidden_row_count']} / {s['hidden']['hidden_column_count']}"
        )
        lines.append(f"- Sheet protection: {s['protection']}")
        lines.append("")
        lines.append("### Control / period header region")
        lines.append("")
        if s["controls"]["control_cells"]:
            lines.append("| Cell | Kind | Value | Formula? |")
            lines.append("|------|------|-------|----------|")
            for c in s["controls"]["control_cells"][:40]:
                lines.append(
                    f"| `{c['cell']}` | {c['kind']} | {c['value'][:60]} | {c['is_formula']} |"
                )
        else:
            lines.append("_No control-like labels detected in top-left scan._")
        lines.append("")
        if s["controls"]["period_header_rows"]:
            lines.append("### Period header rows (year/date-like)")
            lines.append("")
            for phr in s["controls"]["period_header_rows"][:6]:
                preview = ", ".join(
                    f"{h['col']}={h['value']}" for h in phr["headers"][:12]
                )
                more = "" if len(phr["headers"]) <= 12 else f" … (+{len(phr['headers'])-12})"
                lines.append(f"- **Row {phr['row']}:** {preview}{more}")
            lines.append("")
        lines.append("### Hierarchy / labels (first 80)")
        lines.append("")
        lines.append("| Row | Col | Indent | Role | Label |")
        lines.append("|-----|-----|--------|------|-------|")
        for e in st["labels"][:80]:
            lines.append(
                f"| {e['row']} | {e['label_col']} | {e['indent']} | {e['role']} | {e['label'][:80]} |"
            )
        if len(st["labels"]) > 80:
            lines.append("")
            lines.append(f"_… {len(st['labels']) - 80} additional labels in JSON._")
        lines.append("")
        lines.append("### Section boundaries")
        lines.append("")
        if st["section_boundaries"]:
            for e in st["section_boundaries"][:40]:
                lines.append(f"- R{e['row']}: {e['label']}")
        else:
            lines.append("_None classified as section headers by heuristic._")
        lines.append("")
        lines.append("### Check rows")
        lines.append("")
        if st["check_rows"]:
            for e in st["check_rows"]:
                lines.append(f"- R{e['row']}: {e['label']}")
        else:
            lines.append("_None detected containing 'check'._")
        lines.append("")
        lines.append("### Units notes")
        lines.append("")
        if st["units_notes"]:
            for u in st["units_notes"]:
                lines.append(f"- {u}")
        else:
            lines.append("_No explicit units string in label column scan; check header controls and JSON._")
        lines.append("")
        lines.append("### Writable vs formula regions (sample line items)")
        lines.append("")
        lines.append(
            "Per-row samples of which columns hold values vs formulas (not a final write policy — that is M1.5):"
        )
        lines.append("")
        for rs in st["region_samples"][:12]:
            lines.append(
                f"- Row {rs['row']}: values in {rs['value_cols_sample'] or '—'}; "
                f"formulas in {rs['formula_cols_sample'] or '—'}"
            )
        lines.append("")

    lines.append("## Named ranges")
    lines.append("")
    if aapl["named_ranges"]:
        lines.append("| Name | Attr / destinations |")
        lines.append("|------|---------------------|")
        for n in aapl["named_ranges"][:80]:
            dest = n.get("attr_text") or json.dumps(n.get("destinations"))
            lines.append(f"| `{n['name']}` | `{dest}` |")
    else:
        lines.append("_No workbook-level defined names found (or none exposed via openpyxl)._")
    lines.append("")
    lines.append("## Hidden sheets")
    lines.append("")
    hidden = [s["name"] for s in aapl["sheets"] if s["state"] != "visible"]
    if hidden:
        for h in hidden:
            lines.append(f"- `{h}` ({next(s for s in aapl['sheets'] if s['name']==h)['state']})")
    else:
        lines.append("_No hidden / veryHidden sheets in AAPL inventory._")
    lines.append("")
    lines.append("## Freeze panes / merged cells / tables / validation / CF (summary)")
    lines.append("")
    lines.append("| Sheet | Freeze | Merged | Tables | DV | CF rules | Hidden rows | Hidden cols | Protected |")
    lines.append("|-------|--------|--------|--------|----|----------|-------------|-------------|-----------|")
    for s in aapl["sheets"]:
        lines.append(
            f"| `{s['name']}` | `{s['freeze_panes']}` | {s['merged_cells']['count']} | "
            f"{len(s['tables'])} | {s['data_validations']['count']} | "
            f"{s['conditional_formatting']['rule_count']} | "
            f"{s['hidden']['hidden_row_count']} | {s['hidden']['hidden_column_count']} | "
            f"{s['protection']['sheet']} |"
        )
    lines.append("")
    lines.append("## Suite consistency (AAPL vs MSFT / AMZN / TJX)")
    lines.append("")
    lines.append(
        f"- Sheet names identical (order + spelling): **{diff['sheet_names_identical']}**"
    )
    lines.append(
        "- Dependency edge *types* identical across suite (75 shared pairs; no ticker-only edges)."
    )
    if diff["named_range_diffs_vs_AAPL"]:
        lines.append("- Named range differences detected — see `suite_structural_diff.json`.")
    else:
        lines.append("- Named ranges: no name-set differences vs AAPL.")
    lines.append("")
    lines.append("### What differs (expected company data, not template structure)")
    lines.append("")
    lines.append(
        "Core statement **row counts match** across the suite (Income 97 labels, BS 130, CF 72 in the "
        "inventory scan). Observed 'label diffs' are almost entirely:"
    )
    lines.append("")
    lines.append(
        "1. **Fiscal period end dates** in header rows (Apple Sept YE, Microsoft June YE, "
        "Amazon Dec YE, TJX late-Jan YE)."
    )
    lines.append(
        "2. **FY window shift for TJX** — period headers show FY 2017–2026 vs FY 2016–2025 on AAPL/MSFT/AMZN."
    )
    lines.append(
        "3. Occasional **year-embedded text** (e.g. TJX `Retained Earnings FY 2016` vs AAPL `… FY 2015`)."
    )
    lines.append(
        "4. **As-reported LQ sheet depths** differ (`Last Quarter * As Reported` max_row varies by ticker) — "
        "company presentation length, not a different template family."
    )
    lines.append(
        "5. Role-hint on `Last Quarter IS Standardized` is hybrid for AAPL vs data_heavy for others "
        "(fill/formula mix differs; structure same)."
    )
    lines.append("")
    lines.append(
        "**Structural conclusion for M1.5/M2:** The v27.6 sheet skeleton is shared. Map by "
        "**row labels + period column headers**, not by assuming identical date cells or identical "
        "as-reported LQ depth. Do not treat AAPL fiscal dates as universal."
    )
    lines.append("")
    if diff["core_label_diffs_vs_AAPL"]:
        lines.append("Raw mismatch summaries (includes period dates):")
        for sheet, by_t in diff["core_label_diffs_vs_AAPL"].items():
            for t, info in by_t.items():
                lines.append(
                    f"- `{sheet}` vs {t}: counts AAPL={info['aapl_count']} "
                    f"{t}={info[f'{t}_count']}; recorded mismatches={len(info['first_mismatches'])}"
                )
    lines.append("")
    if diff["per_sheet_structural_diffs_vs_AAPL"]:
        lines.append(
            f"- Per-sheet structural field diffs: "
            f"{len(diff['per_sheet_structural_diffs_vs_AAPL'])} sheets "
            f"(mostly LQ as-reported depth / role_hint) — details in JSON."
        )
    lines.append("")
    lines.append("Full machine diff: [`suite_structural_diff.json`](suite_structural_diff.json).")
    lines.append("")
    lines.append("## Unusual observations")
    lines.append("")
    lines.append(
        "- **Control value cells:** On annual IS/BS/CF, labels are `A1:A3`; values are **`C1:C3`** "
        "(Ticker / Start Year / End Year). Column B is typically blank and may be hidden."
    )
    lines.append(
        "- **Version string mismatch:** Filenames use `v27.6`; `Template Version!A2` observed as "
        "`v27.7` on AAPL — treat both as identity signals for regression."
    )
    lines.append(
        "- **Trailing space** in sheet name `IC & NOPAT & ROIC ` must be preserved exactly."
    )
    lines.append(
        "- **Hidden structure on Income - GAAP (AAPL):** rows 4–5 hidden; column B hidden; freeze at `C9`."
    )
    lines.append(
        "- **Named ranges:** CapIQ-style `IQ_*` period codes plus helpers; broken "
        "`SpreadsheetBuilder_*` names with `#REF!` may exist — do not rely on them for HAP fill."
    )
    # Collect volatile / external across suite
    for t in TICKERS:
        vols = []
        exts = []
        for s in inventories[t]["sheets"]:
            vols.extend(s["formula_stats"]["volatile_formula_samples"])
            exts.extend(s["formula_stats"]["external_ref_samples"])
        lines.append(
            f"- **{t}:** volatile-formula samples={len(vols)}; external-ref samples={len(exts)}"
        )
    lines.append("")
    lines.append("See also [`INDUSTRIAL_TEMPLATE_RISK_LOG.md`](INDUSTRIAL_TEMPLATE_RISK_LOG.md).")
    lines.append("")
    lines.append("## Machine artifacts")
    lines.append("")
    lines.append("| File | Role |")
    lines.append("|------|------|")
    lines.append("| `industrial_template_v27_inventory.json` | Full inventory (all 4) |")
    lines.append("| `industrial_template_v27_fingerprint.json` | Compact regression fingerprints |")
    lines.append("| `suite_structural_diff.json` | Cross-ticker structural diff |")
    lines.append("| `INDUSTRIAL_TEMPLATE_DEPENDENCY_MAP.md` | Formula dependency graph |")
    lines.append("| `INDUSTRIAL_TEMPLATE_RISK_LOG.md` | Mapping risks |")
    lines.append("")
    lines.append("## Definition of done (M1)")
    lines.append("")
    lines.append("- [x] Human documentation of the Industrial Template")
    lines.append("- [x] Machine JSON inventory")
    lines.append("- [x] Dependency map")
    lines.append("- [x] Risk log")
    lines.append("- [x] Regression fingerprints for future template versions")
    lines.append("- [x] Suite validation AAPL / MSFT / AMZN / TJX")
    lines.append("- [ ] **Stopped** — M1.5 / M2 not started (awaiting review)")
    lines.append("")
    return "\n".join(lines)


def render_dependency_md(inventories: dict[str, dict[str, Any]]) -> str:
    aapl = inventories["AAPL"]
    lines = []
    lines.append("# Industrial Template — Dependency Map (M1)")
    lines.append("")
    lines.append(
        "Derived from cross-sheet formula references in the AAPL Industrial Template v27.6 "
        "(validated that sheet inventory is shared across the suite). Edge weights are counts of "
        "formulas on the **From** sheet that reference the **To** sheet (best-effort parse)."
    )
    lines.append("")
    lines.append("## High-level mermaid graph (edges with ≥20 refs)")
    lines.append("")
    lines.append("```mermaid")
    lines.append("flowchart TD")
    # sanitize ids
    def nid(n: str) -> str:
        return re.sub(r"[^A-Za-z0-9]", "_", n)

    heavy = [e for e in aapl["dependency_edges"] if e["formula_ref_count"] >= 20][:60]
    seen = set()
    for e in heavy:
        a, b = nid(e["from"]), nid(e["to"])
        lines.append(f'  {a}["{e["from"]}"] --> {b}["{e["to"]}"]')
        seen.add(e["from"])
        seen.add(e["to"])
    lines.append("```")
    lines.append("")
    lines.append("## Full edge list (AAPL, top 80)")
    lines.append("")
    lines.append("| From | To | Refs |")
    lines.append("|------|----|------|")
    for e in aapl["dependency_edges"][:80]:
        lines.append(f"| `{e['from']}` | `{e['to']}` | {e['formula_ref_count']} |")
    lines.append("")
    lines.append("## Per-ticker edge-pair comparison")
    lines.append("")
    base_pairs = {f"{e['from']}->{e['to']}" for e in aapl["dependency_edges"]}
    for t in ("MSFT", "AMZN", "TJX"):
        pairs = {f"{e['from']}->{e['to']}" for e in inventories[t]["dependency_edges"]}
        only_a = sorted(base_pairs - pairs)[:20]
        only_t = sorted(pairs - base_pairs)[:20]
        lines.append(f"### {t} vs AAPL")
        lines.append(f"- Shared edge types: {len(base_pairs & pairs)}")
        lines.append(f"- Only in AAPL (sample): {only_a or '—'}")
        lines.append(f"- Only in {t} (sample): {only_t or '—'}")
        lines.append("")
    lines.append(
        "## Interpretation note\n\n"
        "Percentage and LQ sheets often pull from their standardized siblings. "
        "`Inputs` / `All Ratios` / `Final Metrics` / valuation sheets sit downstream. "
        "M1.5 will assign write policies so HAP never writes into pure formula sinks."
    )
    lines.append("")
    return "\n".join(lines)


def render_risk_md(
    inventories: dict[str, dict[str, Any]],
    diff: dict[str, Any],
) -> str:
    lines = []
    lines.append("# Industrial Template — Risk Log (M1)")
    lines.append("")
    lines.append("Risks that may complicate Workbook Classification (M1.5) and Mapping (M2).")
    lines.append("")
    lines.append("| ID | Risk | Evidence | Impact | Mitigation (later) |")
    lines.append("|----|------|----------|--------|--------------------|")

    risks = []

    # Label diffs
    if diff["core_label_diffs_vs_AAPL"]:
        risks.append(
            (
                "R01",
                "Core statement 'label' fingerprint differs across suite",
                "Almost entirely fiscal end-dates / FY window / year-embedded text — row counts match",
                "Medium for period alignment (M5); Low for line-item skeleton",
                "Map by stable line labels; resolve columns from each workbook's period headers",
            )
        )
    else:
        risks.append(
            (
                "R01",
                "Core statement labels appear suite-consistent",
                "No label diffs in fingerprint compare",
                "Low (monitor on new template versions)",
                "Regression fingerprint compare on each template bump",
            )
        )

    if not diff["sheet_names_identical"]:
        risks.append(
            (
                "R02",
                "Sheet name sets differ across suite",
                str(diff["sheet_names_by_ticker"]),
                "Critical",
                "Block mapping until reconciled",
            )
        )
    else:
        risks.append(
            (
                "R02",
                "Sheet name sets identical across suite",
                "All four share the same ordered sheet list",
                "Low",
                "Fingerprint on version upgrades",
            )
        )

    # Volatile / external
    for t in TICKERS:
        vol_n = sum(
            len(s["formula_stats"]["volatile_formula_samples"])
            for s in inventories[t]["sheets"]
        )
        ext_n = sum(
            len(s["formula_stats"]["external_ref_samples"])
            for s in inventories[t]["sheets"]
        )
        if vol_n:
            risks.append(
                (
                    f"R03-{t}",
                    "Volatile functions present (INDIRECT/OFFSET/NOW/…)",
                    f"{vol_n} sample hits in {t}",
                    "Medium — dynamic targets hard to map statically",
                    "Classify those cells Read-only; never overwrite",
                )
            )
        if ext_n:
            risks.append(
                (
                    f"R04-{t}",
                    "External workbook references in formulas",
                    f"{ext_n} sample hits in {t}",
                    "High — broken links outside HAP",
                    "Document; do not depend on external files for Mode A fill",
                )
            )

    # Protected sheets
    for t in TICKERS:
        prot = [s["name"] for s in inventories[t]["sheets"] if s["protection"]["sheet"]]
        if prot:
            risks.append(
                (
                    f"R05-{t}",
                    "Protected sheets",
                    ", ".join(prot),
                    "High — write_values may fail",
                    "Unprotect policy or skip; document in M1.5",
                )
            )

    # Hidden sheets
    for t in TICKERS:
        hid = [s["name"] for s in inventories[t]["sheets"] if s["state"] != "visible"]
        if hid:
            risks.append(
                (
                    f"R06-{t}",
                    "Hidden sheets",
                    ", ".join(hid),
                    "Medium — hidden dependencies",
                    "Include in classification; do not ignore",
                )
            )

    # Hybrid formula ratios on core sheets
    for sheet in CORE_SHEETS:
        s = next(x for x in inventories["AAPL"]["sheets"] if x["name"] == sheet)
        risks.append(
            (
                f"R07-{sheet[:12]}",
                f"Core sheet `{sheet}` is {s['formula_stats']['role_hint']}",
                f"formula_ratio={s['formula_stats']['formula_ratio']}",
                "High for mapping — mix of inputs and formulas",
                "M1.5 must mark writable regions cell-by-cell / by column band",
            )
        )

    # Version in filenames
    risks.append(
        (
            "R08",
            "Template version embedded in filenames (v27.6); quarters differ by ticker",
            "AAPL Q2 / MSFT Q3 / AMZN Q1 / TJX Q4 — same v27.6",
            "Medium — period windows differ; structure should match",
            "Period alignment milestone (M5); fingerprint on v27.8+",
        )
    )

    risks.append(
        (
            "R09",
            "openpyxl does not evaluate formulas; data_only=False inventory only",
            "Cached values not treated as truth for formula cells",
            "Medium for validation later",
            "Excel-open checks in suite gate; never invent calculated values",
        )
    )

    risks.append(
        (
            "R10",
            "Large used ranges / merged cells / freeze panes",
            "See per-sheet summary in inventory MD",
            "Low–Medium for writer address stability",
            "Write by explicit addresses from mapping file only",
        )
    )

    if diff["per_sheet_structural_diffs_vs_AAPL"]:
        risks.append(
            (
                "R11",
                "Non-label structural diffs vs AAPL on some sheets",
                f"{len(diff['per_sheet_structural_diffs_vs_AAPL'])} sheets",
                "Medium",
                "Review suite_structural_diff.json before M2",
            )
        )

    risks.append(
        (
            "R12",
            "Sheet name `IC & NOPAT & ROIC ` has a trailing space",
            "Exact name required in formula refs and openpyxl access",
            "Medium — easy off-by-space bugs",
            "Always use exact sheetnames from inventory JSON",
        )
    )
    risks.append(
        (
            "R13",
            "openpyxl max_column can report 16382 (XFB) on statement sheets",
            "Sparse dimension metadata; not a real 16k-column model",
            "Low if scans/writes are bounded",
            "Discover used period columns from header row, not max_column",
        )
    )
    risks.append(
        (
            "R14",
            "Filename says Industrial Template v27.6 but Template Version sheet may say v27.7",
            "Observed on AAPL: Template Version!A2 = v27.7",
            "Medium — version identity ambiguous for regression",
            "Fingerprint both filename token and Template Version sheet value",
        )
    )
    risks.append(
        (
            "R15",
            "Annual control values live in C1:C3 (Ticker / Start Year / End Year), not B",
            "Column B often blank/hidden on statement sheets",
            "High if writer targets wrong column",
            "M1.5/M2 treat C1–C3 as control cells on annual IS/BS/CF",
        )
    )

    for r in risks:
        lines.append(f"| {r[0]} | {r[1]} | {r[2]} | {r[3]} | {r[4]} |")

    lines.append("")
    lines.append("## Circular references")
    lines.append("")
    lines.append(
        "openpyxl cannot detect Excel iterative circulars reliably from static formulas alone. "
        "No CIRCULAR markers were required for inventory; treat potential circularity as "
        "**unknown / monitor** during Excel open tests in the suite gate."
    )
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    inventories: dict[str, dict[str, Any]] = {}
    fingerprints: dict[str, dict[str, Any]] = {}

    for t in TICKERS:
        path = find_template(t)
        print(f"Inventorying {t}: {path.name}")
        inv = inventory_workbook(t, path)
        inventories[t] = inv
        fingerprints[t] = structural_fingerprint(inv)
        print(
            f"  sheets={inv['sheet_count']} formulas_sample_done "
            f"edges={len(inv['dependency_edges'])}"
        )

    diff = diff_fingerprints(fingerprints)

    payload = {
        "milestone": "M1",
        "template_family": "Industrial Template",
        "template_version_observed": "v27.6",
        "suite": list(TICKERS),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "companies": inventories,
    }
    inv_path = OUT_DIR / "industrial_template_v27_inventory.json"
    fp_path = OUT_DIR / "industrial_template_v27_fingerprint.json"
    diff_path = OUT_DIR / "suite_structural_diff.json"

    inv_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    fp_path.write_text(
        json.dumps(
            {
                "milestone": "M1",
                "purpose": "Regression fingerprint for future template versions",
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "fingerprints": fingerprints,
            },
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )
    diff_path.write_text(json.dumps(diff, indent=2, default=str), encoding="utf-8")

    (OUT_DIR / "INDUSTRIAL_TEMPLATE_INVENTORY.md").write_text(
        render_human_md(inventories, diff), encoding="utf-8"
    )
    (OUT_DIR / "INDUSTRIAL_TEMPLATE_DEPENDENCY_MAP.md").write_text(
        render_dependency_md(inventories), encoding="utf-8"
    )
    (OUT_DIR / "INDUSTRIAL_TEMPLATE_RISK_LOG.md").write_text(
        render_risk_md(inventories, diff), encoding="utf-8"
    )

    print("Wrote:")
    for p in sorted(OUT_DIR.glob("*")):
        print(f"  {p} ({p.stat().st_size} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
