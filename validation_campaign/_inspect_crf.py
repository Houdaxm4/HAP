"""Deep-inspect Custom_Run_Filter workbooks for CustomRunData design."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"C:\Users\G\HAP\validation_campaign\universe")
TICKERS = ["AAPL", "MSFT", "AMZN", "TJX"]
OUT = Path(r"C:\Users\G\HAP\validation_campaign\_crf_inventory.json")
YEAR_RE = re.compile(r"^(FY\s*)?(19|20)\d{2}$|^(CY\s*)?(19|20)\d{2}$|^\d{4}(-|/)Q[1-4]$|^FY\d{2,4}$", re.I)


def list_folder(ticker: str) -> list[dict]:
    d = ROOT / ticker
    files = []
    for p in sorted(d.iterdir()):
        if p.name.startswith("~$"):
            continue
        files.append({"name": p.name, "bytes": p.stat().st_size, "suffix": p.suffix})
    return files


def cell_str(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    if isinstance(v, (int, float)):
        return str(v)
    return str(v).strip() or None


def looks_like_year(s: str | None) -> bool:
    if not s:
        return False
    s2 = s.strip()
    if YEAR_RE.match(s2):
        return True
    if re.match(r"^(19|20)\d{2}$", s2):
        return True
    if re.match(r"^FY\s*(19|20)\d{2}$", s2, re.I):
        return True
    if re.match(r"^(19|20)\d{2}[A-Z]?$", s2, re.I):
        return True
    return False


def inspect_sheet_preview(ws, max_preview_rows: int = 15, max_cols: int = 30) -> dict:
    """Read first N rows; estimate dimensions by scanning."""
    rows_data: list[list] = []
    used_cols: set[int] = set()
    max_r = 0
    max_c = 0
    row_count = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_count = i
        # find last non-empty in this row
        last = 0
        for j, v in enumerate(row, start=1):
            if v is not None and (not isinstance(v, str) or v.strip()):
                last = j
                used_cols.add(j)
                if j > max_c:
                    max_c = j
        if last:
            max_r = i
        if i <= max_preview_rows:
            # take columns that have any data in preview window later; for now keep first max_cols of non-empty span
            preview = []
            for j in range(1, min(max(last, 1), max_cols) + 1):
                preview.append(cell_str(row[j - 1]) if j - 1 < len(row) else None)
            # trim trailing Nones for compactness but keep alignment
            while preview and preview[-1] is None:
                preview.pop()
            rows_data.append(preview)

    # refine preview to only columns that had data in first 15 rows
    # Re-scan would be costly; use what we have
    preview_used = set()
    for r in rows_data:
        for j, v in enumerate(r, start=1):
            if v is not None:
                preview_used.add(j)
    # rebuild clipped to first 30 used cols
    used_sorted = sorted(preview_used)[:max_cols]

    return {
        "title": ws.title,
        "approx_max_row": max_r or row_count,
        "approx_max_column": max_c,
        "rows_scanned": row_count,
        "preview_used_col_indexes": used_sorted,
        "first_15_rows": rows_data,
    }


def extract_summary_headers(ws) -> dict:
    """Row 1 headers; also sample row 2 types."""
    headers = []
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if row1:
        for j, v in enumerate(row1, start=1):
            s = cell_str(v)
            if s is not None:
                headers.append({"col": j, "label": s})
    # detect year-like headers
    year_headers = [h for h in headers if looks_like_year(h["label"])]
    non_year = [h for h in headers if not looks_like_year(h["label"])]
    # peek a few data rows for shape
    sample_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
        sample_rows.append([cell_str(v) for v in row[: min(len(row), 40)]])
    return {
        "header_count": len(headers),
        "headers": [h["label"] for h in headers],
        "headers_detailed": headers,
        "year_like_headers": [h["label"] for h in year_headers],
        "non_year_headers": [h["label"] for h in non_year],
        "sample_rows_2_6": sample_rows,
    }


def extract_ticker_labels(ws, ticker: str) -> dict:
    """
    Vertical metadata (col A[/B]) + tabular headers further down.
    Heuristic: metadata is key-value until we hit a blank stretch or a row
    that looks like a table header (multiple non-empty cells, year-like).
    """
    meta_kv: list[dict] = []
    col_a_labels: list[str] = []
    table_header_rows: list[dict] = []
    all_col_a: list[dict] = []

    # Scan first ~200 rows carefully, then sample labels further if dense
    year_col_hits = 0
    first_table_row = None
    consecutive_wide = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # normalize row cells
        cells = [cell_str(v) for v in row]
        # trim to used
        last = 0
        for j, v in enumerate(cells, start=1):
            if v is not None:
                last = j
        if last == 0:
            consecutive_wide = 0
            continue

        a = cells[0] if cells else None
        b = cells[1] if len(cells) > 1 else None
        non_empty_count = sum(1 for v in cells[:last] if v is not None)

        if a:
            all_col_a.append({"row": i, "a": a, "b": b, "non_empty": non_empty_count})
            col_a_labels.append(a)

        # Detect table header: row with several year-like cells in columns
        yearish = sum(1 for v in cells[: min(last, 40)] if looks_like_year(v))
        # also treat headers with multiple filled cols and A looking like metric name while B/C are years
        if yearish >= 2 and non_empty_count >= 3:
            headers = [v for v in cells[: min(last, 40)] if v is not None]
            table_header_rows.append(
                {
                    "row": i,
                    "yearish_count": yearish,
                    "non_empty": non_empty_count,
                    "headers_preview": headers[:35],
                    "col_a": a,
                }
            )
            if first_table_row is None:
                first_table_row = i
            year_col_hits += 1

        # Vertical KV: typically 1-2 non-empty cells early, A=label B=value
        if first_table_row is None or i < first_table_row:
            if a and non_empty_count <= 3:
                meta_kv.append(
                    {
                        "row": i,
                        "label": a,
                        "value": b,
                        "extra": [v for v in cells[2:last] if v is not None] if last > 2 else [],
                        "kind": "scalar_candidate",
                    }
                )
            elif a and non_empty_count > 3 and yearish < 2:
                # wide non-year row before table — capture as possible header
                meta_kv.append(
                    {
                        "row": i,
                        "label": a,
                        "value": b,
                        "extra": [v for v in cells[2 : min(last, 20)] if v is not None],
                        "kind": "wide_pre_table",
                    }
                )

        # Don't need every col A after first 5000 for catalog — but we do for proprietary sheet
        # Continue full scan for labels; may be long

    # Unique labels
    unique_a = []
    seen = set()
    for lab in col_a_labels:
        if lab not in seen:
            seen.add(lab)
            unique_a.append(lab)

    # Classify: if label appears in a row that has year columns among neighbors, time series
    # Build set of rows that are under a year-header block
    ts_labels: list[str] = []
    scalar_labels: list[str] = []

    header_rows_set = {t["row"] for t in table_header_rows}
    # After each header row, rows with A set and numeric/other in year cols are series
    # Simpler: labels appearing only in meta_kv with <=2 cols = scalar; labels in all_col_a after first_table_row = series candidates
    meta_labels = {m["label"] for m in meta_kv if m.get("kind") == "scalar_candidate"}
    for lab in unique_a:
        if lab in meta_labels and (first_table_row is None or True):
            # check if also appears after table
            pass

    for entry in all_col_a:
        lab = entry["a"]
        if first_table_row and entry["row"] >= first_table_row and entry["non_empty"] >= 3:
            if lab not in ts_labels:
                ts_labels.append(lab)
        elif entry["non_empty"] <= 3:
            if lab not in scalar_labels and lab not in ts_labels:
                scalar_labels.append(lab)

    # labels only in ts
    scalar_only = [l for l in scalar_labels if l not in ts_labels]
    ts_only = [l for l in ts_labels if l not in scalar_only]

    return {
        "sheet": ws.title,
        "approx_rows_with_col_a": len(all_col_a),
        "unique_col_a_count": len(unique_a),
        "unique_col_a_labels": unique_a,
        "meta_kv_section": meta_kv,
        "first_table_header_row": first_table_row,
        "table_header_rows": table_header_rows[:20],
        "scalar_candidate_labels": scalar_only,
        "time_series_candidate_labels": ts_only,
        "col_a_first_80": all_col_a[:80],
        "col_a_sample_around_table": [
            e for e in all_col_a if first_table_row and first_table_row - 5 <= e["row"] <= first_table_row + 40
        ],
    }


def find_crf(ticker: str) -> Path | None:
    d = ROOT / ticker
    for p in d.glob("Custom_Run_Filter*.xlsx"):
        if not p.name.startswith("~$"):
            return p
    return None


def find_template(ticker: str) -> Path | None:
    d = ROOT / ticker
    for p in d.iterdir():
        if p.name.startswith("~$"):
            continue
        if "Industrial Template" in p.name or "prefilled" in p.name.lower() or "Template" in p.name:
            if p.suffix.lower() == ".xlsx":
                return p
    # also workbook.xlsx
    for p in d.glob("*.xlsx"):
        if p.name.startswith("~$"):
            continue
        if "Custom_Run" not in p.name:
            return p
    return None


def inspect_crf(ticker: str) -> dict:
    path = find_crf(ticker)
    if not path:
        return {"ticker": ticker, "error": "no Custom_Run_Filter"}
    print(f"Inspecting {ticker}: {path.name}", flush=True)
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    sheets_info = []
    summary_catalog = None
    ticker_catalog = None
    proprietary_sheet = None
    summary_sheet = None

    for name in sheet_names:
        ws = wb[name]
        print(f"  sheet: {name}", flush=True)
        preview = inspect_sheet_preview(ws)
        sheets_info.append(preview)
        if name.lower() == "summary" or "summary" in name.lower():
            summary_sheet = name
            # need fresh iter — reopen sheet via new workbook pass later
        if preview["approx_max_row"] >= 4000:
            proprietary_sheet = name
        # ticker-named sheet
        if name.upper() == ticker.upper() or name.upper().startswith(ticker.upper()):
            proprietary_sheet = proprietary_sheet or name

    wb.close()

    # Second pass for catalogs (fresh iterators)
    wb = load_workbook(path, read_only=True, data_only=True)
    for name in sheet_names:
        ws = wb[name]
        if name.lower() == "summary" or name == summary_sheet:
            print(f"  catalog Summary: {name}", flush=True)
            summary_catalog = extract_summary_headers(ws)
            summary_sheet = name
        if name.upper() == ticker.upper() or (
            proprietary_sheet and name == proprietary_sheet and name.lower() != "summary"
        ):
            print(f"  catalog Ticker sheet: {name}", flush=True)
            ticker_catalog = extract_ticker_labels(ws, ticker)
            if name.upper() == ticker.upper():
                proprietary_sheet = name
    # If we haven't identified proprietary, pick max rows
    if not proprietary_sheet:
        best = max(sheets_info, key=lambda s: s["approx_max_row"])
        proprietary_sheet = best["title"]
        if ticker_catalog is None and proprietary_sheet.lower() != "summary":
            ticker_catalog = extract_ticker_labels(wb[proprietary_sheet], ticker)

    wb.close()

    return {
        "ticker": ticker,
        "path": str(path),
        "filename": path.name,
        "sheet_names": sheet_names,
        "sheets": sheets_info,
        "summary_sheet": summary_sheet,
        "proprietary_metric_sheet": proprietary_sheet,
        "summary_catalog": summary_catalog,
        "ticker_sheet_catalog": ticker_catalog,
    }


def inspect_template_aapl() -> dict:
    path = find_template("AAPL")
    if not path:
        return {"error": "no template"}
    print(f"Inspecting AAPL template: {path.name}", flush=True)
    wb = load_workbook(path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    first20 = names[:20]
    # heuristic for IS/BS
    keywords = ("income", "balance", "cash flow", "cashflow", "p&l", "pnl", "statement")
    statement_like = [n for n in names if any(k in n.lower() for k in keywords)]
    return {
        "path": str(path),
        "filename": path.name,
        "sheet_count": len(names),
        "first_20_sheet_names": first20,
        "all_sheet_names": names,
        "statement_like_sheets": statement_like,
        "has_income_style": any("income" in n.lower() for n in names),
        "has_balance_style": any("balance" in n.lower() for n in names),
        "has_cashflow_style": any("cash" in n.lower() and "flow" in n.lower() for n in names),
    }


def compare(inventories: list[dict]) -> dict:
    sheet_sets = {inv["ticker"]: inv.get("sheet_names", []) for inv in inventories}
    summary_headers = {
        inv["ticker"]: (inv.get("summary_catalog") or {}).get("headers", []) for inv in inventories
    }
    # compare header equality
    base = summary_headers.get("AAPL", [])
    header_eq = {t: (hs == base) for t, hs in summary_headers.items()}
    # set diffs vs AAPL
    base_set = set(base)
    header_diffs = {}
    for t, hs in summary_headers.items():
        s = set(hs)
        header_diffs[t] = {
            "missing_vs_aapl": sorted(base_set - s),
            "extra_vs_aapl": sorted(s - base_set),
            "count": len(hs),
        }

    prop_sheets = {inv["ticker"]: inv.get("proprietary_metric_sheet") for inv in inventories}
    prop_rows = {
        inv["ticker"]: next(
            (s["approx_max_row"] for s in inv.get("sheets", []) if s["title"] == inv.get("proprietary_metric_sheet")),
            None,
        )
        for inv in inventories
    }

    ticker_label_sets = {
        inv["ticker"]: set((inv.get("ticker_sheet_catalog") or {}).get("unique_col_a_labels", []))
        for inv in inventories
    }
    aapl_labels = ticker_label_sets.get("AAPL", set())
    label_diffs = {}
    for t, labs in ticker_label_sets.items():
        label_diffs[t] = {
            "unique_count": len(labs),
            "missing_vs_aapl_sample": sorted(aapl_labels - labs)[:50],
            "extra_vs_aapl_sample": sorted(labs - aapl_labels)[:50],
            "jaccard_vs_aapl": round(len(aapl_labels & labs) / len(aapl_labels | labs), 4) if (aapl_labels | labs) else None,
        }

    same_sheets = len({tuple(v) for v in sheet_sets.values()}) == 1

    return {
        "same_sheet_names_across_all": same_sheets,
        "sheet_names_by_ticker": sheet_sets,
        "summary_headers_equal_to_aapl": header_eq,
        "summary_header_diffs_vs_aapl": header_diffs,
        "proprietary_sheet_by_ticker": prop_sheets,
        "proprietary_approx_rows": prop_rows,
        "ticker_label_diffs_vs_aapl": label_diffs,
    }


def main():
    folder_listings = {t: list_folder(t) for t in TICKERS}
    inventories = []
    for t in TICKERS:
        inventories.append(inspect_crf(t))
    template = inspect_template_aapl()
    comparison = compare(inventories)

    payload = {
        "folder_listings": folder_listings,
        "custom_run_filters": inventories,
        "aapl_industrial_template": template,
        "cross_company_comparison": comparison,
    }
    OUT.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {OUT}", flush=True)

    # Also print human-readable digest
    digest = Path(r"C:\Users\G\HAP\validation_campaign\_crf_inventory_digest.md")
    lines = ["# Custom_Run_Filter Inventory Digest", ""]
    for t in TICKERS:
        lines.append(f"## {t} folder")
        for f in folder_listings[t]:
            lines.append(f"- `{f['name']}` ({f['bytes']:,} bytes)")
        lines.append("")
    for inv in inventories:
        t = inv["ticker"]
        lines.append(f"## {t} Custom_Run_Filter")
        lines.append(f"- File: `{inv.get('filename')}`")
        lines.append(f"- Sheets: {inv.get('sheet_names')}")
        lines.append(f"- Summary sheet: {inv.get('summary_sheet')}")
        lines.append(f"- Proprietary/metric sheet: {inv.get('proprietary_metric_sheet')}")
        for s in inv.get("sheets", []):
            lines.append(
                f"  - `{s['title']}`: max_row≈{s['approx_max_row']}, max_col≈{s['approx_max_column']}, scanned={s['rows_scanned']}"
            )
        sc = inv.get("summary_catalog") or {}
        lines.append(f"- Summary headers ({sc.get('header_count')}):")
        for h in sc.get("headers") or []:
            lines.append(f"  - {h}")
        tc = inv.get("ticker_sheet_catalog") or {}
        lines.append(f"- Ticker sheet unique col-A labels: {tc.get('unique_col_a_count')}")
        lines.append(f"- First table header row: {tc.get('first_table_header_row')}")
        lines.append(f"- Scalar candidates: {len(tc.get('scalar_candidate_labels') or [])}")
        lines.append(f"- Time-series candidates: {len(tc.get('time_series_candidate_labels') or [])}")
        if t == "AAPL":
            lines.append("### AAPL meta_kv_section")
            for m in tc.get("meta_kv_section") or []:
                lines.append(f"  - r{m['row']}: {m['label']!r} = {m.get('value')!r} ({m.get('kind')})")
            lines.append("### AAPL all unique col-A labels")
            for lab in tc.get("unique_col_a_labels") or []:
                lines.append(f"  - {lab}")
            lines.append("### AAPL table header rows")
            for th in tc.get("table_header_rows") or []:
                lines.append(f"  - r{th['row']}: {th.get('headers_preview')}")
            lines.append("### AAPL first_15 Summary + proprietary previews")
            for s in inv.get("sheets", []):
                lines.append(f"#### Sheet `{s['title']}` first 15")
                for i, row in enumerate(s.get("first_15_rows") or [], start=1):
                    lines.append(f"  {i}: {row}")
        lines.append("")

    lines.append("## AAPL Industrial Template")
    lines.append(f"- File: `{template.get('filename')}`")
    lines.append(f"- Sheet count: {template.get('sheet_count')}")
    lines.append(f"- First 20: {template.get('first_20_sheet_names')}")
    lines.append(f"- Statement-like: {template.get('statement_like_sheets')}")
    lines.append(f"- Income: {template.get('has_income_style')}, Balance: {template.get('has_balance_style')}, CF: {template.get('has_cashflow_style')}")
    lines.append("")
    lines.append("## Cross-company comparison")
    lines.append("```json")
    lines.append(json.dumps(comparison, indent=2))
    lines.append("```")
    digest.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {digest}", flush=True)


if __name__ == "__main__":
    main()
