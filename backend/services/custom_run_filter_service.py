"""Parse custom_run filter CSV/XLSX mapping files."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from models.custom_run_filter import CustomRunFilterParseResult, FilterMapping

REQUIRED_COLUMNS = {"sheet", "cell", "concept"}


class CustomRunFilterError(Exception):
    """Raised when the custom_run filter file cannot be parsed."""


class CustomRunFilterService:
    """Read analyst-provided fill mappings without recomputing proprietary metrics."""

    def parse(self, filter_path: Path, original_filename: str) -> CustomRunFilterParseResult:
        suffix = filter_path.suffix.lower()
        if suffix == ".csv":
            rows = self._read_csv(filter_path)
        elif suffix in {".xlsx", ".xlsm", ".xls"}:
            rows = self._read_xlsx(filter_path)
        else:
            raise CustomRunFilterError(
                f"Unsupported custom_run filter format: {suffix}. Use .csv or .xlsx."
            )

        mappings: list[FilterMapping] = []
        ticker_override: str | None = None
        company_override: str | None = None

        for row in rows:
            sheet = str(row.get("sheet", "")).strip()
            cell = str(row.get("cell", "")).strip().upper()
            concept = str(row.get("concept", "")).strip()
            if not sheet or not cell or not concept:
                continue

            concept_key = concept.lower().replace(" ", "")
            if concept_key == "ticker":
                ticker_override = str(row.get("value", row.get("period", ""))).strip().upper() or None
                continue
            if concept_key == "company":
                company_override = str(row.get("value", row.get("period", ""))).strip() or None
                continue

            mappings.append(
                FilterMapping(
                    sheet=sheet,
                    cell=cell,
                    concept=concept,
                    period=str(row.get("period", "latest_annual")).strip() or "latest_annual",
                    source=str(row.get("source", "sec_xbrl")).strip() or "sec_xbrl",
                )
            )

        if not mappings:
            raise CustomRunFilterError(
                "custom_run filter contains no fill mappings. Expected columns: sheet, cell, concept."
            )

        return CustomRunFilterParseResult(
            filename=original_filename,
            mappings=mappings,
            ticker_override=ticker_override,
            company_override=company_override,
        )

    def _read_csv(self, path: Path) -> list[dict[str, str]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames is None:
                raise CustomRunFilterError("custom_run filter CSV is missing a header row.")
            headers = {name.strip().lower() for name in reader.fieldnames}
            if not REQUIRED_COLUMNS.issubset(headers):
                raise CustomRunFilterError(
                    "custom_run filter CSV must include sheet, cell, and concept columns."
                )
            return [
                {key.strip().lower(): (value or "").strip() for key, value in row.items() if key}
                for row in reader
            ]

    def _read_xlsx(self, path: Path) -> list[dict[str, str]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not rows:
            raise CustomRunFilterError("custom_run filter workbook is empty.")

        headers = [str(value).strip().lower() for value in rows[0]]
        if not REQUIRED_COLUMNS.issubset(set(headers)):
            raise CustomRunFilterError(
                "custom_run filter workbook must include sheet, cell, and concept columns."
            )

        parsed: list[dict[str, str]] = []
        for raw in rows[1:]:
            if not raw or all(value in (None, "") for value in raw):
                continue
            row = {
                headers[index]: "" if value is None else str(value).strip()
                for index, value in enumerate(raw)
                if index < len(headers)
            }
            parsed.append(row)
        return parsed
