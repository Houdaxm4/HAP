"""Author the institutional HAP Excel workbook (hap_workbook.xlsx).

Creates all 17 locked sheets every run. Populates statement / market foundations
first; analytical and summary sheets are filled from the engine when available.
Never invents financial values. Does not copy the Industrial Template.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analysis_engine.schemas import AnalysisEngineResult
from canonical_model import CompanyFinancialModel
from canonical_model.primitives import FinancialSeries
from models.analysis import Analysis
from models.common import utc_now_iso
from models.custom_run import CustomRunData
from models.provenance import ProvenanceReport
from models.validation import DiscrepancyReport

# Locked sheet order — do not reorder without a product decision.
SHEET_ORDER: tuple[str, ...] = (
    "Cover",
    "Executive Summary",
    "Company Overview",
    "Income Statement",
    "Balance Sheet",
    "Cash Flow Statement",
    "Market Data",
    "Financial Ratios",
    "HAP Metrics",
    "Business Quality",
    "Investment Attractiveness",
    "Valuation",
    "Expected Return",
    "Recommendation",
    "Assumptions",
    "Validation & Provenance",
    "Run Log",
)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="EA580C")
SECTION_FONT = Font(bold=True, size=11)
NOTE_FONT = Font(italic=True, size=9, color="6B7280")
COMING_SOON = "Coming Soon — structure reserved; data pending enrichment"


class HapWorkbookService:
    """Build ``hap_workbook.xlsx`` from CFM + engine + validation artifacts."""

    def write(
        self,
        path: Path,
        *,
        analysis: Analysis,
        model: CompanyFinancialModel,
        engine_result: AnalysisEngineResult | None = None,
        custom_run: CustomRunData | None = None,
        validation_report: DiscrepancyReport | None = None,
        provenance_report: ProvenanceReport | None = None,
        engine_version: str = "0.3.0",
    ) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # Remove default sheet; recreate in locked order.
        default = wb.active
        wb.remove(default)
        sheets = {name: wb.create_sheet(name) for name in SHEET_ORDER}

        annual_periods, quarter_period = self._split_periods(model.periods)

        # Build order: statements → market → analytics → summary/cover last.
        self._write_income_statement(sheets["Income Statement"], model, annual_periods, quarter_period)
        self._write_balance_sheet(sheets["Balance Sheet"], model, annual_periods, quarter_period)
        self._write_cash_flow(sheets["Cash Flow Statement"], model, annual_periods, quarter_period)
        self._write_market_data(sheets["Market Data"], model)
        self._write_financial_ratios(sheets["Financial Ratios"], engine_result)
        self._write_hap_metrics(sheets["HAP Metrics"], engine_result, custom_run)
        self._write_valuation(sheets["Valuation"], engine_result)
        self._write_recommendation(sheets["Recommendation"], engine_result)
        self._write_business_quality(sheets["Business Quality"], engine_result)
        self._write_investment_attractiveness(sheets["Investment Attractiveness"], engine_result)
        self._write_expected_return(sheets["Expected Return"], engine_result)
        self._write_assumptions(sheets["Assumptions"], model, custom_run)
        self._write_company_overview(sheets["Company Overview"], analysis, model, custom_run)
        self._write_validation_provenance(
            sheets["Validation & Provenance"],
            validation_report,
            provenance_report,
        )
        self._write_run_log(sheets["Run Log"], analysis)
        self._write_executive_summary(sheets["Executive Summary"], analysis, model, engine_result)
        self._write_cover(sheets["Cover"], analysis, model, engine_result, engine_version)

        wb.save(path)
        return path

    # ------------------------------------------------------------------ periods

    @staticmethod
    def _split_periods(periods: list[str]) -> tuple[list[str], str | None]:
        annual: list[str] = []
        quarters: list[str] = []
        for period in periods:
            label = str(period)
            upper = label.upper()
            if "Q" in upper and not upper.startswith("FY"):
                quarters.append(label)
            elif upper.startswith("Q") and any(ch.isdigit() for ch in upper):
                quarters.append(label)
            else:
                annual.append(label)
        annual = annual[-10:]  # up to 10 years, oldest→newest already sorted
        latest_quarter = quarters[-1] if quarters else None
        return annual, latest_quarter

    # ------------------------------------------------------------------ helpers

    def _style_header_row(self, ws: Any, row: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def _autosize(self, ws: Any, min_width: int = 12, max_width: int = 42) -> None:
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            length = 0
            for cell in col_cells:
                if cell.value is None:
                    continue
                length = max(length, min(len(str(cell.value)), max_width))
            ws.column_dimensions[letter].width = max(min_width, length + 2)

    def _title(self, ws: Any, text: str) -> None:
        ws["A1"] = text
        ws["A1"].font = TITLE_FONT

    def _coming_soon_note(self, ws: Any, row: int = 3) -> None:
        ws.cell(row=row, column=1, value=COMING_SOON).font = NOTE_FONT

    def _series_value(self, series: FinancialSeries, period: str) -> float | None:
        return series.value_for(period)

    def _write_statement_sheet(
        self,
        ws: Any,
        title: str,
        currency: str,
        lines: list[tuple[str, FinancialSeries]],
        annual_periods: list[str],
        quarter_period: str | None,
    ) -> None:
        self._title(ws, title)
        ws["A2"] = f"Currency: {currency}"
        ws["A2"].font = NOTE_FONT

        headers = ["Line Item", *annual_periods]
        # Reserve empty columns so the sheet always has 10 annual slots.
        for _ in range(max(0, 10 - len(annual_periods))):
            headers.append("")
        headers.append(quarter_period or "Latest Quarter")
        headers.append("Notes")

        header_row = 4
        for col, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col, value=header if header else None)
        self._style_header_row(ws, header_row, len(headers))

        for row_offset, (label, series) in enumerate(lines):
            row = header_row + 1 + row_offset
            ws.cell(row=row, column=1, value=label)
            col = 2
            missing = 0
            for period in annual_periods:
                value = self._series_value(series, period)
                ws.cell(row=row, column=col, value=value)
                if value is None:
                    missing += 1
                col += 1
            for _ in range(max(0, 10 - len(annual_periods))):
                col += 1  # reserved empty FY columns
            if quarter_period:
                ws.cell(row=row, column=col, value=self._series_value(series, quarter_period))
            col += 1
            note = ""
            if missing == len(annual_periods) and not series.points:
                note = "No data"
            elif missing:
                note = f"{missing} period(s) blank"
            ws.cell(row=row, column=col, value=note or None)

        ws.freeze_panes = "B5"
        self._autosize(ws)

    # ------------------------------------------------------------------ sheets

    def _write_income_statement(
        self,
        ws: Any,
        model: CompanyFinancialModel,
        annual: list[str],
        quarter: str | None,
    ) -> None:
        is_ = model.income_statement
        lines = [
            ("Revenue", is_.revenue),
            ("Cost of Revenue", is_.cost_of_revenue),
            ("Gross Profit", is_.gross_profit),
            ("Operating Income", is_.operating_income),
            ("EBIT", is_.ebit),
            ("EBITDA", is_.ebitda),
            ("Interest Expense", is_.interest_expense),
            ("Tax Expense", is_.tax_expense),
            ("Net Income", is_.net_income),
            ("Diluted EPS", is_.diluted_eps),
        ]
        self._write_statement_sheet(
            ws,
            "Income Statement — up to 10Y + Latest Quarter",
            is_.currency or model.reporting_currency,
            lines,
            annual,
            quarter,
        )

    def _write_balance_sheet(
        self,
        ws: Any,
        model: CompanyFinancialModel,
        annual: list[str],
        quarter: str | None,
    ) -> None:
        bs = model.balance_sheet
        lines = [
            ("Cash", bs.cash),
            ("Current Assets", bs.current_assets),
            ("Total Assets", bs.total_assets),
            ("Current Liabilities", bs.current_liabilities),
            ("Total Liabilities", bs.total_liabilities),
            ("Total Debt", bs.total_debt),
            ("Shareholders Equity", bs.shareholders_equity),
            ("Invested Capital", bs.invested_capital),
        ]
        self._write_statement_sheet(
            ws,
            "Balance Sheet — up to 10Y + Latest Quarter",
            bs.currency or model.reporting_currency,
            lines,
            annual,
            quarter,
        )

    def _write_cash_flow(
        self,
        ws: Any,
        model: CompanyFinancialModel,
        annual: list[str],
        quarter: str | None,
    ) -> None:
        cf = model.cash_flow_statement
        lines = [
            ("Operating Cash Flow", cf.operating_cash_flow),
            ("Capital Expenditures", cf.capital_expenditures),
            ("Free Cash Flow", cf.free_cash_flow),
            ("Dividends", cf.dividends),
            ("Share Repurchases", cf.share_repurchases),
            ("Investing Cash Flow", cf.investing_cash_flow),
            ("Financing Cash Flow", cf.financing_cash_flow),
        ]
        self._write_statement_sheet(
            ws,
            "Cash Flow Statement — up to 10Y + Latest Quarter",
            cf.currency or model.reporting_currency,
            lines,
            annual,
            quarter,
        )

    def _write_market_data(self, ws: Any, model: CompanyFinancialModel) -> None:
        self._title(ws, "Market Data")
        md = model.market_data
        rows = [
            ("Currency", md.currency),
            ("As of Date", md.as_of_date),
            ("Share Price", md.share_price),
            ("Shares Outstanding", md.shares_outstanding),
            ("Market Cap", md.market_cap),
            ("Enterprise Value", md.enterprise_value),
            ("Beta", md.beta),
            ("Dividend Yield", md.dividend_yield),
        ]
        ws["A3"] = "Field"
        ws["B3"] = "Value"
        self._style_header_row(ws, 3, 2)
        for idx, (label, value) in enumerate(rows):
            ws.cell(row=4 + idx, column=1, value=label)
            ws.cell(row=4 + idx, column=2, value=value)
        self._autosize(ws)

    def _write_financial_ratios(self, ws: Any, engine: AnalysisEngineResult | None) -> None:
        self._title(ws, "Financial Ratios")
        ws["A3"] = "Code"
        ws["B3"] = "Name"
        ws["C3"] = "Value"
        ws["D3"] = "Unit"
        ws["E3"] = "Period"
        self._style_header_row(ws, 3, 5)
        metrics = list(engine.metrics) if engine and engine.metrics else []
        ratio_like = [
            m
            for m in metrics
            if any(
                token in (m.code or "").upper()
                for token in ("RATIO", "MARGIN", "ROE", "ROA", "ROIC", "YIELD", "COVERAGE", "GROWTH", "CAGR")
            )
        ] or metrics[:40]
        if not ratio_like:
            self._coming_soon_note(ws, 4)
            self._autosize(ws)
            return
        for idx, metric in enumerate(ratio_like):
            row = 4 + idx
            ws.cell(row=row, column=1, value=metric.code)
            ws.cell(row=row, column=2, value=metric.name or metric.code)
            ws.cell(row=row, column=3, value=metric.value)
            ws.cell(row=row, column=4, value=metric.unit)
            ws.cell(row=row, column=5, value=metric.period)
        self._autosize(ws)

    def _write_hap_metrics(
        self,
        ws: Any,
        engine: AnalysisEngineResult | None,
        custom_run: CustomRunData | None,
    ) -> None:
        self._title(ws, "HAP Metrics")
        ws["A3"] = "Source"
        ws["B3"] = "Metric"
        ws["C3"] = "Value"
        ws["D3"] = "Notes"
        self._style_header_row(ws, 3, 4)
        row = 4
        if engine and engine.metrics:
            for metric in engine.metrics:
                ws.cell(row=row, column=1, value="HAP Engine")
                ws.cell(row=row, column=2, value=metric.name or metric.code)
                ws.cell(row=row, column=3, value=metric.value)
                ws.cell(row=row, column=4, value=metric.period)
                row += 1
        if custom_run is not None:
            for key, value in sorted(custom_run.valuation_metrics.items()):
                ws.cell(row=row, column=1, value="Custom Run (imported)")
                ws.cell(row=row, column=2, value=key)
                ws.cell(row=row, column=3, value=value)
                ws.cell(row=row, column=4, value="Imported — not recomputed")
                row += 1
            for key, value in sorted(custom_run.quality_metrics.items()):
                ws.cell(row=row, column=1, value="Custom Run (imported)")
                ws.cell(row=row, column=2, value=key)
                ws.cell(row=row, column=3, value=value)
                ws.cell(row=row, column=4, value="Imported — not recomputed")
                row += 1
        if row == 4:
            self._coming_soon_note(ws, 4)
        self._autosize(ws)

    def _write_module_score_sheet(
        self,
        ws: Any,
        title: str,
        module_name: str,
        engine: AnalysisEngineResult | None,
    ) -> None:
        self._title(ws, title)
        module = None
        if engine is not None:
            for item in engine.modules:
                if item.module_name == module_name:
                    module = item
                    break
        if module is None:
            self._coming_soon_note(ws)
            self._autosize(ws)
            return
        ws["A3"] = "Field"
        ws["B3"] = "Value"
        self._style_header_row(ws, 3, 2)
        rows = [
            ("Module", module.module_name),
            ("Status", module.status),
            ("Score", module.score),
            ("Confidence", module.confidence),
            ("Error", module.error),
        ]
        for idx, (label, value) in enumerate(rows):
            ws.cell(row=4 + idx, column=1, value=label)
            ws.cell(row=4 + idx, column=2, value=value)
        row = 10
        ws.cell(row=row, column=1, value="Findings").font = SECTION_FONT
        row += 1
        for finding in module.findings or []:
            ws.cell(row=row, column=1, value=finding.code)
            ws.cell(row=row, column=2, value=finding.summary)
            row += 1
        if module.status == "skipped":
            ws.cell(row=row + 1, column=1, value=COMING_SOON).font = NOTE_FONT
        self._autosize(ws)

    def _write_valuation(self, ws: Any, engine: AnalysisEngineResult | None) -> None:
        self._write_module_score_sheet(ws, "Valuation", "valuation", engine)

    def _write_expected_return(self, ws: Any, engine: AnalysisEngineResult | None) -> None:
        self._write_module_score_sheet(ws, "Expected Return", "expected_return", engine)

    def _write_aggregator_sheet(
        self,
        ws: Any,
        title: str,
        aggregator: Any | None,
    ) -> None:
        self._title(ws, title)
        if aggregator is None:
            self._coming_soon_note(ws)
            self._autosize(ws)
            return
        ws["A3"] = "Field"
        ws["B3"] = "Value"
        self._style_header_row(ws, 3, 2)
        rows = [
            ("Score", aggregator.score),
            ("Confidence", aggregator.confidence),
            ("Classification", aggregator.classification),
            ("Classification Label", aggregator.classification_label),
        ]
        for idx, (label, value) in enumerate(rows):
            ws.cell(row=4 + idx, column=1, value=label)
            ws.cell(row=4 + idx, column=2, value=value)
        row = 10
        ws.cell(row=row, column=1, value="Module Contributions").font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value="Module")
        ws.cell(row=row, column=2, value="Weight")
        ws.cell(row=row, column=3, value="Score")
        ws.cell(row=row, column=4, value="Status")
        self._style_header_row(ws, row, 4)
        row += 1
        for contrib in aggregator.module_contributions or []:
            ws.cell(row=row, column=1, value=contrib.module_name)
            ws.cell(row=row, column=2, value=contrib.weight)
            ws.cell(row=row, column=3, value=contrib.score)
            ws.cell(row=row, column=4, value=contrib.status)
            row += 1
        self._autosize(ws)

    def _write_business_quality(self, ws: Any, engine: AnalysisEngineResult | None) -> None:
        self._write_aggregator_sheet(
            ws,
            "Business Quality",
            engine.business_quality if engine else None,
        )

    def _write_investment_attractiveness(self, ws: Any, engine: AnalysisEngineResult | None) -> None:
        self._write_aggregator_sheet(
            ws,
            "Investment Attractiveness",
            engine.investment_attractiveness if engine else None,
        )

    def _write_recommendation(self, ws: Any, engine: AnalysisEngineResult | None) -> None:
        self._title(ws, "Recommendation")
        rec = engine.recommendation if engine else None
        if rec is None:
            self._coming_soon_note(ws)
            self._autosize(ws)
            return
        ws["A3"] = "Field"
        ws["B3"] = "Value"
        self._style_header_row(ws, 3, 2)
        rows = [
            ("Recommendation", rec.recommendation),
            ("Label", rec.recommendation_label),
            ("Confidence", rec.confidence),
            ("Business Quality Score", rec.business_quality_score),
            ("Business Quality Class", rec.business_quality_classification),
            ("Investment Attractiveness Score", rec.investment_attractiveness_score),
            ("Investment Attractiveness Class", rec.investment_attractiveness_classification),
        ]
        for idx, (label, value) in enumerate(rows):
            ws.cell(row=4 + idx, column=1, value=label)
            ws.cell(row=4 + idx, column=2, value=value)
        row = 13
        ws.cell(row=row, column=1, value="Reasons").font = SECTION_FONT
        row += 1
        for reason in rec.reasons or []:
            ws.cell(row=row, column=1, value=reason.code)
            ws.cell(row=row, column=2, value=reason.summary)
            row += 1
        self._autosize(ws)

    def _write_assumptions(
        self,
        ws: Any,
        model: CompanyFinancialModel,
        custom_run: CustomRunData | None,
    ) -> None:
        self._title(ws, "Assumptions")
        vi = model.valuation_inputs
        ws["A3"] = "Field"
        ws["B3"] = "Value"
        ws["C3"] = "Source"
        self._style_header_row(ws, 3, 3)
        rows = [
            ("Risk Free Rate", vi.risk_free_rate, "valuation_inputs"),
            ("Equity Risk Premium", vi.equity_risk_premium, "valuation_inputs"),
            ("Cost of Equity", vi.cost_of_equity, "valuation_inputs"),
            ("Cost of Debt", vi.cost_of_debt, "valuation_inputs"),
            ("Tax Rate", vi.tax_rate, "valuation_inputs"),
            ("WACC", vi.wacc, "valuation_inputs"),
            ("Terminal Growth Rate", vi.terminal_growth_rate, "valuation_inputs"),
            ("Forecast Years", vi.forecast_years, "valuation_inputs"),
            ("Net Debt", vi.net_debt, "valuation_inputs"),
        ]
        for idx, (label, value, source) in enumerate(rows):
            ws.cell(row=4 + idx, column=1, value=label)
            ws.cell(row=4 + idx, column=2, value=value)
            ws.cell(row=4 + idx, column=3, value=source)
        row = 4 + len(rows) + 1
        if custom_run and custom_run.assumptions:
            ws.cell(row=row, column=1, value="Custom Run Assumptions").font = SECTION_FONT
            row += 1
            for key, value in custom_run.assumptions.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=3, value="custom_run.assumptions")
                row += 1
        if model.metadata.get("terminal_growth_defaulted"):
            ws.cell(row=row + 1, column=1, value="Note: terminal growth defaulted to 3%").font = NOTE_FONT
        self._autosize(ws)

    def _write_company_overview(
        self,
        ws: Any,
        analysis: Analysis,
        model: CompanyFinancialModel,
        custom_run: CustomRunData | None,
    ) -> None:
        self._title(ws, "Company Overview")
        ws["A3"] = "Field"
        ws["B3"] = "Value"
        self._style_header_row(ws, 3, 2)
        meta = (custom_run.metadata if custom_run else {}) or {}
        summary = (custom_run.summary if custom_run else {}) or {}
        rows = [
            ("Company", analysis.company or model.company),
            ("Ticker", analysis.ticker),
            ("CIK", analysis.cik),
            ("Analysis Type", analysis.analysis_type),
            ("Reporting Currency", model.reporting_currency),
            ("Industry Sector", summary.get("Industry Sector") or meta.get("Industry Sector")),
            ("Industry Subgroup", summary.get("Industry Subgroup") or meta.get("Industry Subgroup")),
            ("Latest Fiscal Quarter", summary.get("Latest Fiscal Quarter") or meta.get("Latest Fiscal Quarter")),
            ("Fiscal Year Closing", meta.get("Fiscal Year Closing")),
        ]
        for idx, (label, value) in enumerate(rows):
            ws.cell(row=4 + idx, column=1, value=label)
            ws.cell(row=4 + idx, column=2, value=value)
        self._autosize(ws)

    def _write_validation_provenance(
        self,
        ws: Any,
        validation_report: DiscrepancyReport | None,
        provenance_report: ProvenanceReport | None,
    ) -> None:
        self._title(ws, "Validation & Provenance")
        row = 3
        ws.cell(row=row, column=1, value="Validation").font = SECTION_FONT
        row += 1
        if validation_report is None:
            ws.cell(row=row, column=1, value=COMING_SOON).font = NOTE_FONT
            row += 2
        else:
            ws.cell(row=row, column=1, value="Pass")
            ws.cell(row=row, column=2, value=validation_report.pass_count)
            row += 1
            ws.cell(row=row, column=1, value="Warn")
            ws.cell(row=row, column=2, value=validation_report.warn_count)
            row += 1
            ws.cell(row=row, column=1, value="Fail")
            ws.cell(row=row, column=2, value=validation_report.fail_count)
            row += 1
            ws.cell(row=row, column=1, value="Summary")
            ws.cell(row=row, column=2, value=validation_report.summary)
            row += 2
            ws.cell(row=row, column=1, value="Concept")
            ws.cell(row=row, column=2, value="Status")
            ws.cell(row=row, column=3, value="Message")
            self._style_header_row(ws, row, 3)
            row += 1
            for check in validation_report.checks:
                ws.cell(row=row, column=1, value=check.concept)
                ws.cell(row=row, column=2, value=check.status)
                ws.cell(row=row, column=3, value=check.message)
                row += 1
            row += 1

        ws.cell(row=row, column=1, value="Provenance").font = SECTION_FONT
        row += 1
        if provenance_report is None or not provenance_report.entries:
            ws.cell(row=row, column=1, value=COMING_SOON).font = NOTE_FONT
        else:
            ws.cell(row=row, column=1, value="Cell Ref")
            ws.cell(row=row, column=2, value="Concept")
            ws.cell(row=row, column=3, value="Period")
            ws.cell(row=row, column=4, value="Value")
            ws.cell(row=row, column=5, value="Source")
            self._style_header_row(ws, row, 5)
            row += 1
            for entry in provenance_report.entries[:500]:
                ws.cell(row=row, column=1, value=entry.cell_ref)
                ws.cell(row=row, column=2, value=entry.concept)
                ws.cell(row=row, column=3, value=entry.period)
                ws.cell(row=row, column=4, value=entry.value)
                ws.cell(row=row, column=5, value=entry.source_document or entry.status)
                row += 1
            if len(provenance_report.entries) > 500:
                ws.cell(
                    row=row,
                    column=1,
                    value=f"Truncated: showing 500 of {len(provenance_report.entries)} entries",
                ).font = NOTE_FONT
        self._autosize(ws)

    def _write_run_log(self, ws: Any, analysis: Analysis) -> None:
        self._title(ws, "Run Log")
        ws["A3"] = "Timestamp"
        ws["B3"] = "Agent"
        ws["C3"] = "Action"
        ws["D3"] = "Detail"
        self._style_header_row(ws, 3, 4)
        if not analysis.decision_log:
            self._coming_soon_note(ws, 4)
            self._autosize(ws)
            return
        for idx, entry in enumerate(analysis.decision_log):
            row = 4 + idx
            ws.cell(row=row, column=1, value=entry.timestamp)
            ws.cell(row=row, column=2, value=entry.agent)
            ws.cell(row=row, column=3, value=entry.action)
            ws.cell(row=row, column=4, value=entry.detail)
        self._autosize(ws)

    def _write_executive_summary(
        self,
        ws: Any,
        analysis: Analysis,
        model: CompanyFinancialModel,
        engine: AnalysisEngineResult | None,
    ) -> None:
        self._title(ws, "Executive Summary")
        rec = engine.recommendation if engine else None
        bq = engine.business_quality if engine else None
        ia = engine.investment_attractiveness if engine else None
        ws["A3"] = "Field"
        ws["B3"] = "Value"
        self._style_header_row(ws, 3, 2)
        rows = [
            ("Company", analysis.company or model.company),
            ("Ticker", analysis.ticker),
            ("Recommendation", rec.recommendation_label if rec else None),
            ("Business Quality", bq.classification_label if bq else None),
            ("Business Quality Score", bq.score if bq else None),
            ("Investment Attractiveness", ia.classification_label if ia else None),
            ("Investment Attractiveness Score", ia.score if ia else None),
            ("Share Price", model.market_data.share_price),
            ("Engine Confidence", engine.confidence if engine else None),
        ]
        for idx, (label, value) in enumerate(rows):
            ws.cell(row=4 + idx, column=1, value=label)
            ws.cell(row=4 + idx, column=2, value=value)
        row = 15
        ws.cell(row=row, column=1, value="Key Reasons").font = SECTION_FONT
        row += 1
        if rec and rec.reasons:
            for reason in rec.reasons[:12]:
                ws.cell(row=row, column=1, value=reason.code)
                ws.cell(row=row, column=2, value=reason.summary)
                row += 1
        else:
            ws.cell(row=row, column=1, value=COMING_SOON).font = NOTE_FONT
        self._autosize(ws)

    def _write_cover(
        self,
        ws: Any,
        analysis: Analysis,
        model: CompanyFinancialModel,
        engine: AnalysisEngineResult | None,
        engine_version: str,
    ) -> None:
        self._title(ws, "HAP Institutional Workbook")
        rec = engine.recommendation if engine else None
        bq = engine.business_quality if engine else None
        ia = engine.investment_attractiveness if engine else None
        ws["A3"] = "Field"
        ws["B3"] = "Value"
        self._style_header_row(ws, 3, 2)
        rows = [
            ("Product", "Houda's Analyst Platform (HAP)"),
            ("Company", analysis.company or model.company),
            ("Ticker", analysis.ticker),
            ("CIK", analysis.cik),
            ("Analysis Type", analysis.analysis_type),
            ("Analysis ID", analysis.analysis_id),
            ("Generated At (UTC)", utc_now_iso()),
            ("Engine Version", engine_version),
            ("Recommendation", rec.recommendation_label if rec else None),
            ("Business Quality Score", bq.score if bq else None),
            ("Business Quality", bq.classification_label if bq else None),
            ("Investment Attractiveness Score", ia.score if ia else None),
            ("Investment Attractiveness", ia.classification_label if ia else None),
            ("Share Price", model.market_data.share_price),
            ("Reporting Currency", model.reporting_currency),
        ]
        for idx, (label, value) in enumerate(rows):
            ws.cell(row=4 + idx, column=1, value=label)
            ws.cell(row=4 + idx, column=2, value=value)
        ws.cell(
            row=22,
            column=1,
            value="Disclaimer: HAP does not invent financial data; missing inputs remain blank.",
        ).font = NOTE_FONT
        ws.cell(
            row=23,
            column=1,
            value="This workbook is authored by HAP. The Industrial Template copy is preserved separately as completed_workbook.xlsx.",
        ).font = NOTE_FONT
        self._autosize(ws)
