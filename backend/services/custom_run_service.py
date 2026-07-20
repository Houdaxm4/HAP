"""Parse Bloomberg-derived Custom_Run_Filter workbooks (HAP v1 product input)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from models.custom_run import CustomRunData, CustomRunPeriods, CustomRunSeries

# Fixed layout of the ticker sheet (identical across Industrial CRF exports).
_META_START = 2
_META_END = 12
_DATE_ROW = 15
_FISCAL_QUARTER_ROW = 16
_SERIES_START = 18
_FISCAL_YEAR_ROW = 146
_SERIES_END = 153
_SCALAR_START = 158
_SCALAR_END = 261
_PE10_HEADER_ROW = 264
_PE10_DATA_START = 265

_MARKET_KEYS = {
    "Current Price (Live Price)",
    "Current Market Capitalization",
    "Current Enterprise Value (not-diluted)",
    "Current Dividend Yield",
    "Current Dividend Rate",
    "Shares Outstanding Diluted Average (MM)",
}

_VALUATION_KEYS = {
    "WACC",
    "Current PE10",
    "Current E10",
    "Current Graham Instrinsic Value",
    "Graham Instrinsic Value in 7 Years",
    "Graham Expected Annualized Return",
    "Expected Return @ Current Price",
    "Expected Return Price Plus Dividends - Given Current Price",
    "Max Current Price to Buy",
    "1st Exit Price",
    "2nd Exit Price",
    "Approximate Residual Earnings Value",
    "High Price in 10 years",
    "Low Price in 10 years",
    "ROC Greenblatt",
    "ROC - WACC",
    "EBIT TTM/EV",
    "Current Max PE10 to Enter (Lowest PE10 or 7PE10)",
}

_QUALITY_KEYS = {
    "Final Score",
    "Franchise Power",
    "Quality_FPFS",
    "P_FS",
    "P_SNOA",
    "P_ROA10",
    "P_ROC10",
    "P_CFOA10",
    "P_MG",
    "P_MS",
    "P_MM",
    "ROA10",
    "ROC10",
    "CFOA",
    "SNOA (Scaled Net Operating Assets)",
    "Profit Margin Growth",
    "Profit Margin Stability",
    "Number of Consecutive Positive Growth Margins",
    "ROCE",
    "ROCE Fiscal Year",
    "ROCE 5 Fiscal Years",
    "ROCE 10 Fiscal Years",
}


class CustomRunParseError(Exception):
    """Raised when the Custom_Run_Filter workbook cannot be parsed."""


class CustomRunService:
    """Load and parse Bloomberg Custom_Run_Filter.xlsx into CustomRunData."""

    def parse(self, file_path: Path, original_filename: str) -> CustomRunData:
        suffix = file_path.suffix.lower()
        if suffix not in {".xlsx", ".xlsm", ".xls"}:
            raise CustomRunParseError(
                f"Unsupported custom_run_filter format '{suffix}'. "
                "HAP v1 expects a Bloomberg-derived Excel workbook (.xlsx)."
            )

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if "Summary" not in workbook.sheetnames:
                raise CustomRunParseError(
                    "Custom_Run_Filter workbook is missing required worksheet 'Summary'."
                )
            ticker_sheet_name = self._resolve_ticker_sheet(workbook.sheetnames)
            ticker_ws = workbook[ticker_sheet_name]
            summary_ws = workbook["Summary"]

            metadata = self._parse_meta_block(ticker_ws)
            periods = self._parse_periods(ticker_ws)
            historical = self._parse_series_block(ticker_ws, periods)
            scalars = self._parse_scalar_block(ticker_ws)
            summary = self._parse_summary(summary_ws)

            ticker = str(
                metadata.get("Ticker")
                or summary.get("Ticker")
                or ticker_sheet_name
            ).strip().upper()
            company = _optional_str(metadata.get("Company") or summary.get("Company"))

            market_data = self._section_from_pools(
                summary, scalars, metadata, keys=_MARKET_KEYS
            )
            # Prefer live price / EV / mkt cap from summary when present.
            for key in list(_MARKET_KEYS):
                if key in summary and summary[key] is not None:
                    market_data[key] = summary[key]

            valuation_metrics = self._section_from_pools(
                summary, scalars, metadata, keys=_VALUATION_KEYS
            )
            quality_metrics = self._section_from_pools(
                summary, scalars, metadata, keys=_QUALITY_KEYS
            )

            proprietary = {
                key: value
                for key, value in {**scalars, **summary}.items()
                if key not in market_data
                and key not in valuation_metrics
                and key not in quality_metrics
                and key not in {"Company", "Ticker"}
            }

            assumptions: dict[str, Any] = {}
            wacc = valuation_metrics.get("WACC")
            if wacc is None:
                wacc = self._latest_series_value(historical, "WACC")
            if wacc is not None:
                assumptions["wacc"] = wacc
                valuation_metrics.setdefault("WACC", wacc)

            return CustomRunData(
                source_filename=original_filename,
                ticker=ticker,
                company=company,
                ticker_sheet_name=ticker_sheet_name,
                metadata=metadata,
                summary=summary,
                market_data=market_data,
                historical_metrics=historical,
                proprietary_metrics=proprietary,
                valuation_metrics=valuation_metrics,
                quality_metrics=quality_metrics,
                assumptions=assumptions,
                scalars=scalars,
                periods=periods,
                period_count=len(periods.fiscal_quarters),
                series_count=len(historical),
                summary_field_count=len(summary),
            )
        finally:
            workbook.close()

    @staticmethod
    def _resolve_ticker_sheet(sheetnames: list[str]) -> str:
        non_summary = [name for name in sheetnames if name != "Summary"]
        if not non_summary:
            raise CustomRunParseError(
                "Custom_Run_Filter workbook has no ticker worksheet (only Summary found)."
            )
        return non_summary[0]

    def _parse_meta_block(self, worksheet: Any) -> dict[str, Any]:
        meta: dict[str, Any] = {}
        for row_idx in range(_META_START, _META_END + 1):
            label = worksheet.cell(row_idx, 1).value
            value = worksheet.cell(row_idx, 2).value
            if label is None or str(label).strip() == "":
                continue
            meta[str(label).strip()] = _normalize_value(value)
        return meta

    def _parse_periods(self, worksheet: Any) -> CustomRunPeriods:
        dates = self._row_values(worksheet, _DATE_ROW)
        quarters = self._row_values(worksheet, _FISCAL_QUARTER_ROW)
        years = self._row_values(worksheet, _FISCAL_YEAR_ROW)
        # Column A is the header label; periods start at column B (index 0 after drop).
        return CustomRunPeriods(
            dates=[_stringify_period(v) for v in dates],
            fiscal_quarters=[_stringify_period(v) for v in quarters],
            fiscal_years=[_stringify_period(v) for v in years],
        )

    def _parse_series_block(
        self,
        worksheet: Any,
        periods: CustomRunPeriods,
    ) -> dict[str, CustomRunSeries]:
        expected = len(periods.fiscal_quarters) or len(periods.dates)
        series: dict[str, CustomRunSeries] = {}
        for row_idx in range(_SERIES_START, _SERIES_END + 1):
            if row_idx == _FISCAL_YEAR_ROW:
                continue
            label = worksheet.cell(row_idx, 1).value
            if label is None or str(label).strip() == "":
                continue
            label_text = str(label).strip()
            if label_text.lower() in {"date", "fiscal quarter", "fiscal year"}:
                continue
            values = self._row_values(worksheet, row_idx)
            if expected and len(values) > expected:
                values = values[:expected]
            while expected and len(values) < expected:
                values.append(None)
            numeric = [_as_float(v) for v in values]
            populated = sum(1 for v in numeric if v is not None)
            if populated == 0:
                continue
            kind = "annual_aligned" if "fiscal year" in label_text.lower() else "quarterly"
            series[label_text] = CustomRunSeries(
                label=label_text,
                values=numeric,
                kind=kind,
            )
        return series

    def _parse_scalar_block(self, worksheet: Any) -> dict[str, Any]:
        scalars: dict[str, Any] = {}
        for row_idx in range(_SCALAR_START, _SCALAR_END + 1):
            label = worksheet.cell(row_idx, 1).value
            value = worksheet.cell(row_idx, 2).value
            if label is None or str(label).strip() == "":
                continue
            # Trailer indices are numeric labels — skip once we hit pure index runs.
            if isinstance(label, (int, float)) and value is not None and row_idx >= _PE10_HEADER_ROW:
                break
            label_text = str(label).strip()
            if label_text.replace(".", "", 1).isdigit():
                continue
            scalars[label_text] = _normalize_value(value)
        return scalars

    def _parse_summary(self, worksheet: Any) -> dict[str, Any]:
        headers = [
            str(cell.value).strip()
            for cell in worksheet[1]
            if cell.value is not None and str(cell.value).strip() != ""
        ]
        if not headers:
            raise CustomRunParseError("Summary worksheet is missing a header row.")
        values_row = list(worksheet[2])
        summary: dict[str, Any] = {}
        for index, header in enumerate(headers):
            raw = values_row[index].value if index < len(values_row) else None
            summary[header] = _normalize_value(raw)
        return summary

    @staticmethod
    def _row_values(worksheet: Any, row_idx: int) -> list[Any]:
        """Return values from column B onward until trailing empties dominate."""
        values: list[Any] = []
        empty_streak = 0
        col = 2
        # CRF period width is ~102; allow headroom.
        while col <= 120 and empty_streak < 5:
            value = worksheet.cell(row_idx, col).value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                empty_streak += 1
                values.append(None)
            else:
                empty_streak = 0
                values.append(value)
            col += 1
        while values and values[-1] is None:
            values.pop()
        return values

    @staticmethod
    def _section_from_pools(
        *pools: dict[str, Any],
        keys: set[str],
    ) -> dict[str, Any]:
        out: dict[str, Any] = {}
        for key in keys:
            for pool in pools:
                if key in pool and pool[key] is not None:
                    out[key] = pool[key]
                    break
        return out

    @staticmethod
    def _latest_series_value(
        historical: dict[str, CustomRunSeries],
        label: str,
    ) -> float | None:
        series = historical.get(label)
        if series is None:
            return None
        for value in reversed(series.values):
            if value is not None:
                return value
        return None


def _normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:  # noqa: BLE001
            return str(value)
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _stringify_period(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:  # noqa: BLE001
            return str(value)
    return str(value).strip()


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
