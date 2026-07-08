"""Workbook parsing and SEC-driven filling."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.custom_run_filter import CustomRunFilterParseResult, FilterMapping
from models.phase1 import FillRecord
from models.sec import SecCompanyProfile
from models.workbook_parse import WorkbookCell, WorkbookParseResult
from services.sec_edgar_service import SecEdgarService

TICKER_PATTERN = re.compile(r"^[A-Z]{1,5}$")


class WorkbookFillError(Exception):
    """Raised when workbook filling fails."""


class WorkbookFillService:
    """Map SEC facts into analyst template cells without overwriting formulas."""

    def __init__(self, sec_service: SecEdgarService | None = None) -> None:
        self.sec_service = sec_service or SecEdgarService()

    def parse_workbook(self, workbook_path: Path, original_filename: str) -> WorkbookParseResult:
        workbook = load_workbook(workbook_path, read_only=False, data_only=False)
        try:
            worksheet_names = list(workbook.sheetnames)
            visible_sheets: list[str] = []
            blank_cells: list[WorkbookCell] = []
            detected_ticker: str | None = None
            detected_company: str | None = None

            for name in worksheet_names:
                sheet = workbook[name]
                if sheet.sheet_state in ("hidden", "veryHidden"):
                    continue
                visible_sheets.append(name)
                blank_cells.extend(self._scan_blank_cells(sheet))
                ticker, company = self._scan_identity_fields(sheet)
                detected_ticker = detected_ticker or ticker
                detected_company = detected_company or company

            return WorkbookParseResult(
                workbook_filename=original_filename,
                worksheet_names=worksheet_names,
                visible_sheets=visible_sheets,
                blank_cells=blank_cells,
                detected_ticker=detected_ticker,
                detected_company=detected_company,
            )
        finally:
            workbook.close()

    def fill_workbook(
        self,
        workbook_path: Path,
        filter_parse: CustomRunFilterParseResult,
        profile: SecCompanyProfile,
        output_path: Path,
    ) -> list[FillRecord]:
        workbook = load_workbook(workbook_path, read_only=False, data_only=False)
        fills: list[FillRecord] = []

        try:
            for mapping in filter_parse.mappings:
                fill = self._apply_mapping(workbook, mapping, profile)
                if fill is not None:
                    fills.append(fill)

            output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_path)
            return fills
        finally:
            workbook.close()

    def validate_fills(
        self,
        mappings: list[FilterMapping],
        fills: list[FillRecord],
    ) -> tuple[bool, str]:
        if not fills:
            return (
                False,
                "No SEC-backed values could be mapped into the workbook template.",
            )

        filled_cells = {(fill.sheet, fill.cell) for fill in fills}
        missing = [
            f"{mapping.sheet}!{mapping.cell}"
            for mapping in mappings
            if (mapping.sheet, mapping.cell.upper()) not in filled_cells
        ]
        if missing:
            return (
                True,
                f"Applied {len(fills)} SEC-backed fills. Missing values for "
                f"{len(missing)} mapped cells: {', '.join(missing[:5])}",
            )
        return True, f"Validated {len(fills)} SEC-backed fills across mapped template cells."

    def _apply_mapping(
        self,
        workbook,
        mapping: FilterMapping,
        profile: SecCompanyProfile,
    ) -> FillRecord | None:
        if mapping.sheet not in workbook.sheetnames:
            return None

        sheet = workbook[mapping.sheet]
        cell = sheet[mapping.cell.upper()]
        if cell.data_type == "f" or (cell.value not in (None, "") and str(cell.value).strip() != ""):
            return None

        value, xbrl_tag = self.sec_service.lookup_fact(profile, mapping.concept, mapping.period)
        if value is None:
            return None

        cell.value = value
        return FillRecord(
            sheet=mapping.sheet,
            cell=mapping.cell.upper(),
            concept=mapping.concept,
            period=mapping.period,
            value=value,
            source="sec_edgar_companyfacts",
            xbrl_tag=xbrl_tag,
        )

    def _scan_blank_cells(self, sheet: Worksheet) -> list[WorkbookCell]:
        blanks: list[WorkbookCell] = []
        for row in sheet.iter_rows():
            for cell in row:
                is_formula = cell.data_type == "f"
                is_blank = cell.value in (None, "") or (
                    isinstance(cell.value, str) and cell.value.strip() == ""
                )
                if is_blank and not is_formula:
                    blanks.append(
                        WorkbookCell(
                            sheet=sheet.title,
                            cell=cell.coordinate,
                            value=None,
                            is_formula=False,
                            is_blank=True,
                        )
                    )
        return blanks

    def _scan_identity_fields(self, sheet: Worksheet) -> tuple[str | None, str | None]:
        ticker: str | None = None
        company: str | None = None

        for row in sheet.iter_rows(min_row=1, max_row=40, min_col=1, max_col=6):
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                label = cell.value.strip().lower()
                neighbor = sheet.cell(row=cell.row, column=cell.column + 1)
                neighbor_value = neighbor.value
                if label in {"ticker", "symbol"} and isinstance(neighbor_value, str):
                    candidate = neighbor_value.strip().upper()
                    if TICKER_PATTERN.match(candidate):
                        ticker = candidate
                if label in {"company", "company name", "issuer"} and isinstance(neighbor_value, str):
                    company = neighbor_value.strip()

        return ticker, company
