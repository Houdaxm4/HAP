"""File upload handling for analysis workbooks."""

from __future__ import annotations

import re
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook

from models.analysis import Analysis, AnalysisFiles, UploadedFileMetadata
from models.common import utc_now_iso
from services.custom_run_service import CustomRunParseError, CustomRunService

UPLOADS_DIR = Path(__file__).resolve().parent.parent / "storage" / "uploads"

# Form field names accepted by the upload endpoint.
PREFILLED_WORKBOOK_FIELD = "prefilled_workbook"
PREVIOUS_WORKBOOK_FIELD = "previous_workbook"
CUSTOM_RUN_FILTER_FIELD = "custom_run_filter"

WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
CUSTOM_RUN_EXTENSIONS = {".csv", ".xlsx", ".xlsm", ".xls"}


class FileUploadError(Exception):
    """Raised when a required upload is missing or invalid."""


class FileService:
    """Save uploaded workbooks under per-analysis directories."""

    def __init__(
        self,
        uploads_dir: Path | None = None,
        custom_run_service: CustomRunService | None = None,
    ) -> None:
        self.uploads_dir = uploads_dir or UPLOADS_DIR
        self.uploads_dir.mkdir(parents=True, exist_ok=True)
        self.custom_run_service = custom_run_service or CustomRunService()

    def analysis_upload_dir(self, analysis_id: str) -> Path:
        """Return (and create) the upload directory for an analysis."""
        directory = self.uploads_dir / analysis_id
        directory.mkdir(parents=True, exist_ok=True)
        return directory

    async def save_upload(
        self,
        analysis_id: str,
        upload: UploadFile,
        stored_filename: str,
    ) -> UploadedFileMetadata:
        """Write an uploaded file to disk and return its metadata."""
        directory = self.analysis_upload_dir(analysis_id)
        destination = directory / stored_filename

        content = await upload.read()
        destination.write_bytes(content)

        original_name = upload.filename or stored_filename
        return UploadedFileMetadata(
            filename=original_name,
            stored_filename=stored_filename,
            size_bytes=len(content),
            uploaded_at=utc_now_iso(),
        )

    async def handle_uploads(
        self,
        analysis: Analysis,
        prefilled_workbook: UploadFile | None,
        previous_workbook: UploadFile | None = None,
        custom_run_filter: UploadFile | None = None,
    ) -> Analysis:
        """
        Validate and save required uploads, then merge metadata into the analysis.

        Both ``prefilled_workbook`` and ``custom_run_filter`` are required.
        """
        if prefilled_workbook is None or not prefilled_workbook.filename:
            raise FileUploadError("prefilled_workbook is required.")
        if custom_run_filter is None or not custom_run_filter.filename:
            raise FileUploadError("custom_run_filter is required.")

        self._assert_extension(
            prefilled_workbook.filename,
            WORKBOOK_EXTENSIONS,
            "prefilled_workbook",
        )
        self._assert_extension(
            custom_run_filter.filename,
            CUSTOM_RUN_EXTENSIONS,
            "custom_run_filter",
        )

        analysis_id = analysis.analysis_id
        files = AnalysisFiles(
            prefilled_workbook=await self.save_upload(
                analysis_id,
                prefilled_workbook,
                "prefilled_workbook.xlsx",
            )
        )

        if previous_workbook is not None and previous_workbook.filename:
            self._assert_extension(
                previous_workbook.filename,
                WORKBOOK_EXTENSIONS,
                "previous_workbook",
            )
            files.previous_workbook = await self.save_upload(
                analysis_id,
                previous_workbook,
                "previous_workbook.xlsx",
            )

        stored_name = self._stored_filename(
            custom_run_filter.filename,
            default_stem="custom_run_filter",
            allowed_extensions=CUSTOM_RUN_EXTENSIONS,
        )
        files.custom_run_filter = await self.save_upload(
            analysis_id,
            custom_run_filter,
            stored_name,
        )

        # Validate file contents after they are on disk.
        workbook_path = self.analysis_upload_dir(analysis_id) / files.prefilled_workbook.stored_filename
        custom_run_path = (
            self.analysis_upload_dir(analysis_id) / files.custom_run_filter.stored_filename
        )
        self._validate_workbook_file(workbook_path)
        self._validate_custom_run_file(custom_run_path, files.custom_run_filter.filename)

        analysis.files = files
        analysis.status = "uploaded"
        analysis.updated_at = utc_now_iso()
        return analysis

    def get_custom_run_filter_path(self, analysis: Analysis) -> Path:
        """Resolve the on-disk path to the custom_run filter file."""
        if analysis.files.custom_run_filter is None:
            raise FileUploadError("No custom_run_filter has been uploaded for this analysis.")

        path = (
            self.analysis_upload_dir(analysis.analysis_id)
            / analysis.files.custom_run_filter.stored_filename
        )
        if not path.exists():
            raise FileUploadError("custom_run_filter file is missing on disk.")
        return path

    @staticmethod
    def _assert_extension(filename: str, allowed: set[str], field_name: str) -> None:
        suffix = Path(filename).suffix.lower()
        if suffix not in allowed:
            raise FileUploadError(
                f"Unsupported file extension '{suffix}' for {field_name}. "
                f"Allowed: {', '.join(sorted(allowed))}"
            )

    @staticmethod
    def _validate_workbook_file(path: Path) -> None:
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                if not workbook.sheetnames:
                    raise FileUploadError("prefilled_workbook contains no worksheets.")
            finally:
                workbook.close()
        except FileUploadError:
            raise
        except Exception as exc:  # noqa: BLE001 — surface openpyxl errors as upload errors
            raise FileUploadError(f"prefilled_workbook is not a valid Excel workbook: {exc}") from exc

    def _validate_custom_run_file(self, path: Path, original_filename: str) -> None:
        try:
            self.custom_run_service.parse(path, original_filename)
        except CustomRunParseError as exc:
            raise FileUploadError(str(exc)) from exc

    @staticmethod
    def _stored_filename(
        original_filename: str,
        default_stem: str,
        allowed_extensions: set[str],
    ) -> str:
        suffix = Path(original_filename).suffix.lower()
        if suffix not in allowed_extensions:
            raise FileUploadError(
                f"Unsupported file extension '{suffix}' for {default_stem}. "
                f"Allowed: {', '.join(sorted(allowed_extensions))}"
            )
        safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(original_filename).stem) or default_stem
        return f"{safe_stem}{suffix}"

    def get_prefilled_workbook_path(self, analysis: Analysis) -> Path:
        """Resolve the on-disk path to the prefilled workbook."""
        if analysis.files.prefilled_workbook is None:
            raise FileUploadError("No prefilled workbook has been uploaded for this analysis.")

        path = (
            self.analysis_upload_dir(analysis.analysis_id)
            / analysis.files.prefilled_workbook.stored_filename
        )
        if not path.exists():
            raise FileUploadError("Prefilled workbook file is missing on disk.")
        return path
