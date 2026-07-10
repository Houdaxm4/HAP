"""Deep workbook parsing and safe value-only writes."""

from __future__ import annotations

import re
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from models.provenance import CellProvenance
from models.workbook_schema import (
    CellFormatting,
    CellInfo,
    NamedRangeInfo,
    WorkbookMetadata,
    WorkbookStructure,
    WorkbookSummary,
    WorksheetInfo,
)

CELL_REF_PATTERN = re.compile(r"^([A-Za-z][A-Za-z0-9_ ]+)!([A-Z]+[0-9]+)$")


class WorkbookParseError(Exception):
    """Raised when a workbook cannot be parsed safely."""


class WorkbookService:
    """Inspect, parse, and safely update Excel workbooks."""

    def read_summary(self, workbook_path: Path, original_filename: str) -> WorkbookSummary:
        """Return high-level workbook statistics without modifying the file."""
        structure = self.parse_structure(workbook_path, original_filename)
        return WorkbookSummary(
            workbook_filename=structure.workbook_filename,
            worksheet_names=structure.worksheet_names,
            sheet_count=len(structure.worksheet_names),
            visible_sheets=structure.visible_sheets,
            hidden_sheets=structure.hidden_sheets,
            formula_count=structure.formula_count,
            editable_cell_count=structure.editable_cell_count,
            non_empty_cell_count=structure.non_empty_cell_count,
            named_range_count=len(structure.named_ranges),
        )

    def parse_structure(self, workbook_path: Path, original_filename: str) -> WorkbookStructure:
        """
        Parse workbook structure into JSON-serializable metadata.

        Detects worksheets, editable cells, formula cells, named ranges,
        hidden sheets, document properties, and cell formatting.

        The source workbook is opened read-only for values/formulas and is
        never modified. Formatting is captured so later writes can preserve it.
        """
        # data_only=False keeps formulas; we never save this workbook handle.
        workbook = load_workbook(workbook_path, read_only=False, data_only=False)
        try:
            worksheets: list[WorksheetInfo] = []
            visible_sheets: list[str] = []
            hidden_sheets: list[str] = []
            formula_count = 0
            editable_cell_count = 0
            non_empty_cell_count = 0
            flat_formula_cells: list[str] = []
            flat_editable_cells: list[str] = []

            for index, name in enumerate(workbook.sheetnames):
                sheet = workbook[name]
                visibility = self._sheet_visibility(sheet)
                if visibility == "visible":
                    visible_sheets.append(name)
                else:
                    hidden_sheets.append(name)

                sheet_info = self._parse_worksheet(sheet, index=index)
                formula_count += sheet_info.formula_count
                editable_cell_count += sheet_info.editable_cell_count
                non_empty_cell_count += sheet_info.non_empty_cell_count
                for cell in sheet_info.formula_cells:
                    flat_formula_cells.append(f"{name}!{cell.address}")
                for cell in sheet_info.editable_cells:
                    flat_editable_cells.append(f"{name}!{cell.address}")
                worksheets.append(sheet_info)

            named_ranges = self._parse_named_ranges(workbook)
            metadata = self._parse_metadata(workbook, named_range_count=len(named_ranges))

            return WorkbookStructure(
                workbook_filename=original_filename,
                metadata=metadata,
                worksheet_names=list(workbook.sheetnames),
                visible_sheets=visible_sheets,
                hidden_sheets=hidden_sheets,
                named_ranges=named_ranges,
                worksheets=worksheets,
                formula_count=formula_count,
                editable_cell_count=editable_cell_count,
                non_empty_cell_count=non_empty_cell_count,
                formula_cells=flat_formula_cells,
                editable_cells=flat_editable_cells,
            )
        finally:
            workbook.close()

    def _parse_worksheet(self, sheet: Worksheet, index: int) -> WorksheetInfo:
        cells: list[CellInfo] = []
        formula_cells: list[CellInfo] = []
        editable_cells: list[CellInfo] = []
        sheet_formula_count = 0
        sheet_value_count = 0
        sheet_blank_count = 0
        sheet_editable_count = 0

        max_row = sheet.max_row or 0
        max_column = sheet.max_column or 0

        if max_row > 0 and max_column > 0:
            for row in sheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_column,
            ):
                for cell in row:
                    cell_info = self._parse_cell(cell)
                    if cell_info.data_type == "blank":
                        sheet_blank_count += 1
                        # Skip storing empty unformatted blanks to keep JSON lean.
                        if cell_info.formatting is None:
                            continue
                        # Styled blanks are still editable input candidates.
                        sheet_editable_count += 1
                        editable_cells.append(cell_info)
                        cells.append(cell_info)
                        continue

                    cells.append(cell_info)
                    if cell_info.is_formula:
                        sheet_formula_count += 1
                        formula_cells.append(cell_info)
                    else:
                        sheet_value_count += 1
                        sheet_editable_count += 1
                        editable_cells.append(cell_info)

        return WorksheetInfo(
            name=sheet.title,
            visibility=self._sheet_visibility(sheet),
            index=index,
            dimensions=sheet.dimensions if max_row and max_column else None,
            max_row=max_row or None,
            max_column=max_column or None,
            formula_count=sheet_formula_count,
            editable_cell_count=sheet_editable_count,
            value_count=sheet_value_count,
            blank_count=sheet_blank_count,
            non_empty_cell_count=sheet_value_count + sheet_formula_count,
            formula_cells=formula_cells,
            editable_cells=editable_cells,
            cells=cells,
        )

    def _parse_cell(self, cell: Cell) -> CellInfo:
        formatting = self._extract_formatting(cell)
        is_blank = cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == "")
        is_formula = cell.data_type == "f" or (
            isinstance(cell.value, str) and cell.value.startswith("=")
        )

        if is_blank and not is_formula:
            return CellInfo(
                address=cell.coordinate,
                row=cell.row,
                column=cell.column,
                data_type="blank",
                is_formula=False,
                is_editable=True,
                formatting=formatting,
            )

        if is_formula:
            formula = str(cell.value) if cell.value is not None else None
            return CellInfo(
                address=cell.coordinate,
                row=cell.row,
                column=cell.column,
                value=cell.value,
                data_type="formula",
                formula=formula,
                is_formula=True,
                is_editable=False,
                formatting=formatting,
            )

        return CellInfo(
            address=cell.coordinate,
            row=cell.row,
            column=cell.column,
            value=self._serialize_value(cell.value),
            data_type="value",
            is_formula=False,
            is_editable=True,
            formatting=formatting,
        )

    def _extract_formatting(self, cell: Cell) -> CellFormatting | None:
        """Capture non-default formatting needed to preserve appearance later."""
        number_format = cell.number_format if cell.number_format not in (None, "General") else None

        font = cell.font
        font_bold = bool(font.bold) if font and font.bold else None
        font_italic = bool(font.italic) if font and font.italic else None
        # Ignore default Calibri/11 theme fonts; keep explicit customizations.
        font_name = None
        font_size = None
        font_color = None
        if font and (font_bold or font_italic or (font.name and font.name != "Calibri") or (font.size and float(font.size) != 11.0)):
            font_name = font.name if font.name and font.name != "Calibri" else None
            font_size = float(font.size) if font.size is not None and float(font.size) != 11.0 else None
            font_color = self._color_value(getattr(font, "color", None))

        fill = cell.fill
        fill_pattern = None
        fill_color = None
        if fill is not None and fill.fill_type and fill.fill_type != "none":
            fill_pattern = str(fill.fill_type)
            fill_color = self._color_value(getattr(fill, "fgColor", None))

        alignment = cell.alignment
        horizontal = alignment.horizontal if alignment and alignment.horizontal else None
        vertical = alignment.vertical if alignment and alignment.vertical else None
        wrap_text = True if alignment and alignment.wrap_text else None

        protection = cell.protection
        # Excel defaults locked=True; only record explicit unlocks.
        locked = False if protection and protection.locked is False else None

        formatting = CellFormatting(
            number_format=number_format,
            font_name=font_name,
            font_size=font_size,
            font_bold=font_bold,
            font_italic=font_italic,
            font_color=font_color,
            fill_pattern=fill_pattern,
            fill_color=fill_color,
            horizontal_alignment=horizontal,
            vertical_alignment=vertical,
            wrap_text=wrap_text,
            locked=locked,
        )

        # Omit empty formatting objects to keep JSON compact.
        if all(value is None for value in formatting.model_dump().values()):
            return None
        return formatting

    @staticmethod
    def _color_value(color: Any) -> str | None:
        if color is None:
            return None
        rgb = getattr(color, "rgb", None)
        if isinstance(rgb, str) and rgb and rgb != "00000000":
            return rgb
        theme = getattr(color, "theme", None)
        if theme is not None:
            return f"theme:{theme}"
        indexed = getattr(color, "indexed", None)
        if indexed is not None:
            return f"indexed:{indexed}"
        return None

    @staticmethod
    def _serialize_value(value: Any) -> Any:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return value

    def write_values(
        self,
        source_workbook_path: Path,
        destination_workbook_path: Path,
        provenance_entries: list[CellProvenance],
    ) -> tuple[int, int, int]:
        """
        Copy the source workbook and write only value cells from provenance.

        Returns (filled_count, blank_count, skipped_formula_count).
        Never overwrites formula cells. Only ``cell.value`` is assigned so
        existing formatting (fonts, fills, number formats) is preserved.
        """
        destination_workbook_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_workbook_path, destination_workbook_path)

        workbook = load_workbook(destination_workbook_path)
        try:
            filled_count = 0
            blank_count = 0
            skipped_formula_count = 0

            for entry in provenance_entries:
                sheet = self._get_sheet(workbook, entry.worksheet)
                cell = sheet[entry.cell]

                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and str(cell.value).startswith("=")
                ):
                    entry.status = "skipped_formula"
                    entry.failure_reason = "Cell contains a formula and was not overwritten."
                    skipped_formula_count += 1
                    continue

                if entry.value is None or entry.status != "filled":
                    blank_count += 1
                    continue

                # Assign value only — do not clear or replace cell styles.
                cell.value = entry.value
                filled_count += 1

            workbook.save(destination_workbook_path)
            return filled_count, blank_count, skipped_formula_count
        finally:
            workbook.close()

    def get_cell_value(self, workbook_path: Path, worksheet: str, cell: str) -> Any:
        """Read a single cell value from a workbook."""
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            sheet = self._get_sheet(workbook, worksheet)
            return sheet[cell.upper()].value
        finally:
            workbook.close()

    def cell_contains_formula(self, structure: WorkbookStructure, worksheet: str, cell: str) -> bool:
        """Return True if the target cell is a formula in the parsed structure."""
        target = cell.upper()
        for sheet in structure.worksheets:
            if sheet.name != worksheet:
                continue
            for cell_info in sheet.formula_cells:
                if cell_info.address == target:
                    return True
            for cell_info in sheet.cells:
                if cell_info.address == target:
                    return cell_info.is_formula
        return False

    @staticmethod
    def make_cell_ref(worksheet: str, cell: str) -> str:
        """Build a stable cell reference key used in provenance and validation."""
        return f"{worksheet}!{cell.upper()}"

    @staticmethod
    def parse_cell_ref(cell_ref: str) -> tuple[str, str]:
        """Split a cell reference into worksheet and cell address."""
        match = CELL_REF_PATTERN.match(cell_ref)
        if not match:
            raise WorkbookParseError(f"Invalid cell reference: {cell_ref}")
        return match.group(1), match.group(2)

    @staticmethod
    def _get_sheet(workbook: Any, worksheet_name: str) -> Worksheet:
        if worksheet_name not in workbook.sheetnames:
            raise WorkbookParseError(f"Worksheet '{worksheet_name}' not found in workbook.")
        return workbook[worksheet_name]

    @staticmethod
    def _sheet_visibility(sheet: Worksheet) -> str:
        state = sheet.sheet_state
        if state == "veryHidden":
            return "veryHidden"
        if state == "hidden":
            return "hidden"
        return "visible"

    @staticmethod
    def _parse_named_ranges(workbook: Any) -> list[NamedRangeInfo]:
        named_ranges: list[NamedRangeInfo] = []
        defined_names = getattr(workbook, "defined_names", None)
        if defined_names is None:
            return named_ranges

        for name in defined_names:
            definition = defined_names[name]
            destinations: list[str] = []
            attr_text = None
            if definition is not None:
                attr_text = getattr(definition, "attr_text", None)
                try:
                    for worksheet_title, coordinate in definition.destinations:
                        destinations.append(f"{worksheet_title}!{coordinate}")
                except (TypeError, ValueError, AttributeError):
                    if attr_text:
                        destinations.append(str(attr_text))
            named_ranges.append(
                NamedRangeInfo(
                    name=name,
                    destinations=destinations,
                    attr_text=str(attr_text) if attr_text is not None else None,
                )
            )
        return named_ranges

    @staticmethod
    def _parse_metadata(workbook: Any, named_range_count: int) -> WorkbookMetadata:
        props = getattr(workbook, "properties", None)

        def _prop(name: str) -> Any:
            if props is None:
                return None
            value = getattr(props, name, None)
            if isinstance(value, (datetime, date)):
                return value.isoformat()
            if value is None:
                return None
            text = str(value).strip()
            return text or None

        excel_base_date = None
        epoch = getattr(workbook, "epoch", None)
        if epoch is not None:
            excel_base_date = epoch.isoformat() if hasattr(epoch, "isoformat") else str(epoch)

        return WorkbookMetadata(
            title=_prop("title"),
            subject=_prop("subject"),
            creator=_prop("creator"),
            description=_prop("description"),
            keywords=_prop("keywords"),
            category=_prop("category"),
            last_modified_by=_prop("lastModifiedBy"),
            created=_prop("created"),
            modified=_prop("modified"),
            content_status=_prop("contentStatus"),
            revision=_prop("revision"),
            excel_base_date=excel_base_date,
            sheet_count=len(workbook.sheetnames),
            defined_name_count=named_range_count,
        )
