"""Tests for the institutional HAP workbook writer."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from analysis_engine.runner import AnalysisEngine
from canonical_model.primitives import FinancialPoint
from canonical_model import build_company_financial_model
from models.analysis import Analysis
from models.common import utc_now_iso
from models.custom_run import CustomRunData, CustomRunPeriods
from models.pipeline import DecisionLogEntry
from models.provenance import ProvenanceReport
from models.validation import DiscrepancyReport
from services.hap_workbook_service import SHEET_ORDER, HapWorkbookService


def test_hap_workbook_contains_all_seventeen_sheets(tmp_path: Path):
    analysis = Analysis(
        analysis_id="wb-test-1",
        company="Apple Inc.",
        ticker="AAPL",
        analysis_type="new_company",
        status="complete",
        created_at=utc_now_iso(),
        updated_at=utc_now_iso(),
        decision_log=[
            DecisionLogEntry(
                agent="Test",
                action="unit_test",
                detail="Workbook writer unit test",
            )
        ],
        cik="0000320193",
    )
    custom_run = CustomRunData(
        source_filename="Custom_Run_Filter_AAPL.xlsx",
        ticker="AAPL",
        company="APPLE INC",
        ticker_sheet_name="AAPL",
        metadata={"Industry Sector": "Technology"},
        summary={"Industry Sector": "Technology", "Latest Fiscal Quarter": "2026 Q2"},
        market_data={"Current Price (Live Price)": 276.83},
        valuation_metrics={"WACC": 0.09, "Current PE10": 28.5},
        quality_metrics={"Final Score": 80},
        assumptions={"wacc": 0.09},
        periods=CustomRunPeriods(
            dates=["2020-01-01"],
            fiscal_quarters=["2020 Q1"],
            fiscal_years=["FY2020"],
        ),
        period_count=1,
        series_count=0,
        summary_field_count=2,
    )
    model = build_company_financial_model(
        analysis_id=analysis.analysis_id,
        ticker="AAPL",
        company="Apple Inc.",
        analysis_type="new_company",
        provenance_report=ProvenanceReport(analysis_id=analysis.analysis_id, ticker="AAPL"),
        discrepancy_report=DiscrepancyReport(
            analysis_id=analysis.analysis_id,
            ticker="AAPL",
            summary="ok",
        ),
        company_facts={},
        custom_run=custom_run,
    )
    # Seed minimal statement history so statement sheets are non-empty.
    model.income_statement.revenue.upsert(
        FinancialPoint(period="FY2024", value=100.0, currency="USD", source="test")
    )
    model.income_statement.revenue.upsert(
        FinancialPoint(period="FY2025", value=110.0, currency="USD", source="test")
    )
    model.income_statement.net_income.upsert(
        FinancialPoint(period="FY2024", value=20.0, currency="USD", source="test")
    )
    model.income_statement.net_income.upsert(
        FinancialPoint(period="FY2025", value=22.0, currency="USD", source="test")
    )
    model.balance_sheet.total_assets.upsert(
        FinancialPoint(period="FY2025", value=300.0, currency="USD", source="test")
    )
    model.cash_flow_statement.free_cash_flow.upsert(
        FinancialPoint(period="FY2025", value=25.0, currency="USD", source="test")
    )
    model.market_data.share_price = 276.83
    model.refresh_periods()

    engine_result = AnalysisEngine().run(model)
    out = tmp_path / "hap_workbook.xlsx"
    HapWorkbookService().write(
        out,
        analysis=analysis,
        model=model,
        engine_result=engine_result,
        custom_run=custom_run,
        validation_report=DiscrepancyReport(
            analysis_id=analysis.analysis_id,
            ticker="AAPL",
            pass_count=1,
            summary="ok",
        ),
        provenance_report=ProvenanceReport(analysis_id=analysis.analysis_id, ticker="AAPL"),
    )

    assert out.exists()
    wb = load_workbook(out, read_only=True, data_only=True)
    assert tuple(wb.sheetnames) == SHEET_ORDER
    assert wb["Income Statement"]["A5"].value == "Revenue"
    assert wb["Cover"]["A1"].value == "HAP Institutional Workbook"
    assert wb["Executive Summary"]["A1"].value == "Executive Summary"
    wb.close()
