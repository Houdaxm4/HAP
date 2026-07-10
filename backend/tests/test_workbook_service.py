"""Unit tests for workbook parsing and safe writes."""

from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from models.provenance import CellProvenance
from services.workbook_service import WorkbookService


def test_parse_workbook_structure(sample_workbook):
    service = WorkbookService()
    structure = service.parse_structure(sample_workbook, "prefilled_workbook.xlsx")

    assert "Income Statement" in structure.worksheet_names
    assert structure.formula_count == 1
    assert structure.editable_cell_count >= 2
    assert "Income Statement!B10" in structure.formula_cells
    assert "Income Statement!A5" in structure.editable_cells
    # Unformatted blank cells are not listed as editable candidates.
    assert "Income Statement!B5" not in structure.formula_cells
    assert service.cell_contains_formula(structure, "Income Statement", "B10") is True
    assert service.cell_contains_formula(structure, "Income Statement", "B5") is False
    assert structure.metadata.sheet_count == len(structure.worksheet_names)


def test_parse_detects_hidden_sheets_named_ranges_and_formatting(tmp_path):
    from openpyxl.workbook.defined_name import DefinedName

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Model"
    sheet["A1"] = "Revenue"
    sheet["A1"].font = Font(bold=True, name="Calibri", size=12)
    sheet["B1"] = None
    sheet["B1"].number_format = "#,##0.00"
    sheet["B1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    sheet["B1"].alignment = Alignment(horizontal="right")
    sheet["C1"] = "=B1*1.1"
    sheet["C1"].font = Font(italic=True)

    hidden = workbook.create_sheet("Assumptions")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "WACC"

    workbook.defined_names.add(DefinedName(name="RevenueInput", attr_text="'Model'!$B$1"))
    workbook.properties.title = "HAP Sample Model"
    workbook.properties.creator = "HAP Tests"

    path = tmp_path / "rich_workbook.xlsx"
    workbook.save(path)
    workbook.close()

    structure = WorkbookService().parse_structure(path, "rich_workbook.xlsx")

    assert structure.hidden_sheets == ["Assumptions"]
    assert structure.visible_sheets == ["Model"]
    assert any(item.name == "RevenueInput" for item in structure.named_ranges)
    assert structure.metadata.title == "HAP Sample Model"
    assert structure.metadata.creator == "HAP Tests"
    assert structure.formula_count == 1
    assert "Model!C1" in structure.formula_cells
    assert "Model!B1" in structure.editable_cells

    model = next(ws for ws in structure.worksheets if ws.name == "Model")
    b1 = next(cell for cell in model.editable_cells if cell.address == "B1")
    assert b1.is_editable is True
    assert b1.formatting is not None
    assert b1.formatting.number_format == "#,##0.00"
    assert b1.formatting.fill_pattern == "solid"

    a1 = next(cell for cell in model.editable_cells if cell.address == "A1")
    assert a1.formatting is not None
    assert a1.formatting.font_bold is True

    c1 = next(cell for cell in model.formula_cells if cell.address == "C1")
    assert c1.is_editable is False
    assert c1.formula == "=B1*1.1"


def test_write_values_preserves_formulas_and_formatting(sample_workbook, tmp_path):
    service = WorkbookService()

    # Add formatting to the source editable cell before writing.
    source = load_workbook(sample_workbook)
    try:
        cell = source["Income Statement"]["B5"]
        cell.number_format = '#,##0'
        cell.font = Font(bold=True, name="Arial")
        source.save(sample_workbook)
    finally:
        source.close()

    destination = tmp_path / "completed_workbook.xlsx"
    provenance = [
        CellProvenance(
            cell_ref="Income Statement!B5",
            worksheet="Income Statement",
            cell="B5",
            concept="Revenue",
            period="FY2024",
            value=1000,
            status="filled",
        ),
        CellProvenance(
            cell_ref="Income Statement!B10",
            worksheet="Income Statement",
            cell="B10",
            concept="Total",
            period="FY2024",
            value=999,
            status="filled",
        ),
    ]

    filled, blank, skipped = service.write_values(sample_workbook, destination, provenance)
    assert filled == 1
    assert skipped == 1
    assert service.get_cell_value(destination, "Income Statement", "B5") == 1000

    workbook = load_workbook(destination, data_only=False)
    try:
        assert workbook["Income Statement"]["B10"].data_type == "f"
        written = workbook["Income Statement"]["B5"]
        assert written.value == 1000
        assert written.number_format == "#,##0"
        assert written.font.bold is True
        assert written.font.name == "Arial"
    finally:
        workbook.close()
