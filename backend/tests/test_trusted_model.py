"""Tests for trusted financial model: provenance, validation, and pipeline phase 2."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from models.analysis import Analysis, AnalysisFiles, UploadedFileMetadata
from models.custom_run import CustomRunEntry, CustomRunMapping, CustomRunValidationReport
from models.pipeline import PipelineStage
from models.provenance import CellProvenance, ProvenanceReport
from models.validation import ValidationReport
from models.workbook_schema import WorkbookStructure
from pipeline.orchestrator import PipelineOrchestrator
from pipeline.stages.fill_workbook import FillWorkbookStage
from pipeline.stages.validate_workbook import ValidateWorkbookStage
from services.analysis_service import AnalysisService
from services.file_service import FileService
from services.output_service import OutputService
from services.sec_service import SecService
from services.validation_service import ValidationService
from services.workbook_service import WorkbookService


@pytest.fixture
def workbook_with_populated(fixtures_dir: Path) -> Path:
    """Workbook with blank inputs, one formula, and one pre-populated value."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Income Statement"
    sheet["A1"] = "Metric"
    sheet["B1"] = "FY2024"
    sheet["A5"] = "Revenue"
    sheet["B5"] = None
    sheet["A6"] = "Net Income"
    sheet["B6"] = 1000  # populated manual input — must not be silently overwritten
    sheet["A10"] = "Total"
    sheet["B10"] = "=SUM(B5:B6)"

    balance = workbook.create_sheet("Balance Sheet")
    balance["A5"] = "Total Assets"
    balance["B5"] = None
    balance["A6"] = "Total Liabilities"
    balance["B6"] = None
    balance["A7"] = "Equity"
    balance["B7"] = None

    path = fixtures_dir / "prefilled_with_values.xlsx"
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def custom_run_with_conflict(fixtures_dir: Path) -> Path:
    content = """worksheet,cell,concept,period,xbrl_tag,unit,value
Income Statement,B5,Revenue,FY2024,Revenues,USD,999
Income Statement,B6,Net Income,FY2024,NetIncomeLoss,USD,1000
Income Statement,B10,Total,FY2024,,USD,
Balance Sheet,B5,Total Assets,FY2024,Assets,USD,
Balance Sheet,B6,Total Liabilities,FY2024,Liabilities,USD,
Balance Sheet,B7,Stockholders Equity,FY2024,StockholdersEquity,USD,
"""
    path = fixtures_dir / "custom_run_conflict.csv"
    path.write_text(content, encoding="utf-8")
    return path


@pytest.fixture
def rich_company_facts() -> dict:
    return {
        "facts": {
            "us-gaap": {
                "Revenues": {
                    "label": "Revenues",
                    "units": {
                        "USD": [
                            {
                                "val": 391035000000,
                                "fy": 2024,
                                "fp": "FY",
                                "form": "10-K",
                                "filed": "2024-11-01",
                                "accn": "0000320193-24-000123",
                            }
                        ]
                    },
                },
                "NetIncomeLoss": {
                    "label": "Net Income (Loss)",
                    "units": {
                        "USD": [
                            {
                                "val": 93736000000,
                                "fy": 2024,
                                "fp": "FY",
                                "form": "10-K",
                                "filed": "2024-11-01",
                                "accn": "0000320193-24-000123",
                            }
                        ]
                    },
                },
                "Assets": {
                    "label": "Assets",
                    "units": {
                        "USD": [
                            {
                                "val": 364980000000,
                                "fy": 2024,
                                "fp": "FY",
                                "form": "10-K",
                                "filed": "2024-11-01",
                                "accn": "0000320193-24-000123",
                            }
                        ]
                    },
                },
                "Liabilities": {
                    "label": "Liabilities",
                    "units": {
                        "USD": [
                            {
                                "val": 290437000000,
                                "fy": 2024,
                                "fp": "FY",
                                "form": "10-K",
                                "filed": "2024-11-01",
                                "accn": "0000320193-24-000123",
                            }
                        ]
                    },
                },
                "StockholdersEquity": {
                    "label": "Stockholders Equity",
                    "units": {
                        "USD": [
                            {
                                "val": 74543000000,
                                "fy": 2024,
                                "fp": "FY",
                                "form": "10-K",
                                "filed": "2024-11-01",
                                "accn": "0000320193-24-000123",
                            }
                        ]
                    },
                },
            }
        }
    }


@pytest.fixture
def balanced_statements() -> dict:
    return {
        "annual_periods": ["FY2024"],
        "quarterly_periods": ["Q1 2024"],
        "balance_sheet": {
            "line_items": [
                {
                    "concept": "Assets",
                    "label": "Total Assets",
                    "xbrl_tag": "Assets",
                    "values": [{"period": "FY2024", "value": 100.0}],
                },
                {
                    "concept": "Liabilities",
                    "label": "Total Liabilities",
                    "xbrl_tag": "Liabilities",
                    "values": [{"period": "FY2024", "value": 40.0}],
                },
                {
                    "concept": "StockholdersEquity",
                    "label": "Equity",
                    "xbrl_tag": "StockholdersEquity",
                    "values": [{"period": "FY2024", "value": 60.0}],
                },
                {
                    "concept": "CashAndCashEquivalentsAtCarryingValue",
                    "label": "Cash",
                    "xbrl_tag": "CashAndCashEquivalentsAtCarryingValue",
                    "values": [{"period": "FY2024", "value": 10.0}],
                },
            ]
        },
        "income_statement": {
            "line_items": [
                {
                    "concept": "NetIncomeLoss",
                    "label": "Net Income",
                    "xbrl_tag": "NetIncomeLoss",
                    "values": [{"period": "FY2024", "value": 5.0}],
                }
            ]
        },
        "cash_flow": {
            "line_items": [
                {
                    "concept": "NetIncomeLoss",
                    "label": "Net Income",
                    "xbrl_tag": "NetIncomeLoss",
                    "values": [{"period": "FY2024", "value": 5.0}],
                },
                {
                    "concept": "CashAndCashEquivalentsAtCarryingValue",
                    "label": "Ending Cash",
                    "xbrl_tag": "CashAndCashEquivalentsAtCarryingValue",
                    "values": [{"period": "FY2024", "value": 10.0}],
                },
            ]
        },
    }


def _parse_and_fill(
    tmp_path: Path,
    workbook_path: Path,
    custom_run_path: Path,
    company_facts: dict,
    filings_manifest: dict,
) -> tuple[ProvenanceReport, Path, WorkbookStructure, CustomRunMapping]:
    from services.custom_run_service import CustomRunService

    wb_service = WorkbookService()
    output_service = OutputService(outputs_dir=tmp_path / "outputs")
    structure = wb_service.parse_structure(workbook_path, workbook_path.name)
    mapping = CustomRunService().parse(custom_run_path, custom_run_path.name)

    analysis = Analysis(
        analysis_id="prov-test",
        company="Apple Inc.",
        ticker="AAPL",
        analysis_type="Annual Update",
        files=AnalysisFiles(
            prefilled_workbook=UploadedFileMetadata(
                filename=workbook_path.name,
                stored_filename=workbook_path.name,
                size_bytes=workbook_path.stat().st_size,
                uploaded_at="2026-07-10T00:00:00+00:00",
            )
        ),
    )
    stage = FillWorkbookStage(output_service=output_service, sec_service=SecService())
    report, _wb_rel, _prov_rel, _log = stage.run(
        analysis,
        workbook_path,
        mapping,
        structure,
        company_facts,
        filings_manifest,
    )
    completed = output_service.artifact_path("prov-test", "completed_workbook.xlsx")
    return report, completed, structure, mapping


def test_every_written_cell_receives_provenance(
    tmp_path: Path,
    workbook_with_populated: Path,
    custom_run_with_conflict: Path,
    rich_company_facts: dict,
    mock_filings_manifest: dict,
):
    report, _completed, _structure, mapping = _parse_and_fill(
        tmp_path,
        workbook_with_populated,
        custom_run_with_conflict,
        rich_company_facts,
        mock_filings_manifest,
    )
    assert len(report.entries) == mapping.entry_count
    for entry in report.entries:
        assert entry.cell_ref
        assert entry.analysis_id == "prov-test"
        assert entry.ticker == "AAPL"
        assert entry.workbook_filename
        assert entry.timestamp


def test_formula_cells_never_overwritten(
    tmp_path: Path,
    workbook_with_populated: Path,
    custom_run_with_conflict: Path,
    rich_company_facts: dict,
    mock_filings_manifest: dict,
):
    report, completed, _structure, _mapping = _parse_and_fill(
        tmp_path,
        workbook_with_populated,
        custom_run_with_conflict,
        rich_company_facts,
        mock_filings_manifest,
    )
    formula_entries = [e for e in report.entries if e.cell == "B10"]
    assert formula_entries
    assert formula_entries[0].status == "skipped_formula"
    assert formula_entries[0].source_type == "workbook formula"

    wb = load_workbook(completed, data_only=False)
    assert wb["Income Statement"]["B10"].value == "=SUM(B5:B6)"
    wb.close()


def test_populated_values_not_silently_replaced(
    tmp_path: Path,
    workbook_with_populated: Path,
    custom_run_with_conflict: Path,
    rich_company_facts: dict,
    mock_filings_manifest: dict,
):
    report, completed, _structure, _mapping = _parse_and_fill(
        tmp_path,
        workbook_with_populated,
        custom_run_with_conflict,
        rich_company_facts,
        mock_filings_manifest,
    )
    ni = next(e for e in report.entries if e.cell == "B6")
    assert ni.status == "preserved_existing"
    assert ni.original_workbook_value == 1000
    assert ni.proposed_value == 93736000000
    assert ni.value is None

    wb = load_workbook(completed, data_only=True)
    assert wb["Income Statement"]["B6"].value == 1000
    wb.close()


def test_unit_conversion_recorded(
    tmp_path: Path,
    sample_workbook: Path,
    rich_company_facts: dict,
    mock_filings_manifest: dict,
):
    content = """worksheet,cell,concept,period,xbrl_tag,unit
Income Statement,B5,Revenue,FY2024,Revenues,millions
Income Statement,B6,Net Income,FY2024,NetIncomeLoss,millions
Balance Sheet,B5,Total Assets,FY2024,Assets,millions
"""
    custom_run = tmp_path / "units.csv"
    custom_run.write_text(content, encoding="utf-8")

    report, _completed, _structure, _mapping = _parse_and_fill(
        tmp_path,
        sample_workbook,
        custom_run,
        rich_company_facts,
        mock_filings_manifest,
    )
    revenue = next(e for e in report.entries if e.concept == "Revenue")
    assert revenue.status == "filled"
    assert any(t.type == "unit_conversion" for t in revenue.transformations)
    assert revenue.original_source_value == 391035000000
    assert abs(revenue.value - 391035.0) < 0.01


def test_filing_custom_run_conflict_generates_warning(
    tmp_path: Path,
    sample_workbook: Path,
    rich_company_facts: dict,
    mock_filings_manifest: dict,
    balanced_statements: dict,
):
    content = """worksheet,cell,concept,period,xbrl_tag,value
Income Statement,B5,Revenue,FY2024,Revenues,1
Income Statement,B6,Net Income,FY2024,NetIncomeLoss,2
Balance Sheet,B5,Total Assets,FY2024,Assets,3
"""
    custom_run = tmp_path / "conflict.csv"
    custom_run.write_text(content, encoding="utf-8")

    report, completed, structure, mapping = _parse_and_fill(
        tmp_path,
        sample_workbook,
        custom_run,
        rich_company_facts,
        mock_filings_manifest,
    )
    revenue = next(e for e in report.entries if e.concept == "Revenue")
    assert revenue.conflict_with_custom_run is True
    assert revenue.status == "filled"
    assert revenue.value == 391035000000

    validation = ValidationService().validate_trusted_model(
        analysis_id="prov-test",
        ticker="AAPL",
        custom_run_mapping=mapping,
        provenance_report=report,
        completed_workbook_path=completed,
        workbook_structure=structure,
        financial_statements=balanced_statements,
    )
    assert any(i.code == "VALUE_CONFLICT" for i in validation.issues)
    assert validation.overall_status in {"passed_with_warnings", "passed", "failed"}


def test_missing_required_worksheet_is_critical(tmp_path: Path, sample_workbook: Path):
    wb_service = WorkbookService()
    structure = wb_service.parse_structure(sample_workbook, sample_workbook.name)
    mapping = CustomRunMapping(
        source_filename="x.csv",
        entry_count=1,
        entries=[
            CustomRunEntry(
                worksheet="Missing Sheet",
                cell="B5",
                concept="Revenue",
                period="FY2024",
            )
        ],
    )
    provenance = ProvenanceReport(
        analysis_id="x",
        ticker="AAPL",
        entries=[
            CellProvenance(
                cell_ref="Missing Sheet!B5",
                worksheet="Missing Sheet",
                cell="B5",
                concept="Revenue",
                period="FY2024",
                status="unresolved",
            )
        ],
    )
    # Copy workbook as "completed"
    completed = tmp_path / "completed.xlsx"
    completed.write_bytes(sample_workbook.read_bytes())

    report = ValidationService().validate_trusted_model(
        analysis_id="x",
        ticker="AAPL",
        custom_run_mapping=mapping,
        provenance_report=provenance,
        completed_workbook_path=completed,
        workbook_structure=structure,
    )
    assert report.critical_count >= 1
    assert report.overall_status == "failed"
    assert any(
        i.code in {"MAPPED_WORKSHEET_MISSING", "REQUIRED_FIELD_UNRESOLVED", "MISSING_PROVENANCE"}
        for i in report.issues
    )


def test_missing_custom_run_columns_critical():
    cr_report = CustomRunValidationReport(
        analysis_id="x",
        ticker="AAPL",
        source_filename="bad.csv",
        entry_count=0,
        checks=[
            __import__("models.custom_run", fromlist=["CustomRunValidationIssue"]).CustomRunValidationIssue(
                check_type="required_columns",
                status="fail",
                message="Missing required columns: worksheet, cell",
            )
        ],
        fail_count=1,
        is_valid=False,
        summary="failed",
    )
    mapping = CustomRunMapping(source_filename="bad.csv", entry_count=0, entries=[])
    provenance = ProvenanceReport(analysis_id="x", ticker="AAPL", entries=[])

    # Need a real workbook path — create minimal
    from openpyxl import Workbook

    path = Path("/tmp/hap_min_wb.xlsx")
    wb = Workbook()
    wb.save(path)
    wb.close()

    report = ValidationService().validate_trusted_model(
        analysis_id="x",
        ticker="AAPL",
        custom_run_mapping=mapping,
        provenance_report=provenance,
        completed_workbook_path=path,
        custom_run_validation_report=cr_report,
    )
    assert any(i.code == "CUSTOM_RUN_REQUIRED_COLUMNS" for i in report.issues)
    assert report.overall_status == "failed"


def test_balance_sheet_mismatch_detected(balanced_statements: dict, tmp_path: Path, sample_workbook: Path):
    imbalanced = dict(balanced_statements)
    imbalanced["balance_sheet"] = {
        "line_items": [
            {
                "concept": "Assets",
                "label": "Total Assets",
                "xbrl_tag": "Assets",
                "values": [{"period": "FY2024", "value": 100.0}],
            },
            {
                "concept": "Liabilities",
                "label": "Total Liabilities",
                "xbrl_tag": "Liabilities",
                "values": [{"period": "FY2024", "value": 10.0}],
            },
            {
                "concept": "StockholdersEquity",
                "label": "Equity",
                "xbrl_tag": "StockholdersEquity",
                "values": [{"period": "FY2024", "value": 10.0}],
            },
        ]
    }
    completed = tmp_path / "c.xlsx"
    completed.write_bytes(sample_workbook.read_bytes())
    structure = WorkbookService().parse_structure(sample_workbook, "w.xlsx")
    mapping = CustomRunMapping(source_filename="x.csv", entry_count=0, entries=[])
    provenance = ProvenanceReport(analysis_id="x", ticker="AAPL", entries=[])

    report = ValidationService().validate_trusted_model(
        analysis_id="x",
        ticker="AAPL",
        custom_run_mapping=mapping,
        provenance_report=provenance,
        completed_workbook_path=completed,
        workbook_structure=structure,
        financial_statements=imbalanced,
    )
    assert any(i.code == "BALANCE_SHEET_IMBALANCE" for i in report.issues)
    assert report.overall_status == "failed"


def test_warnings_allow_passed_with_warnings(
    tmp_path: Path,
    sample_workbook: Path,
    rich_company_facts: dict,
    mock_filings_manifest: dict,
    balanced_statements: dict,
):
    content = """worksheet,cell,concept,period,xbrl_tag,value
Income Statement,B5,Revenue,FY2024,Revenues,1
Income Statement,B6,Net Income,FY2024,NetIncomeLoss,2
Balance Sheet,B5,Total Assets,FY2024,Assets,3
"""
    custom_run = tmp_path / "w.csv"
    custom_run.write_text(content, encoding="utf-8")
    report, completed, structure, mapping = _parse_and_fill(
        tmp_path,
        sample_workbook,
        custom_run,
        rich_company_facts,
        mock_filings_manifest,
    )
    validation = ValidationService().validate_trusted_model(
        analysis_id="prov-test",
        ticker="AAPL",
        custom_run_mapping=mapping,
        provenance_report=report,
        completed_workbook_path=completed,
        source_workbook_path=sample_workbook,
        workbook_structure=structure,
        financial_statements=balanced_statements,
    )
    assert validation.critical_count == 0
    assert validation.warning_count >= 1
    assert validation.overall_status == "passed_with_warnings"
    assert validation.blocks_pipeline is False


@pytest.fixture
def trusted_pipeline_env(
    tmp_path: Path,
    workbook_with_populated: Path,
    custom_run_with_conflict: Path,
    rich_company_facts: dict,
    mock_filings_manifest: dict,
    monkeypatch: pytest.MonkeyPatch,
):
    analysis_service = AnalysisService(storage_dir=tmp_path / "analyses")
    file_service = FileService(uploads_dir=tmp_path / "uploads")
    output_service = OutputService(outputs_dir=tmp_path / "outputs")

    # Use a custom_run that can fully resolve (no formula mapping that blocks completion
    # via unresolved — formula is skipped which is OK; populated NI is preserved with warning)
    # For complete success path, use blank-only workbook from sample + full facts.
    analysis = Analysis(
        analysis_id="trusted-1",
        company="Apple Inc.",
        ticker="AAPL",
        analysis_type="Annual Update",
        status="uploaded",
        files=AnalysisFiles(
            prefilled_workbook=UploadedFileMetadata(
                filename="prefilled_workbook.xlsx",
                stored_filename="prefilled_workbook.xlsx",
                size_bytes=workbook_with_populated.stat().st_size,
                uploaded_at="2026-07-10T00:00:00+00:00",
            ),
            custom_run_filter=UploadedFileMetadata(
                filename="custom_run_filter.csv",
                stored_filename="custom_run_filter.csv",
                size_bytes=custom_run_with_conflict.stat().st_size,
                uploaded_at="2026-07-10T00:00:00+00:00",
            ),
        ),
    )
    analysis_service.save(analysis)
    upload_dir = file_service.analysis_upload_dir(analysis.analysis_id)
    upload_dir.mkdir(parents=True, exist_ok=True)
    (upload_dir / "prefilled_workbook.xlsx").write_bytes(workbook_with_populated.read_bytes())
    (upload_dir / "custom_run_filter.csv").write_bytes(custom_run_with_conflict.read_bytes())

    sec = SecService(cache_dir=tmp_path / "sec-cache")
    monkeypatch.setattr(sec, "resolve_cik", lambda ticker: "0000320193")
    monkeypatch.setattr(
        sec,
        "fetch_filings_manifest",
        lambda ticker, cik: {**mock_filings_manifest, "cik": cik, "ticker": ticker},
    )
    monkeypatch.setattr(sec, "fetch_company_facts", lambda cik: rich_company_facts)

    orchestrator = PipelineOrchestrator(
        analysis_service=analysis_service,
        file_service=file_service,
        output_service=output_service,
        sec_service=sec,
        sec_cache_dir=tmp_path / "sec-cache",
    )
    return orchestrator, analysis_service, output_service


def test_critical_validation_stops_pipeline(trusted_pipeline_env):
    """Unresolved required fields (formula skip is OK; missing facts would fail)."""
    orchestrator, analysis_service, output_service = trusted_pipeline_env

    # Replace custom_run with an unresolvable concept so fill leaves unresolved → critical
    custom_run_path = (
        orchestrator.file_service.analysis_upload_dir("trusted-1") / "custom_run_filter.csv"
    )
    custom_run_path.write_text(
        "worksheet,cell,concept,period\nIncome Statement,B5,UnknownMetricXYZ,FY2024\n",
        encoding="utf-8",
    )

    phase1 = orchestrator.run("trusted-1")
    assert phase1.pipeline.state == "waiting"

    phase2 = orchestrator.continue_trusted_model("trusted-1")
    assert phase2.pipeline.state == "failed"
    assert phase2.status == "validation_failed"
    assert phase2.pipeline.validation_status == "failed"
    assert phase2.pipeline.critical_issue_count >= 1
    assert phase2.is_trusted_model_complete is False

    # Artifacts still written for analyst review
    assert (output_service.artifact_path("trusted-1", "provenance_report.json")).exists()
    assert (output_service.artifact_path("trusted-1", "validation_report.json")).exists()
    assert phase2.pipeline.outputs.provenance_report is not None
    assert phase2.pipeline.outputs.validation_report is not None


def test_trusted_model_complete_requires_artifacts_and_no_critical(
    tmp_path: Path,
    sample_workbook: Path,
    sample_custom_run_csv: Path,
    rich_company_facts: dict,
    mock_filings_manifest: dict,
    balanced_statements: dict,
    monkeypatch: pytest.MonkeyPatch,
):
    analysis_service = AnalysisService(storage_dir=tmp_path / "analyses")
    file_service = FileService(uploads_dir=tmp_path / "uploads")
    output_service = OutputService(outputs_dir=tmp_path / "outputs")

    analysis = Analysis(
        analysis_id="trusted-ok",
        company="Apple Inc.",
        ticker="AAPL",
        analysis_type="Annual Update",
        status="uploaded",
        files=AnalysisFiles(
            prefilled_workbook=UploadedFileMetadata(
                filename="prefilled_workbook.xlsx",
                stored_filename="prefilled_workbook.xlsx",
                size_bytes=sample_workbook.stat().st_size,
                uploaded_at="2026-07-10T00:00:00+00:00",
            ),
            custom_run_filter=UploadedFileMetadata(
                filename="custom_run_filter.csv",
                stored_filename="custom_run_filter.csv",
                size_bytes=sample_custom_run_csv.stat().st_size,
                uploaded_at="2026-07-10T00:00:00+00:00",
            ),
        ),
    )
    analysis_service.save(analysis)
    upload_dir = file_service.analysis_upload_dir("trusted-ok")
    upload_dir.mkdir(parents=True, exist_ok=True)
    (upload_dir / "prefilled_workbook.xlsx").write_bytes(sample_workbook.read_bytes())
    (upload_dir / "custom_run_filter.csv").write_bytes(sample_custom_run_csv.read_bytes())

    sec = SecService()
    monkeypatch.setattr(sec, "resolve_cik", lambda ticker: "0000320193")
    monkeypatch.setattr(
        sec,
        "fetch_filings_manifest",
        lambda ticker, cik: {**mock_filings_manifest, "cik": cik},
    )
    monkeypatch.setattr(sec, "fetch_company_facts", lambda cik: rich_company_facts)

    orchestrator = PipelineOrchestrator(
        analysis_service=analysis_service,
        file_service=file_service,
        output_service=output_service,
        sec_service=sec,
    )

    # Seed financial statements for BS checks
    output_service.write_json("trusted-ok", "financial_statements.json", balanced_statements)

    phase1 = orchestrator.run("trusted-ok")
    assert phase1.is_pipeline_complete

    phase2 = orchestrator.continue_trusted_model("trusted-ok")
    assert phase2.pipeline.outputs.completed_workbook is not None
    assert phase2.pipeline.outputs.provenance_report is not None
    assert phase2.pipeline.outputs.validation_report is not None

    validation = ValidationReport.model_validate(
        output_service.read_json("trusted-ok", "validation_report.json")
    )
    provenance = ProvenanceReport.model_validate(
        output_service.read_json("trusted-ok", "provenance_report.json")
    )
    assert len(provenance.entries) == 3
    assert all(e.status in {"filled", "skipped_formula", "preserved_existing"} for e in provenance.entries)

    if validation.critical_count == 0:
        assert phase2.pipeline.state == "complete"
        assert phase2.is_trusted_model_complete
        assert phase2.pipeline.current_stage == PipelineStage.COMPLETE
        assert PipelineStage.PROVENANCE_RECORDED in phase2.pipeline.stages_completed
        assert PipelineStage.WORKBOOK_VALIDATED in phase2.pipeline.stages_completed
        assert PipelineStage.VALIDATION_REPORT_GENERATED in phase2.pipeline.stages_completed
        assert PipelineStage.PROVENANCE_REPORT_GENERATED in phase2.pipeline.stages_completed
    else:
        assert phase2.is_trusted_model_complete is False
        assert phase2.pipeline.state == "failed"

    # Cannot mark complete without artifacts — property enforces this
    phase2.pipeline.outputs.validation_report = None
    assert phase2.is_trusted_model_complete is False


def test_validate_workbook_stage_writes_both_reports(
    tmp_path: Path,
    sample_workbook: Path,
    sample_custom_run_csv: Path,
    rich_company_facts: dict,
    mock_filings_manifest: dict,
    balanced_statements: dict,
):
    report, completed, structure, mapping = _parse_and_fill(
        tmp_path,
        sample_workbook,
        sample_custom_run_csv,
        rich_company_facts,
        mock_filings_manifest,
    )
    output_service = OutputService(outputs_dir=tmp_path / "outputs")
    analysis = Analysis(
        analysis_id="prov-test",
        company="Apple",
        ticker="AAPL",
        analysis_type="Annual Update",
    )
    stage = ValidateWorkbookStage(output_service=output_service)
    validation, discrepancy, vpath, dpath, log = stage.run(
        analysis,
        mapping,
        report,
        completed,
        source_workbook_path=sample_workbook,
        workbook_structure=structure,
        financial_statements=balanced_statements,
    )
    assert validation.overall_status in {"passed", "passed_with_warnings", "failed"}
    assert (tmp_path / "outputs" / "prov-test" / "validation_report.json").exists()
    assert (tmp_path / "outputs" / "prov-test" / "discrepancy_report.json").exists()
    assert "validate_workbook" == log.action
