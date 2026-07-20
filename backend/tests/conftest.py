"""Shared pytest fixtures for HAP backend pipeline tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"

# Fields that appear on the Bloomberg Summary sheet (truncated for fixtures).
_SUMMARY_HEADERS = [
    "Company",
    "Ticker",
    "Current Price (Live Price)",
    "Current Market Capitalization",
    "Current Enterprise Value (not-diluted)",
    "Current Dividend Yield",
    "Current Dividend Rate",
    "WACC",
    "Current PE10",
    "Current E10",
    "Current Graham Instrinsic Value",
    "Graham Instrinsic Value in 7 Years",
    "Graham Expected Annualized Return",
    "Expected Return @ Current Price",
    "Expected Return Price Plus Dividends - Given Current Price",
    "Max Current Price to Buy",
    "1st Exit Price",
    "2nd Exit Price",
    "Approximate Residual Earnings Value",
    "High Price in 10 years",
    "Low Price in 10 years",
    "ROC Greenblatt",
    "ROC - WACC",
    "EBIT TTM/EV",
    "Current Max PE10 to Enter (Lowest PE10 or 7PE10)",
    "Final Score",
    "Franchise Power",
    "Quality_FPFS",
    "P_FS",
    "P_SNOA",
    "P_ROA10",
    "P_ROC10",
    "P_CFOA10",
    "P_MG",
    "P_MS",
    "P_MM",
    "ROA10",
    "ROC10",
    "CFOA",
    "SNOA (Scaled Net Operating Assets)",
    "Profit Margin Growth",
    "Profit Margin Stability",
    "Number of Consecutive Positive Growth Margins",
    "ROCE",
    "ROCE Fiscal Year",
    "ROCE 5 Fiscal Years",
    "ROCE 10 Fiscal Years",
    "Tax Rate",
    "Shares Outstanding Diluted Average (MM)",
    "Sector",
    "Industry",
] + [f"Extra Field {i}" for i in range(1, 61)]


@pytest.fixture
def fixtures_dir(tmp_path: Path) -> Path:
    directory = tmp_path / "fixtures"
    directory.mkdir()
    return directory


@pytest.fixture
def sample_workbook(fixtures_dir: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Income Statement"
    sheet["A1"] = "Metric"
    sheet["B1"] = "FY2024"
    sheet["A5"] = "Revenue"
    sheet["B5"] = None
    sheet["A6"] = "Net Income"
    sheet["B6"] = None
    sheet["A10"] = "Total"
    sheet["B10"] = "=SUM(B5:B6)"

    balance = workbook.create_sheet("Balance Sheet")
    balance["A5"] = "Total Assets"
    balance["B5"] = None

    path = fixtures_dir / "prefilled_workbook.xlsx"
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def sample_custom_run_xlsx(fixtures_dir: Path) -> Path:
    """Minimal Bloomberg-shaped Custom_Run_Filter workbook (HAP v1 product input)."""
    workbook = Workbook()
    ticker_ws = workbook.active
    ticker_ws.title = "AAPL"

    ticker_ws["A2"] = "Company"
    ticker_ws["B2"] = "Apple Inc."
    ticker_ws["A3"] = "Ticker"
    ticker_ws["B3"] = "AAPL"
    ticker_ws["A4"] = "Sector"
    ticker_ws["B4"] = "Technology"

    # Period axes (rows 15–16, fiscal years on 146) — 24 quarters.
    ticker_ws["A15"] = "Date"
    ticker_ws["A16"] = "Fiscal Quarter"
    ticker_ws["A146"] = "Fiscal Year"
    for i in range(24):
        col = i + 2
        ticker_ws.cell(15, col, f"2020-{(i % 12) + 1:02d}-01")
        ticker_ws.cell(16, col, f"Q{(i % 4) + 1} FY{2020 + i // 4}")
        ticker_ws.cell(146, col, 2020 + i // 4)

    series_labels = [
        "Revenue",
        "Net Income",
        "Operating Income",
        "Free Cash Flow",
        "Shares Outstanding Diluted Average (MM)",
        "Gross Margin",
        "Operating Margin",
        "ROE",
        "ROA",
        "WACC",
        "EPS Diluted",
        "Book Value Per Share",
    ]
    for offset, label in enumerate(series_labels):
        row = 18 + offset
        if row == 146:
            continue
        ticker_ws.cell(row, 1, label)
        for i in range(24):
            ticker_ws.cell(row, i + 2, 100.0 + offset + i)

    # Scalar block
    scalars = {
        158: ("Current Price (Live Price)", 190.0),
        159: ("Current Market Capitalization", 2_900_000_000_000.0),
        160: ("Current Enterprise Value (not-diluted)", 2_950_000_000_000.0),
        161: ("Current Dividend Yield", 0.005),
        162: ("WACC", 0.09),
        163: ("Current PE10", 28.5),
        164: ("Current E10", 6.5),
        165: ("Final Score", 8.2),
    }
    for row, (label, value) in scalars.items():
        ticker_ws.cell(row, 1, label)
        ticker_ws.cell(row, 2, value)

    summary = workbook.create_sheet("Summary")
    values = {
        "Company": "Apple Inc.",
        "Ticker": "AAPL",
        "Current Price (Live Price)": 190.0,
        "Current Market Capitalization": 2_900_000_000_000.0,
        "Current Enterprise Value (not-diluted)": 2_950_000_000_000.0,
        "Current Dividend Yield": 0.005,
        "WACC": 0.09,
        "Current PE10": 28.5,
        "Current E10": 6.5,
        "Final Score": 8.2,
        "Franchise Power": 0.8,
    }
    for col, header in enumerate(_SUMMARY_HEADERS, start=1):
        summary.cell(1, col, header)
        summary.cell(2, col, values.get(header, col * 0.01))

    path = fixtures_dir / "Custom_Run_Filter_AAPL.xlsx"
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def mock_company_facts() -> dict:
    return {
        "facts": {
            "us-gaap": {
                "Revenues": {
                    "label": "Revenues",
                    "units": {
                        "USD": [
                            {
                                "val": 391035000000,
                                "fy": 2024,
                                "fp": "FY",
                                "form": "10-K",
                                "filed": "2024-11-01",
                                "accn": "0000320193-24-000123",
                            }
                        ]
                    },
                },
                "NetIncomeLoss": {
                    "label": "Net Income (Loss)",
                    "units": {
                        "USD": [
                            {
                                "val": 93736000000,
                                "fy": 2024,
                                "fp": "FY",
                                "form": "10-K",
                                "filed": "2024-11-01",
                                "accn": "0000320193-24-000123",
                            }
                        ]
                    },
                },
                "Assets": {
                    "label": "Assets",
                    "units": {
                        "USD": [
                            {
                                "val": 364980000000,
                                "fy": 2024,
                                "fp": "FY",
                                "form": "10-K",
                                "filed": "2024-11-01",
                                "accn": "0000320193-24-000123",
                            }
                        ]
                    },
                },
            }
        }
    }


@pytest.fixture
def mock_filings_manifest() -> dict:
    return {
        "ticker": "AAPL",
        "cik": "0000320193",
        "selected_filings": [
            {
                "accession_number": "0000320193-24-000123",
                "filing_type": "10-K",
                "filing_date": "2024-11-01",
                "report_date": "2024-09-28",
                "primary_document": "aapl-20240928.htm",
                "document_url": (
                    "https://www.sec.gov/Archives/edgar/data/320193/"
                    "000032019324000123/aapl-20240928.htm"
                ),
                "fiscal_year": 2024,
            }
        ],
    }
