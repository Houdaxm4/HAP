"""Introspect Bloomberg Custom_Run_Filter workbooks without assuming layout."""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class CellSample:
    address: str
    value: Any


@dataclass
class RowSample:
    row_number: int
    cells: list[CellSample]


@dataclass
class WorksheetProfile:
    name: str
    max_row: int
    max_column: int
    preview_rows: list[RowSample] = field(default_factory=list)
    non_empty_row_count: int = 0


@dataclass
class WorkbookIntrospection:
    source_path: str
    worksheets: list[WorksheetProfile]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def to_json(self, *, indent: int = 2) -> str:
        return json.dumps(self.to_dict(), indent=indent, default=str)


class WorkbookIntrospector:
    """Dump the actual structure of a Custom_Run_Filter workbook for reverse-engineering."""

    def __init__(self, *, preview_rows: int = 25, preview_cols: int = 12) -> None:
        self.preview_rows = preview_rows
        self.preview_cols = preview_cols

    def inspect(self, file_path: Path) -> WorkbookIntrospection:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheets = [
                self._inspect_sheet(workbook[sheet_name]) for sheet_name in workbook.sheetnames
            ]
            return WorkbookIntrospection(
                source_path=str(file_path),
                worksheets=worksheets,
            )
        finally:
            workbook.close()

    def inspect_to_file(self, file_path: Path, output_path: Path) -> WorkbookIntrospection:
        report = self.inspect(file_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(report.to_json(), encoding="utf-8")
        return report

    def _inspect_sheet(self, sheet: Worksheet) -> WorksheetProfile:
        max_row = sheet.max_row or 0
        max_column = sheet.max_column or 0
        preview_rows: list[RowSample] = []
        non_empty_rows = 0

        for row_number, row in enumerate(
            sheet.iter_rows(
                min_row=1,
                max_row=min(max_row, self.preview_rows) if max_row else self.preview_rows,
                max_col=self.preview_cols,
                values_only=True,
            ),
            start=1,
        ):
            cells: list[CellSample] = []
            has_value = False
            for column_index, value in enumerate(row, start=1):
                if value is not None and str(value).strip() != "":
                    has_value = True
                address = f"{self._column_letter(column_index)}{row_number}"
                cells.append(CellSample(address=address, value=value))
            if has_value:
                non_empty_rows += 1
            preview_rows.append(RowSample(row_number=row_number, cells=cells))

        return WorksheetProfile(
            name=sheet.title,
            max_row=max_row,
            max_column=max_column,
            preview_rows=preview_rows,
            non_empty_row_count=non_empty_rows,
        )

    @staticmethod
    def _column_letter(column_index: int) -> str:
        letters = ""
        while column_index:
            column_index, remainder = divmod(column_index - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters
