"""Parse Bloomberg-derived Custom_Run_Filter workbooks into CustomRunData."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ingestion.models.custom_run_data import CustomRunData, MetricSeries
from ingestion.production_workbook_profile import (
    ProductionWorkbookProfile,
    SheetSectionProfile,
    load_production_profile,
    production_profile_available,
)
from ingestion.workbook_introspector import WorkbookIntrospector


class CustomRunParseError(Exception):
    """Raised when a Custom_Run_Filter workbook cannot be parsed."""


class CustomRunParser:
    """
    Parse the production Bloomberg Custom_Run_Filter workbook using an
    evidence-based profile. The profile is reverse-engineered from the real
    AAPL workbook and generalized only when other tickers prove identical.
    """

    def __init__(
        self,
        *,
        profile: ProductionWorkbookProfile | None = None,
        introspector: WorkbookIntrospector | None = None,
    ) -> None:
        self.profile = profile
        self.introspector = introspector or WorkbookIntrospector()

    def parse(self, file_path: Path, original_filename: str) -> CustomRunData:
        suffix = file_path.suffix.lower()
        if suffix not in {".xlsx", ".xlsm"}:
            raise CustomRunParseError(
                "Custom_Run_Filter must be a Bloomberg-derived Excel workbook (.xlsx). "
                "Mapping CSV files are not part of the HAP v1 product specification."
            )

        profile = self._resolve_profile(file_path)
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet_map = {name: name for name in workbook.sheetnames}
            missing_sheets = [
                section.sheet_name
                for section in profile.sections
                if section.sheet_name not in sheet_map
            ]
            if missing_sheets:
                introspection = self.introspector.inspect(file_path)
                raise CustomRunParseError(
                    "Custom_Run_Filter workbook does not match the verified production profile. "
                    f"Missing worksheets: {', '.join(missing_sheets)}. "
                    f"Found: {', '.join(workbook.sheetnames)}. "
                    "Run backend/scripts/inspect_custom_run_workbook.py on the production "
                    "workbook and update the evidence-based profile before parsing."
                    f"\n\nIntrospected worksheets: "
                    f"{', '.join(sheet.name for sheet in introspection.worksheets)}"
                )

            metadata: dict[str, str] = {}
            market_data: dict[str, float | str | None] = {}
            historical_metrics: list[MetricSeries] = []
            proprietary_metrics: list[MetricSeries] = []
            valuation_metrics: list[MetricSeries] = []
            quality_metrics: list[MetricSeries] = []
            assumptions: dict[str, float | str | None] = {}
            raw_sections: dict[str, Any] = {}

            for section in profile.sections:
                sheet = workbook[section.sheet_name]
                parsed = self._parse_section(sheet, section)
                raw_sections[section.section] = parsed

                if section.section == "metadata":
                    metadata = {k: str(v) for k, v in parsed.items() if v is not None}
                elif section.section == "market_data":
                    market_data = parsed
                elif section.section == "historical_metrics":
                    historical_metrics = parsed
                elif section.section == "proprietary_metrics":
                    proprietary_metrics = parsed
                elif section.section == "valuation_metrics":
                    valuation_metrics = parsed
                elif section.section == "quality_metrics":
                    quality_metrics = parsed
                elif section.section == "assumptions":
                    assumptions = parsed

            return CustomRunData(
                source_filename=original_filename,
                metadata=metadata,
                market_data=market_data,
                historical_metrics=historical_metrics,
                proprietary_metrics=proprietary_metrics,
                valuation_metrics=valuation_metrics,
                quality_metrics=quality_metrics,
                assumptions=assumptions,
                worksheets_found=list(workbook.sheetnames),
                raw_sections=raw_sections,
            )
        finally:
            workbook.close()

    def _resolve_profile(self, file_path: Path) -> ProductionWorkbookProfile:
        if self.profile is not None:
            return self.profile
        if not production_profile_available():
            introspection = self.introspector.inspect(file_path)
            worksheet_names = ", ".join(sheet.name for sheet in introspection.worksheets)
            raise CustomRunParseError(
                "No verified production Custom_Run_Filter profile is available. "
                "The parser cannot assume worksheet names or layout. "
                "Commit the real AAPL workbook to "
                "backend/fixtures/production/custom_run_filter_aapl.xlsx, run "
                "backend/scripts/inspect_custom_run_workbook.py, author "
                "backend/fixtures/production/custom_run_filter_aapl.profile.json from "
                "that evidence, then rerun the parser."
                f"\n\nUploaded workbook worksheets: {worksheet_names}"
            )
        return load_production_profile()

    def _parse_section(
        self,
        sheet: Worksheet,
        section: SheetSectionProfile,
    ) -> Any:
        if section.layout == "key_value":
            numeric = section.section in {"market_data", "assumptions"}
            return self._parse_key_value_sheet(
                sheet,
                label_column=section.label_column,
                value_column=section.value_column,
                start_row=section.data_start_row,
                numeric=numeric,
            )
        if section.layout == "time_series":
            return self._parse_time_series_sheet(
                sheet,
                header_row=section.header_row,
                label_column=section.label_column,
                period_start_column=section.period_start_column,
                data_start_row=section.data_start_row,
            )
        if section.layout == "metric_value":
            return self._parse_metric_value_sheet(
                sheet,
                label_column=section.label_column,
                value_column=section.value_column,
                start_row=section.data_start_row,
            )
        return self._parse_raw_grid(sheet)

    def _parse_key_value_sheet(
        self,
        sheet: Worksheet,
        *,
        label_column: int,
        value_column: int,
        start_row: int,
        numeric: bool = False,
    ) -> dict[str, float | str | None]:
        result: dict[str, float | str | None] = {}
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            label_index = label_column - 1
            value_index = value_column - 1
            if not row or label_index >= len(row) or row[label_index] is None:
                continue
            field = str(row[label_index]).strip()
            if not field:
                continue
            raw_value = row[value_index] if value_index < len(row) else None
            result[field] = self._coerce_value(raw_value, numeric=numeric)
        return result

    def _parse_time_series_sheet(
        self,
        sheet: Worksheet,
        *,
        header_row: int,
        label_column: int,
        period_start_column: int,
        data_start_row: int,
    ) -> list[MetricSeries]:
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < header_row:
            return []

        header = rows[header_row - 1]
        period_headers = [
            self._normalize_period(value)
            for value in header[period_start_column - 1 :]
        ]
        periods = [period for period in period_headers if period]
        if not periods:
            raise CustomRunParseError(
                f"Worksheet '{sheet.title}' has no period headers on row {header_row}."
            )

        series_list: list[MetricSeries] = []
        label_index = label_column - 1
        for row in rows[data_start_row - 1 :]:
            if not row or label_index >= len(row) or row[label_index] is None:
                continue
            metric = str(row[label_index]).strip()
            if not metric:
                continue
            values: dict[str, float | str | None] = {}
            for index, period in enumerate(periods):
                cell_index = period_start_column - 1 + index
                raw = row[cell_index] if cell_index < len(row) else None
                values[period] = self._coerce_value(raw, numeric=True)
            series_list.append(MetricSeries(metric=metric, values=values))
        return series_list

    def _parse_metric_value_sheet(
        self,
        sheet: Worksheet,
        *,
        label_column: int,
        value_column: int,
        start_row: int,
    ) -> list[MetricSeries]:
        series_list: list[MetricSeries] = []
        label_index = label_column - 1
        value_index = value_column - 1
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            if not row or label_index >= len(row) or row[label_index] is None:
                continue
            metric = str(row[label_index]).strip()
            if not metric:
                continue
            raw = row[value_index] if value_index < len(row) else None
            value = self._coerce_value(raw, numeric=True)
            series_list.append(MetricSeries(metric=metric, values={"current": value}))
        return series_list

    @staticmethod
    def _parse_raw_grid(sheet: Worksheet) -> list[list[Any]]:
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    @staticmethod
    def _normalize_period(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        fy_match = re.match(r"^FY\s*'?(\d{2}|\d{4})$", text, re.IGNORECASE)
        if fy_match:
            year = fy_match.group(1)
            if len(year) == 2:
                year = f"20{year}"
            return f"FY{year}"
        if re.fullmatch(r"20\d{2}|19\d{2}", text):
            return f"FY{text}"
        return text

    @staticmethod
    def _coerce_value(raw: Any, *, numeric: bool = False) -> float | str | None:
        if raw is None:
            return None
        if isinstance(raw, str) and not raw.strip():
            return None
        if numeric:
            try:
                return float(raw)
            except (TypeError, ValueError):
                if isinstance(raw, str):
                    cleaned = raw.replace(",", "").replace("$", "").strip()
                    if cleaned.endswith("%"):
                        cleaned = cleaned[:-1]
                    try:
                        return float(cleaned)
                    except ValueError:
                        return raw.strip()
                return str(raw).strip()
        return str(raw).strip() if not isinstance(raw, (int, float)) else raw
